import sys
import os
import json
import shutil
import shlex
import re
import subprocess
import fnmatch
import webbrowser
import zipfile
import stat
import hashlib
import logging
import tempfile
from concurrent.futures import ThreadPoolExecutor
import threading
import queue
from concurrent.futures import ThreadPoolExecutor
from collections import OrderedDict, defaultdict
from dataclasses import dataclass
from typing import Dict, List, Optional, Sequence, Set, Tuple
from difflib import SequenceMatcher
from functools import partial
from urllib.parse import quote_plus
import urllib.request
import importlib
import importlib.util
try:
    from bs4 import BeautifulSoup  # type: ignore
except Exception:
    BeautifulSoup = None
from datetime import datetime, time, date
from openpyxl import Workbook
from pathlib import Path

from modules.log_manager import LogManagerDialog, LogAnalyzerHooks
from modules.ts4script_search import Ts4ScriptSearchDialog
from modules.id_conflict_viewer import IDConflictViewerDialog as IDConflictViewerDialogV2

# Qt bindings import shim (supports PyQt5/PySide2/PySide6/PyQt6)
try:
    from PyQt5 import QtWidgets, QtCore, QtGui  # type: ignore
    QT_LIB = "PyQt5"
except Exception:
    try:
        from PySide2 import QtWidgets, QtCore, QtGui  # type: ignore
        QT_LIB = "PySide2"
    except Exception:
        try:
            from PySide6 import QtWidgets, QtCore, QtGui  # type: ignore
            QT_LIB = "PySide6"
        except Exception:
            from PyQt6 import QtWidgets, QtCore, QtGui  # type: ignore
            QT_LIB = "PyQt6"

# Compatibility shims for Qt6 enum/API differences used in this codebase
if 'QT_LIB' in globals() and QT_LIB in ("PyQt6", "PySide6"):
    # QDialogButtonBox: map StandardButton members to direct attributes (Ok, Cancel, ...)
    if hasattr(QtWidgets.QDialogButtonBox, "StandardButton"):
        _STD = QtWidgets.QDialogButtonBox.StandardButton
        for _name in ("Ok", "Cancel", "Yes", "No", "Apply", "Save", "Open", "Close"):
            if not hasattr(QtWidgets.QDialogButtonBox, _name) and hasattr(_STD, _name):
                setattr(QtWidgets.QDialogButtonBox, _name, getattr(_STD, _name))

    # QtCore.Qt enum aliases commonly used by PyQt5-era code
    if hasattr(QtCore, "Qt"):
        _Qt = QtCore.Qt
        # Alignment flags
        if hasattr(_Qt, "AlignmentFlag"):
            if not hasattr(_Qt, "AlignRight"):
                _Qt.AlignRight = _Qt.AlignmentFlag.AlignRight
            if not hasattr(_Qt, "AlignCenter"):
                _Qt.AlignCenter = _Qt.AlignmentFlag.AlignCenter
        # Text interaction
        if hasattr(_Qt, "TextInteractionFlag") and not hasattr(_Qt, "TextSelectableByMouse"):
            _Qt.TextSelectableByMouse = _Qt.TextInteractionFlag.TextSelectableByMouse
        # Context menu policy
        if hasattr(_Qt, "ContextMenuPolicy") and not hasattr(_Qt, "CustomContextMenu"):
            _Qt.CustomContextMenu = _Qt.ContextMenuPolicy.CustomContextMenu
        # Orientation
        if hasattr(_Qt, "Orientation") and not hasattr(_Qt, "Horizontal"):
            _Qt.Horizontal = _Qt.Orientation.Horizontal
        # Check state
        if hasattr(_Qt, "CheckState") and not hasattr(_Qt, "Checked"):
            _Qt.Checked = _Qt.CheckState.Checked
        # Item flags
        if hasattr(_Qt, "ItemFlag"):
            if not hasattr(_Qt, "ItemIsEditable"):
                _Qt.ItemIsEditable = _Qt.ItemFlag.ItemIsEditable
            if not hasattr(_Qt, "ItemIsEnabled"):
                _Qt.ItemIsEnabled = _Qt.ItemFlag.ItemIsEnabled
            if not hasattr(_Qt, "ItemIsSelectable"):
                _Qt.ItemIsSelectable = _Qt.ItemFlag.ItemIsSelectable
        # Item data roles
        if hasattr(_Qt, "ItemDataRole") and not hasattr(_Qt, "UserRole"):
            _Qt.UserRole = _Qt.ItemDataRole.UserRole
    # Widgets enum aliasing
    try:
        if hasattr(QtWidgets, "QAbstractItemView") and hasattr(QtWidgets.QAbstractItemView, "SelectionMode"):
            if not hasattr(QtWidgets.QAbstractItemView, "NoSelection"):
                QtWidgets.QAbstractItemView.NoSelection = QtWidgets.QAbstractItemView.SelectionMode.NoSelection
    except Exception:
        pass
    try:
        if hasattr(QtWidgets, "QComboBox") and hasattr(QtWidgets.QComboBox, "SizeAdjustPolicy"):
            if not hasattr(QtWidgets.QComboBox, "AdjustToContents"):
                QtWidgets.QComboBox.AdjustToContents = QtWidgets.QComboBox.SizeAdjustPolicy.AdjustToContents
    except Exception:
        pass
    try:
        if hasattr(QtWidgets, "QHeaderView") and hasattr(QtWidgets.QHeaderView, "ResizeMode"):
            if not hasattr(QtWidgets.QHeaderView, "Stretch"):
                QtWidgets.QHeaderView.Stretch = QtWidgets.QHeaderView.ResizeMode.Stretch
            if not hasattr(QtWidgets.QHeaderView, "ResizeToContents"):
                QtWidgets.QHeaderView.ResizeToContents = QtWidgets.QHeaderView.ResizeMode.ResizeToContents
    except Exception:
        pass
    # exec_ compatibility for Qt6
    try:
        for _cls in (
            getattr(QtWidgets, "QDialog", None),
            getattr(QtWidgets, "QApplication", None),
            getattr(QtWidgets, "QMenu", None),
        ):
            if _cls is not None and hasattr(_cls, "exec") and not hasattr(_cls, "exec_"):
                setattr(_cls, "exec_", getattr(_cls, "exec"))
    except Exception:
        pass
    # QDialog.DialogCode aliases
    try:
        if hasattr(QtWidgets, "QDialog") and hasattr(QtWidgets.QDialog, "DialogCode"):
            _DC = QtWidgets.QDialog.DialogCode
            if not hasattr(QtWidgets.QDialog, "Accepted"):
                QtWidgets.QDialog.Accepted = _DC.Accepted
            if not hasattr(QtWidgets.QDialog, "Rejected"):
                QtWidgets.QDialog.Rejected = _DC.Rejected
    except Exception:
        pass

# Optional: new Mod Root archive installer
try:
    from mod_root_zip import (
        install_zip as mr_install_zip,
        plan_zip as mr_plan_zip,
        install_extracted_dir as mr_install_extracted_dir,
        plan_extracted_dir as mr_plan_extracted_dir,
    )
except Exception:
    mr_install_zip = None
    mr_plan_zip = None
    mr_install_extracted_dir = None
    mr_plan_extracted_dir = None

# Cross-binding Slot decorator shim (PyQt5 uses pyqtSlot, PySide uses Slot)
try:
    SLOT = getattr(QtCore, 'pyqtSlot')
except Exception:
    SLOT = getattr(QtCore, 'Slot', None)
if SLOT is None:
    def SLOT(*_args, **_kwargs):
        def _decorator(func):
            return func
        return _decorator

# Cross-binding Signal shim
try:
    SIGNAL = getattr(QtCore, 'pyqtSignal')
except Exception:
    SIGNAL = getattr(QtCore, 'Signal', None)

# Optional watchdog (used to monitor mod_scan_cache.json)
try:
    from watchdog.observers import Observer  # type: ignore
    from watchdog.events import FileSystemEventHandler  # type: ignore
except Exception:
    Observer = None  # type: ignore
    FileSystemEventHandler = object  # type: ignore

class ScanWorker(QtCore.QObject):
    pass

SETTINGS_PATH = "settings.json"
IGNORE_LIST_PATH = "ignorelist.txt"
VERSION_RELEASE_PATH = "version_release.json"
APP_VERSION = "v3.48.0"
APP_VERSION_DATE = "02/11/2025 18:00 UTC"
INSTALLED_MODS_PATH = "installed_mods.json"
MOD_SCAN_CACHE_PATH = "mod_scan_cache.json"
MOD_MARKER_FILENAME = ".s4mt_mod_marker.json"
LOG_FILE_PATH = "sims4modtool.log"
ID_INDEX_CACHE_PATH = "id_index_cache.json"

# --- Logging setup (default DEBUG, configurable via settings) ---
_LEVEL_MAP = {
    "CRITICAL": logging.CRITICAL,
    "ERROR": logging.ERROR,
    "WARNING": logging.WARNING,
    "INFO": logging.INFO,
    "DEBUG": logging.DEBUG,
}

def _normalize_log_level(value):
    if isinstance(value, int):
        return value
    try:
        key = str(value or "").strip().upper()
    except Exception:
        key = "DEBUG"
    return _LEVEL_MAP.get(key, logging.DEBUG)

def setup_logging(level="DEBUG"):
    level_no = _normalize_log_level(level)
    logger = logging.getLogger("Sims4ModTool")
    logger.setLevel(level_no)
    logger.propagate = False

    # Ensure handlers only added once
    if not logger.handlers:
        formatter = logging.Formatter(
            fmt="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        # Console handler
        ch = logging.StreamHandler()
        ch.setLevel(level_no)
        ch.setFormatter(formatter)
        logger.addHandler(ch)
        try:
            # File handler (best-effort)
            fh = logging.FileHandler(LOG_FILE_PATH, encoding="utf-8")
            fh.setLevel(level_no)
            fh.setFormatter(formatter)
            logger.addHandler(fh)
        except Exception:
            # If file logging fails, continue with console only
            pass

    # Also set root logger to the same level for libraries if needed
    logging.getLogger().setLevel(level_no)
    return logger

# Initialize logging early with default DEBUG; runtime settings can override later
setup_logging("DEBUG")

SUPPORTED_INSTALL_EXTENSIONS = {".package", ".ts4script", ".zip", ".7z", ".rar"}

IGNORED_ARCHIVE_PREFIXES = {"__MACOSX"}
IGNORED_ARCHIVE_PREFIX_MATCHES = (".git",)
IGNORED_ARCHIVE_FILENAMES = {"thumbs.db", ".ds_store"}
IGNORED_ARCHIVE_NAME_PREFIXES = ("readme", "license")
DISALLOWED_ARCHIVE_EXTENSIONS = {
    ".exe",
    ".dll",
    ".bat",
    ".cmd",
    ".com",
    ".msi",
    ".ps1",
    ".vbs",
    ".js",
    ".jar",
    ".scr",
    ".pif",
    ".apk",
    ".app",
    ".sh",
    ".bash",
    ".py",
    ".rb",
    ".php",
}
MAX_RELATIVE_DEPTH = 2

MOD_NAME_SANITIZE_RE = re.compile(r"[^a-z0-9]+")
MIN_SIMILARITY_RATIO = 0.7
MIN_NAME_LENGTH = 4
MAX_SIMILARITY_CANDIDATES = 80


@dataclass(frozen=True)
class ZipInstallEntry:
    member_name: str
    relative_parts: Tuple[str, ...]


@dataclass
class ZipInstallPlan:
    mod_folder_name: str
    target_folder: str
    entries: List[ZipInstallEntry]
    warnings: List[str]


@dataclass
class ZipPlanResult:
    success: bool
    plan: Optional[ZipInstallPlan]
    message: str = ""


def normalize_addon_metadata(addons):
    normalized = []
    if not isinstance(addons, list):
        return normalized

    def _append_or_merge(target_list, candidate):
        label = candidate.get("label", "")
        if not label:
            return
        key = label.casefold()
        for existing in target_list:
            if existing.get("label", "").casefold() == key:
                existing_paths = existing.setdefault("paths", [])
                for path in candidate.get("paths", []):
                    if path and path not in existing_paths:
                        existing_paths.append(path)
                if candidate.get("added_at") and not existing.get("added_at"):
                    existing["added_at"] = candidate.get("added_at")
                return
        target_list.append(candidate)

    for addon in addons:
        label = ""
        paths = []
        added_at = ""
        if isinstance(addon, dict):
            label = str(
                addon.get("label")
                or addon.get("name")
                or addon.get("source")
                or addon.get("title")
                or ""
            ).strip()
            raw_paths = addon.get("paths", [])
            if isinstance(raw_paths, (list, tuple)):
                for path in raw_paths:
                    path_str = str(path).replace("\\", "/").strip()
                    if path_str:
                        if path_str.startswith("..") or "/../" in path_str:
                            continue
                        if path_str.endswith("/"):
                            base_path = os.path.normpath(path_str[:-1])
                            normalized_path = base_path.replace("\\", "/") + "/"
                        else:
                            base_path = os.path.normpath(path_str)
                            normalized_path = base_path.replace("\\", "/")
                        if normalized_path not in paths:
                            paths.append(normalized_path)
            added_at = str(addon.get("added_at") or "").strip()
        else:
            label = str(addon).strip()

        if not label:
            continue

        _append_or_merge(
            normalized,
            {
                "label": label,
                "paths": paths,
                "added_at": added_at,
            },
        )

    return normalized


DEFAULT_VERSION_RELEASES = [
    {"version": "1.109.185.1030", "release_date": "2024-09-18"},
    {"version": "1.110.265.1030", "release_date": "2024-10-22"},
    {"version": "1.110.311.1020", "release_date": "2024-11-07"},
    {"version": "1.111.102.1030", "release_date": "2024-12-03"},
    {"version": "1.115.216.1030", "release_date": "2025-05-27"},
    {"version": "1.115.253.1020", "release_date": "2025-06-17"},
    {"version": "1.116.202.1030", "release_date": "2025-07-01"},
    {"version": "1.116.223.1030", "release_date": "2025-07-10"},
    {"version": "1.116.232.1030", "release_date": "2025-07-16"},
    {"version": "1.116.240.1020", "release_date": "2025-07-29"},
    {"version": "1.117.227.1030", "release_date": "2025-08-19"},
    {"version": "1.117.244.1020", "release_date": "2025-09-04"},
    {"version": "1.118.242.1030", "release_date": "2025-09-18"},
    {"version": "1.118.257.1020", "release_date": "2025-10-02"},
]


def parse_release_date(date_str):
    if not date_str:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def format_release_date(date_obj):
    if not date_obj:
        return ""
    return date_obj.strftime("%d/%m/%Y")


def load_custom_version_releases(path=VERSION_RELEASE_PATH):
    if not os.path.exists(path):
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError):
        return {}

    releases = {}
    if isinstance(data, dict):
        items = data.items()
    elif isinstance(data, list):
        items = ((entry.get("version"), entry.get("release_date")) for entry in data)
    else:
        return {}

    for version, date_str in items:
        if not version:
            continue
        parsed = parse_release_date(date_str)
        if parsed is None:
            continue
        releases[str(version)] = parsed
    return releases


def save_custom_version_releases(releases, path=VERSION_RELEASE_PATH):
    entries = [
        {"version": version, "release_date": date.isoformat()}
        for version, date in sorted(releases.items(), key=lambda item: (item[1], item[0]))
    ]
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(entries, f, indent=4, ensure_ascii=False)


def merge_version_releases(custom_releases):
    merged = {}
    for entry in DEFAULT_VERSION_RELEASES:
        version = entry.get("version")
        date_str = entry.get("release_date")
        if not version or not date_str:
            continue
        parsed = parse_release_date(date_str)
        if parsed is None:
            continue
        merged[version] = parsed
    for version, date_obj in custom_releases.items():
        if not version or date_obj is None:
            continue
        merged[version] = date_obj
    sorted_items = sorted(merged.items(), key=lambda item: (item[1], item[0]))
    return OrderedDict(sorted_items)


def estimate_version_from_dates(package_date, script_date, version_releases):
    candidates = [value for value in (package_date, script_date) if value is not None]
    if not candidates:
        return ""
    latest_datetime = max(candidates)
    latest_date = latest_datetime.date()
    estimated_version = ""
    for version, release_date in version_releases.items():
        if release_date <= latest_date:
            estimated_version = version
        elif not estimated_version:
            return version
        else:
            break
    return estimated_version


def format_datetime(value):
    if not value:
        return ""
    return value.strftime("%d/%m/%Y %H:%M")


def load_ignore_list(path=IGNORE_LIST_PATH):
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


def save_ignore_list(ignored_mods, path=IGNORE_LIST_PATH):
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        for mod_name in sorted(set(ignored_mods)):
            f.write(f"{mod_name}\n")


def load_installed_mods(path=INSTALLED_MODS_PATH):
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
    except (json.JSONDecodeError, OSError):
        return []

    normalized_entries = []
    for entry in data if isinstance(data, list) else []:
        if not isinstance(entry, dict):
            continue
        name = str(entry.get("name") or entry.get("mod_name") or "").strip()
        target_folder = str(entry.get("target_folder") or "").strip()
        if not name or not target_folder:
            continue
        normalized_entries.append({
            "name": name,
            "type": str(entry.get("type") or "").strip(),
            "installed_at": str(entry.get("installed_at") or "").strip(),
            "target_folder": target_folder,
            "source": str(entry.get("source") or "").strip(),
            "addons": normalize_addon_metadata(entry.get("addons", [])),
            "files": [str(p).replace("\\", "/").strip() for p in (entry.get("files", []) or []) if str(p).strip()],
            "mod_version": str(entry.get("mod_version") or "").strip(),
            "url": str(entry.get("url") or "").strip(),
            "atf": bool(entry.get("atf", False)),
            "disabled": bool(entry.get("disabled", False)),
            "disabled_path": str(entry.get("disabled_path") or "").strip(),
        })

    normalized_entries.sort(key=lambda item: item.get("installed_at", ""), reverse=True)
    return normalized_entries


def save_installed_mods(installed_mods, path=INSTALLED_MODS_PATH):
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(installed_mods, handle, indent=4, ensure_ascii=False)


def load_mod_scan_cache(path=MOD_SCAN_CACHE_PATH):
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
    except (json.JSONDecodeError, OSError):
        return None
    if not isinstance(data, dict):
        return None
    entries = data.get("entries")
    if not isinstance(entries, list):
        return None
    normalized_entries = []
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        path_value = str(entry.get("path") or "").strip()
        if not path_value:
            continue
        normalized_entries.append({
            "path": path_value.replace("\\", "/"),
            "mtime": int(entry.get("mtime", 0)),
            "size": int(entry.get("size", 0)),
            "type": str(entry.get("type") or ""),
        })
    normalized_entries.sort(key=lambda item: item["path"].casefold())
    normalized_root = str(data.get("root") or "").replace("\\", "/").strip()
    cache = {
        "root": normalized_root,
        "entries": normalized_entries,
    }
    try:
        logging.getLogger("Sims4ModTool").debug(
            "Loaded mod_scan_cache.json: root=%s, entries=%d",
            cache.get("root"), len(normalized_entries)
        )
    except Exception:
        pass
    return cache


def save_mod_scan_cache(snapshot, path=MOD_SCAN_CACHE_PATH):
    if not isinstance(snapshot, dict):
        return
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    serializable = {
        "root": str(snapshot.get("root") or ""),
        "generated_at": snapshot.get("generated_at", ""),
        "entries": list(snapshot.get("entries", [])),
    }
    start_time = datetime.utcnow()
    try:
        logging.getLogger("Sims4ModTool").info(
            "Writing cache -> %s (entries=%d)",
            os.path.abspath(path), len(serializable.get("entries", []))
        )
    except Exception:
        pass
    with open(path, "w", encoding="utf-8") as handle:
        json.dump(serializable, handle, indent=2, ensure_ascii=False)
    try:
        size = os.path.getsize(path)
    except Exception:
        size = -1
    try:
        elapsed = (datetime.utcnow() - start_time).total_seconds() * 1000.0
        logging.getLogger("Sims4ModTool").info(
            "Cache written: size=%s bytes, elapsed=%.1fms, generated_at=%s",
            size, elapsed, serializable.get("generated_at", "")
        )
    except Exception:
        pass


def mod_scan_snapshots_equal(first, second):
    if not first or not second:
        return False
    first_entries = first.get("entries") or []
    second_entries = second.get("entries") or []
    if len(first_entries) != len(second_entries):
        return False
    for a, b in zip(first_entries, second_entries):
        if (
            a.get("path") != b.get("path")
            or int(a.get("mtime", 0)) != int(b.get("mtime", 0))
            or int(a.get("size", 0)) != int(b.get("size", 0))
            or (a.get("type") or "") != (b.get("type") or "")
        ):
            return False
    return True


def sanitize_mod_folder_name(file_name):
    base_name = os.path.splitext(os.path.basename(file_name))[0]
    sanitized = re.sub(r"[\\/:*?\"<>|]", "_", base_name).strip()
    return sanitized or "mod"


def sanitize_archive_member_path(member_name):
    normalized = str(member_name).replace("\\", "/").strip()
    normalized = re.sub(r"/+", "/", normalized.lstrip("/"))
    if not normalized:
        return ""

    parts = []
    for raw_part in normalized.split("/"):
        part = raw_part.strip()
        if not part or part in {".", ".."}:
            continue
        safe_part = re.sub(r"[\\/:*?\"<>|]", "_", part)
        if safe_part:
            parts.append(safe_part)

    return "/".join(parts)


def _zipinfo_is_symlink(info):
    mode = (info.external_attr >> 16) & 0xFFFF
    return stat.S_ISLNK(mode)


def _member_should_be_skipped(parts, name):
    if not parts:
        return True
    first = parts[0].casefold()
    if first in {prefix.casefold() for prefix in IGNORED_ARCHIVE_PREFIXES}:
        return True
    for prefix in IGNORED_ARCHIVE_PREFIX_MATCHES:
        if first.startswith(prefix.casefold()):
            return True
    lowered = name.casefold()
    if lowered in IGNORED_ARCHIVE_FILENAMES:
        return True
    for prefix in IGNORED_ARCHIVE_NAME_PREFIXES:
        if lowered.startswith(prefix):
            return True
    return False


def _collapse_folder_name(parts):
    filtered = [part for part in parts if part]
    if not filtered:
        return ""
    if len(filtered) == 1:
        return filtered[0]
    return "-".join(filtered)


def _ensure_depth(parts):
    # Preserve last segments instead of collapsing names with hyphens
    # Example: ["betterbuy", "data", "file"] with MAX_RELATIVE_DEPTH=2 -> ["data", "file"]
    if len(parts) <= MAX_RELATIVE_DEPTH:
        return list(parts)
    return list(parts[-MAX_RELATIVE_DEPTH:])


def _register_directory(existing_dirs, parts, warnings):
    if not parts:
        return
    normalized = tuple(part.casefold() for part in parts)
    if normalized in existing_dirs:
        return
    existing_dirs.add(normalized)


def _ensure_unique_parts(existing_paths, existing_dirs, parts, warnings, *, is_dir=False):
    normalized = tuple(part.casefold() for part in parts)
    registry = existing_dirs if is_dir else existing_paths
    if normalized not in registry:
        registry.add(normalized)
        if not is_dir and parts[:-1]:
            _register_directory(existing_dirs, parts[:-1], warnings)
        return list(parts)

    base_name = parts[-1]
    if is_dir:
        stem = base_name
        suffix = ""
    else:
        stem, suffix = os.path.splitext(base_name)
    counter = 2
    while True:
        candidate_name = f"{stem}-v{counter}{suffix}"
        candidate_parts = list(parts[:-1]) + [candidate_name]
        normalized_candidate = tuple(part.casefold() for part in candidate_parts)
        if normalized_candidate not in registry:
            registry.add(normalized_candidate)
            if not is_dir:
                _register_directory(existing_dirs, candidate_parts[:-1], warnings)
            warnings.append(
                f"Conflit détecté pour '{base_name}', renommé en '{candidate_name}'."
            )
            return candidate_parts
        counter += 1


def _resolve_file_conflicts(target_root, relative_parts, warnings, written_paths):
    base_name = relative_parts[-1]
    stem, ext = os.path.splitext(base_name)
    if not stem:
        stem = "fichier"
    counter = 1
    renamed_target = None
    candidate_parts = list(relative_parts)

    while True:
        candidate_path = os.path.join(target_root, *candidate_parts)
        candidate_key = tuple(part.casefold() for part in candidate_parts)
        if candidate_key not in written_paths and not os.path.exists(candidate_path):
            written_paths.add(candidate_key)
            if renamed_target is not None:
                warnings.append(
                    f"Fichier existant détecté pour '{'/'.join(relative_parts)}', renommé en '{'/'.join(candidate_parts)}'."
                )
            return candidate_parts

        counter += 1
        renamed_target = f"{stem}-v{counter}{ext}"
        candidate_parts = list(relative_parts[:-1]) + [renamed_target]


def _preferred_parent_parts(entry):
    # Keep only the immediate parent directory to preserve original folder names
    if len(entry["adjusted_parts"]) <= 1:
        return []
    parents = entry["adjusted_parts"][:-1]
    return [parents[-1]]


def _organize_zip_entries(entries):
    plan_entries: List[ZipInstallEntry] = []
    warnings: List[str] = []
    existing_paths: Set[Tuple[str, ...]] = set()
    existing_dirs: Set[Tuple[str, ...]] = set()
    script_parents: Dict[str, Tuple[str, ...]] = {}
    scripts_by_base: Dict[str, List[Dict[str, object]]] = defaultdict(list)

    for entry in entries:
        if entry["is_ts4script"]:
            scripts_by_base[entry["base_name"]].append(entry)

    for entry in entries:
        if not entry["is_ts4script"]:
            continue
        normalized_name = (entry["name"].casefold(),)
        if normalized_name in existing_paths:
            module_base = sanitize_mod_folder_name(entry.get("raw_base") or entry["name"])
            module_base = module_base or "module"
            folder_candidate = module_base
            counter = 2
            while (folder_candidate.casefold(),) in existing_dirs:
                folder_candidate = f"{module_base}-v{counter}"
                counter += 1
            final_parts = _ensure_unique_parts(
                existing_paths,
                existing_dirs,
                [folder_candidate, entry["name"]],
                warnings,
            )
            script_parent = (final_parts[0],)
        else:
            final_parts = _ensure_unique_parts(
                existing_paths,
                existing_dirs,
                [entry["name"]],
                warnings,
            )
            script_parent = tuple(final_parts[:-1])
        script_parents[entry["member_name"]] = script_parent
        plan_entries.append(ZipInstallEntry(entry["member_name"], tuple(final_parts)))

    for entry in entries:
        if not entry["is_package"]:
            continue
        base_key = entry["base_name"]
        parent: Sequence[str] = []
        candidate_scripts = scripts_by_base.get(base_key, [])
        if candidate_scripts:
            script_entry = candidate_scripts[0]
            parent = script_parents.get(script_entry["member_name"], ())
        else:
            parent = _preferred_parent_parts(entry)
        final_parts = list(parent) + [entry["name"]]
        final_parts = _ensure_depth(final_parts)
        final_parts = _ensure_unique_parts(existing_paths, existing_dirs, final_parts, warnings)
        plan_entries.append(ZipInstallEntry(entry["member_name"], tuple(final_parts)))

    for entry in entries:
        if entry["is_ts4script"] or entry["is_package"]:
            continue
        parent = _preferred_parent_parts(entry)
        final_parts = parent + [entry["name"]]
        final_parts = _ensure_depth(final_parts)
        final_parts = _ensure_unique_parts(existing_paths, existing_dirs, final_parts, warnings)
        plan_entries.append(ZipInstallEntry(entry["member_name"], tuple(final_parts)))

    return plan_entries, warnings


def build_zip_install_plan(
    file_path,
    *,
    mod_directory,
    default_mod_name,
    existing_target=None,
):
    try:
        archive = zipfile.ZipFile(file_path, "r")
    except zipfile.BadZipFile as exc:
        return ZipPlanResult(False, None, f"Archive zip invalide : {exc}")

    with archive:
        entries: List[Dict[str, object]] = []
        for info in archive.infolist():
            sanitized = sanitize_archive_member_path(info.filename)
            if not sanitized:
                continue
            parts = sanitized.split("/")
            if not parts:
                continue
            name = parts[-1]
            if info.is_dir():
                if _member_should_be_skipped(parts, name):
                    continue
                continue
            if _zipinfo_is_symlink(info):
                return ZipPlanResult(False, None, f"Lien symbolique détecté dans l'archive : {info.filename}")
            extension = os.path.splitext(name)[1].lower()
            if extension in DISALLOWED_ARCHIVE_EXTENSIONS:
                return ZipPlanResult(False, None, f"Fichier interdit détecté : {name}")
            if _member_should_be_skipped(parts, name):
                continue
            raw_base = os.path.splitext(name)[0]
            entry = {
                "member_name": info.filename,
                "parts": parts,
                "name": name,
                "extension": extension,
                "is_ts4script": extension == ".ts4script",
                "is_package": extension == ".package",
                "base_name": raw_base.casefold(),
                "raw_base": raw_base,
            }
            entries.append(entry)

        if not entries:
            return ZipPlanResult(False, None, "Aucun fichier exploitable dans l'archive.")

        root_dirs: Set[str] = set()
        root_files_present = False
        for entry in entries:
            parts = entry["parts"]
            if len(parts) == 1:
                root_files_present = True
            else:
                root_dirs.add(parts[0])

        drop_segments = 0
        unique_root_name = ""
        if len(root_dirs) == 1 and not root_files_present:
            unique_root_name = next(iter(root_dirs))
            drop_segments = 1

        if existing_target:
            target_folder = existing_target
            mod_folder_name = os.path.basename(existing_target.rstrip("/\\")) or default_mod_name
        else:
            mod_folder_name = default_mod_name
            if unique_root_name:
                candidate = sanitize_mod_folder_name(unique_root_name)
                if candidate:
                    mod_folder_name = candidate
            target_folder = os.path.join(mod_directory, mod_folder_name)

        adjusted_entries: List[Dict[str, object]] = []
        for entry in entries:
            parts = entry["parts"][drop_segments:]
            if not parts:
                continue
            adjusted_entry = dict(entry)
            adjusted_entry["adjusted_parts"] = parts
            adjusted_entries.append(adjusted_entry)

        plan_entries, warnings = _organize_zip_entries(adjusted_entries)
        if not plan_entries:
            return ZipPlanResult(False, None, "Aucun fichier valide après normalisation de l'archive.")

        plan = ZipInstallPlan(
            mod_folder_name=mod_folder_name,
            target_folder=target_folder,
            entries=plan_entries,
            warnings=warnings,
        )
        return ZipPlanResult(True, plan, "")


def build_extracted_install_plan(
    extracted_root,
    *,
    mod_directory,
    default_mod_name,
    existing_target=None,
):
    entries: List[Dict[str, object]] = []
    root = os.path.abspath(extracted_root)
    # Walk extracted tree and build entries similar to build_zip_install_plan
    for cur, _dirs, files in os.walk(root):
        for file in files:
            rel_path = os.path.relpath(os.path.join(cur, file), root).replace("\\", "/")
            sanitized = sanitize_archive_member_path(rel_path)
            if not sanitized:
                continue
            parts = sanitized.split("/")
            if not parts:
                continue
            name = parts[-1]
            extension = os.path.splitext(name)[1].lower()
            if extension in DISALLOWED_ARCHIVE_EXTENSIONS:
                continue
            if _member_should_be_skipped(parts, name):
                continue
            raw_base = os.path.splitext(name)[0]
            entry = {
                "member_name": "/".join(parts),  # relative path from root
                "parts": parts,
                "name": name,
                "extension": extension,
                "is_ts4script": extension == ".ts4script",
                "is_package": extension == ".package",
                "base_name": raw_base.casefold(),
                "raw_base": raw_base,
            }
            entries.append(entry)

    if not entries:
        return ZipPlanResult(False, None, "Aucun fichier exploitable après extraction.")

    root_dirs: Set[str] = set()
    root_files_present = False
    for entry in entries:
        parts = entry["parts"]
        if len(parts) == 1:
            root_files_present = True
        else:
            root_dirs.add(parts[0])

    drop_segments = 0
    unique_root_name = ""
    if len(root_dirs) == 1 and not root_files_present:
        unique_root_name = next(iter(root_dirs))
        drop_segments = 1

    if existing_target:
        target_folder = existing_target
        mod_folder_name = os.path.basename(existing_target.rstrip("/\\")) or default_mod_name
    else:
        mod_folder_name = default_mod_name
        if unique_root_name:
            candidate = sanitize_mod_folder_name(unique_root_name)
            if candidate:
                mod_folder_name = candidate
        target_folder = os.path.join(mod_directory, mod_folder_name)

    adjusted_entries: List[Dict[str, object]] = []
    for entry in entries:
        parts = entry["parts"][drop_segments:]
        if not parts:
            continue
        adjusted_entry = dict(entry)
        adjusted_entry["adjusted_parts"] = parts
        adjusted_entries.append(adjusted_entry)

    plan_entries, warnings = _organize_zip_entries(adjusted_entries)
    if not plan_entries:
        return ZipPlanResult(False, None, "Aucun fichier valide après normalisation de l'extraction.")

    plan = ZipInstallPlan(
        mod_folder_name=mod_folder_name,
        target_folder=target_folder,
        entries=plan_entries,
        warnings=warnings,
    )
    # Reuse ZipPlanResult for simplicity
    return ZipPlanResult(True, plan, "")
def format_installation_display(iso_value):
    if not iso_value:
        return ""
    try:
        parsed = datetime.fromisoformat(iso_value)
    except ValueError:
        return iso_value
    return parsed.strftime("%d/%m/%Y %H:%M UTC")

def get_file_date(file_path):
    timestamp = os.path.getmtime(file_path)
    return datetime.fromtimestamp(timestamp)

def load_settings(path=SETTINGS_PATH):
    try:
        with open(path, "r", encoding="utf-8") as f:
            settings = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        settings = {}
    defaults = {
        "version_filter_start": "",
        "version_filter_end": "",
        "enable_version_filters": True,
        "show_package_mods": True,
        "show_ts4script_mods": True,
        "mod_directory": "",
        "sims_cache_directory": "",
        "backups_directory": "",
        "xls_file_path": "",
        "sims_executable_path": "",
        "sims_executable_arguments": "",
        "log_extra_extensions": [],
        "grab_logs_ignore_files": [],
        "ignored_mods": [],
        "show_ignored": False,
        "show_search_results": True,
        "instant_search": True,
        "file_filter_mode": "both",
        "auto_scan_on_start": True,
        "hidden_columns": [],
        "installer_hidden_columns": [],
        "show_disabled_only": False,
        "log_level": "DEBUG",
        "background_image_path": "",
        "splash_background_image_path": "",
        # UI appearance
        "ui_frame_opacity": 100,
        "language": "fr-fr",
        # Mod Installer – archive handling
        "installer_use_mod_root": True,
        "installer_include_extras": False,
        # Filters
        "hide_installer_mods": False,
        # AI
        "ai_enabled": False,
        "ai_auto_train": True,
        "ai_model_path": "mod_ai.json",
        "ai_group_overrides": {},
        "ai_train_min_samples": 2,
        "ai_train_rare_strategy": "drop",
        "ai_train_class_weight_balanced": True,
        "ai_train_cv_folds": 0,
        "ai_train_mlp_hidden": 128,
        "ai_train_mlp_epochs": 30,
        # Log Manager
        "last_log_path": "",
        # Web interface
        "web_enabled": True,
        "web_host": "127.0.0.1",
        "web_port": 5000,
        "web_debug": False,
    }
    for key, value in defaults.items():
        settings.setdefault(key, value)

    show_search_pref = settings.get("show_search_results")
    if isinstance(show_search_pref, str):
        normalized = show_search_pref.strip().lower()
        settings["show_search_results"] = normalized not in {"false", "0", "non", "no", "off"}
    else:
        settings["show_search_results"] = bool(show_search_pref)

    settings["enable_version_filters"] = bool(settings.get("enable_version_filters", True))
    legacy_combined_filter = settings.pop("filter_package_and_ts4script", None)
    if legacy_combined_filter is True:
        settings["show_package_mods"] = True
        settings["show_ts4script_mods"] = True
    legacy_hide_post_118 = settings.pop("hide_post_118", None)
    legacy_filter_range = settings.pop("filter_116_to_118", None)
    if not settings.get("version_filter_start") and legacy_filter_range:
        settings["version_filter_start"] = "1.116.202.1030"
    if not settings.get("version_filter_end") and (legacy_filter_range or legacy_hide_post_118):
        settings["version_filter_end"] = "1.118.242.1030"

    if isinstance(settings.get("log_extra_extensions"), str):
        settings["log_extra_extensions"] = [part.strip() for part in settings["log_extra_extensions"].split(",") if part.strip()]
    extra_extensions = []
    for entry in settings.get("log_extra_extensions", []):
        if not entry:
            continue
        ext = entry if entry.startswith(".") else f".{entry}"
        extra_extensions.append(ext.lower())
    settings["log_extra_extensions"] = sorted(set(extra_extensions))

    ignore_entries = settings.get("grab_logs_ignore_files", [])
    if isinstance(ignore_entries, str):
        raw_parts = re.split(r"[,;\n]+", ignore_entries)
    else:
        raw_parts = ignore_entries
    normalized_ignore = []
    seen = set()
    for part in raw_parts:
        if not part:
            continue
        cleaned = str(part).strip()
        if not cleaned or cleaned.lower() in seen:
            continue
        seen.add(cleaned.lower())
        normalized_ignore.append(cleaned)
    settings["grab_logs_ignore_files"] = normalized_ignore

    ignored_from_file = load_ignore_list()
    if not ignored_from_file and settings.get("ignored_mods"):
        ignored_from_file = settings.get("ignored_mods", [])
        save_ignore_list(ignored_from_file)
    settings["ignored_mods"] = ignored_from_file
    try:
        logging.getLogger("Sims4ModTool").debug("Settings loaded from %s", path)
    except Exception:
        pass
    return settings

def save_settings(settings, path=SETTINGS_PATH):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=4, ensure_ascii=False)
    try:
        logging.getLogger("Sims4ModTool").debug("Settings saved to %s", path)
    except Exception:
        pass


# ---------- Mod AI (lightweight learner) ----------
class ModAI:
    def __init__(self, data=None):
        self.data = data or {"mods": {}, "token_to_mod": {}}
        self.logger = logging.getLogger("Sims4ModTool.AI")
        self._tfidf_model = None
        self._tfidf_loaded_path = None
        self._mlp_model = None
        self._mlp_loaded_path = None

    @staticmethod
    def _tokenize(text: str):
        return [t for t in re.split(r"[^a-z0-9]+", (text or "").lower()) if t and len(t) > 1]

    @classmethod
    def load(cls, path: str):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = None
        inst = cls(data)
        inst.logger.info("AI model loaded from %s (mods=%d, tokens=%d)", path, len(inst.data.get('mods', {})), len(inst.data.get('token_to_mod', {})))
        return inst

    def save(self, path: str):
        tmp = path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)
        try:
            os.replace(tmp, path)
        except Exception:
            shutil.move(tmp, path)
        self.logger.info("AI model saved to %s (mods=%d, tokens=%d)", path, len(self.data.get('mods', {})), len(self.data.get('token_to_mod', {})))

    def update_from_rows(self, rows: List[Dict[str, object]]):
        mods = self.data.setdefault("mods", {})
        tok2 = self.data.setdefault("token_to_mod", {})
        for row in rows or []:
            name = str(row.get("group") or "").strip()
            if not name:
                continue
            entry = mods.get(name)
            if not isinstance(entry, dict):
                entry = {"tokens": {}, "seen": []}
                mods[name] = entry
            tokens_map = entry.setdefault("tokens", {})
            # filenames tokens
            for key in ("package", "script"):
                base = os.path.splitext(str(row.get(key) or ""))[0]
                for tok in self._tokenize(base):
                    tokens_map[tok] = int(tokens_map.get(tok, 0)) + 1
                    m = tok2.setdefault(tok, {})
                    m[name] = int(m.get(name, 0)) + 1

    def update_from_log_results(self, results: List[Dict[str, object]]):
        mods = self.data.setdefault("mods", {})
        tok2 = self.data.setdefault("token_to_mod", {})
        for item in results or []:
            mod = str(item.get("mod") or "").strip()
            if not mod:
                continue
            entry = mods.get(mod)
            if not isinstance(entry, dict):
                entry = {"tokens": {}, "seen": []}
                mods[mod] = entry
            tokens_map = entry.setdefault("tokens", {})
            text = f"{item.get('type') or ''} {item.get('message') or ''}"
            for tok in self._tokenize(text):
                tokens_map[tok] = int(tokens_map.get(tok, 0)) + 1
                m = tok2.setdefault(tok, {})
                m[mod] = int(m.get(mod, 0)) + 1

    def guess_from_paths_and_text(self, paths: List[str], text: str) -> Tuple[str, float]:
        # Path-based first
        for p in paths or []:
            norm = ("/" + str(p).replace("\\", "/").lstrip("/")).lower()
            m = re.search(r"/mods/([^/]+)/", norm)
            if m:
                # Deterministic when a clear Mods/<folder>/ segment exists
                return m.group(1), 1.0
        # Prefer neural / TF‑IDF models when available
        pred = self._predict_mlp(paths, text)
        if pred is not None:
            return pred
        pred = self._predict_tfidf(paths, text)
        if pred is not None:
            return pred
        # Token voting fallback
        tokens = set(self._tokenize(text))
        votes = {}
        for tok in tokens:
            for name, w in self.data.get("token_to_mod", {}).get(tok, {}).items():
                votes[name] = votes.get(name, 0.0) + float(w)
        if not votes:
            return "", 0.0
        best = max(votes.items(), key=lambda kv: kv[1])
        total = sum(votes.values()) or 1.0
        return best[0], float(best[1]) / float(total)

    def _load_tfidf_model(self):
        try:
            ml = self.data.get("ml", {}) or {}
            path = str(ml.get("tfidf_model_path") or "").strip()
            if not path:
                return None
            if self._tfidf_model is not None and self._tfidf_loaded_path == path:
                return self._tfidf_model
            try:
                import joblib  # type: ignore
            except Exception:
                return None
            model = joblib.load(path)
            self._tfidf_model = model
            self._tfidf_loaded_path = path
            return model
        except Exception:
            return None

    def _predict_tfidf(self, paths: List[str], text: str) -> Optional[Tuple[str, float]]:
        model = self._load_tfidf_model()
        if model is None:
            return None
        # Compose text similar to training corpus: message + basenames of paths
        parts = [text or ""]
        for p in (paths or []):
            try:
                parts.append(os.path.basename(str(p)))
            except Exception:
                continue
        sample = " ".join([t for t in parts if t]).strip()
        if not sample:
            return None
        try:
            pred = model.predict([sample])
            label = str(pred[0]) if isinstance(pred, (list, tuple)) else str(pred)
            conf = 0.75
            try:
                if hasattr(model, "decision_function"):
                    df = model.decision_function([sample])
                    # LinearSVC: decision_function returns margins; handle 1D/2D
                    import math
                    if hasattr(df, "shape"):
                        # If binary, df is shape (n_samples,) else (n_samples, n_classes)
                        if len(getattr(df, "shape", [])) == 1:
                            margin = float(abs(df[0]))
                        else:
                            margin = float(max(df[0]))
                        conf = 1.0 / (1.0 + math.exp(-margin))
            except Exception:
                pass
            return label, max(0.0, min(1.0, float(conf)))
        except Exception:
            return None

    def _load_mlp_model(self):
        try:
            ml = self.data.get("ml", {}) or {}
            path = str(ml.get("mlp_model_path") or "").strip()
            if not path:
                return None
            if self._mlp_model is not None and self._mlp_loaded_path == path:
                return self._mlp_model
            try:
                import joblib  # type: ignore
            except Exception:
                return None
            model = joblib.load(path)
            self._mlp_model = model
            self._mlp_loaded_path = path
            return model
        except Exception:
            return None

    def _predict_mlp(self, paths: List[str], text: str) -> Optional[Tuple[str, float]]:
        model = self._load_mlp_model()
        if model is None:
            return None
        parts = [text or ""]
        for p in (paths or []):
            try:
                parts.append(os.path.basename(str(p)))
            except Exception:
                continue
        sample = " ".join([t for t in parts if t]).strip()
        if not sample:
            return None
        try:
            label = ""
            conf = 0.0
            if hasattr(model, "predict_proba"):
                proba = model.predict_proba([sample])
                if proba is not None and len(proba):
                    probs = proba[0]
                    if hasattr(probs, "tolist"):
                        probs_list = probs.tolist()
                    else:
                        probs_list = list(probs)
                    if probs_list:
                        idx = max(range(len(probs_list)), key=lambda i: probs_list[i])
                        conf = float(probs_list[idx])
                        classes = getattr(model, "classes_", None)
                        if classes is not None and len(classes) > idx:
                            label = str(classes[idx])
            if not label:
                pred = model.predict([sample])
                label = str(pred[0]) if isinstance(pred, (list, tuple)) else str(pred)
                if hasattr(model, "predict_proba") and not conf:
                    try:
                        proba = model.predict_proba([sample])
                        probs = proba[0]
                        if hasattr(probs, "tolist"):
                            probs_list = probs.tolist()
                        else:
                            probs_list = list(probs)
                        if probs_list:
                            conf = float(max(probs_list))
                    except Exception:
                        conf = 0.0
                classes = getattr(model, "classes_", None)
                if hasattr(model, "decision_function"):
                    df = model.decision_function([sample])
                    if hasattr(df, "shape"):
                        import math
                        if len(getattr(df, "shape", [])) == 1:
                            margin = float(abs(df[0]))
                        else:
                            margin = float(max(df[0]))
                        conf = 1.0 / (1.0 + math.exp(-margin))
                if conf <= 0.0:
                    conf = 0.6
            return (label, max(0.0, min(1.0, float(conf)))) if label else None
        except Exception:
            return None

    def update_from_index_entries(self, entries: List[Dict[str, object]]):
        """Update model from Updates Checker index entries.
        Expects dicts with at least 'title'. Tokens are derived from title (and creator when present)."""
        mods = self.data.setdefault('mods', {})
        tok2 = self.data.setdefault('token_to_mod', {})
        for ent in entries or []:
            name = str(ent.get('title') or '').strip()
            if not name:
                continue
            entry = mods.get(name)
            if not isinstance(entry, dict):
                entry = {"tokens": {}, "seen": []}
                mods[name] = entry
            tokens_map = entry.setdefault('tokens', {})
            base_text = name
            creator = str(ent.get('creator') or '').strip()
            if creator:
                base_text += ' ' + creator
            for tok in self._tokenize(base_text):
                tokens_map[tok] = int(tokens_map.get(tok, 0)) + 1
                m = tok2.setdefault(tok, {})
                m[name] = int(m.get(name, 0)) + 1


# ---------- AI helpers (deterministic grouping and reinstall checks) ----------
def ai_resolve_group_from_paths(mods_root: str, paths: List[str]) -> Tuple[str, bool]:
    """Return (group_name, deterministic) using path rules only.
    Deterministic when all files map to the same first-level folder under Mods.
    """
    try:
        root = os.path.abspath(mods_root or "")
    except Exception:
        root = mods_root or ""
    segments = set()
    for p in (paths or []):
        if not p:
            continue
        try:
            ap = os.path.abspath(p)
        except Exception:
            ap = p
        if root and ap.lower().startswith(os.path.abspath(root).lower()):
            rel = os.path.relpath(ap, root).replace("\\", "/")
            first = (rel.split("/") or [""])[0].strip()
        else:
            first = os.path.basename(os.path.dirname(ap))
        if first:
            segments.add(first)
    if len(segments) == 1:
        return segments.pop(), True
    return "", False


# ---------- Log parsing helpers ----------
def _strip_html_to_text(content: str) -> str:
    try:
        if BeautifulSoup is not None:
            soup = BeautifulSoup(content, "html.parser")
            return soup.get_text("\n", strip=False)
    except Exception:
        pass
    return re.sub(r"<[^>]+>", "", content)

def _extract_exception_blocks(lines: List[str]) -> List[Dict[str, object]]:
    blocks = []
    i = 0
    n = len(lines)
    tp = re.compile(r"^\s*Traceback \(most recent call last\):")
    fp = re.compile(r"\bFile \"([^\"]+)\", line (\d+)")
    ep = re.compile(r"^\s*([A-Za-z_][A-Za-z0-9_]*(?:Error|Exception))\s*:\s*(.*)$")
    while i < n:
        if not tp.search(lines[i] or ""):
            i += 1
            continue
        ctx = [lines[i]]
        i += 1
        paths = []
        while i < n and lines[i].strip():
            ctx.append(lines[i])
            m = fp.search(lines[i])
            if m:
                paths.append(m.group(1))
            i += 1
        exc_type = ""; exc_msg = ""
        j = i
        while j < min(n, i + 10):
            m = ep.match(lines[j] or "")
            if m:
                exc_type, exc_msg = m.group(1), m.group(2)
                ctx.append(lines[j])
                break
            j += 1
        blocks.append({"type": exc_type, "message": exc_msg, "paths": paths, "context": ctx})
        i = j + 1
    return blocks

def analyze_last_exception_html(content: str) -> Dict[str, object]:
    text = _strip_html_to_text(content)
    lines = text.splitlines()
    sims_version = ""
    for ln in lines[:500]:
        if "Sims 4 Version:" in ln:
            sims_version = ln.split(":", 1)[-1].strip()
            break
    results = []
    for blk in _extract_exception_blocks(lines):
        paths = blk.get("paths") or []
        mod = ""
        for p in paths:
            m = re.search(r"(?i)/Mods/([^/]+)/", ("/" + p.replace("\\", "/").lstrip("/")))
            if m:
                mod = m.group(1)
                break
        results.append({"type": blk.get("type", ""), "message": blk.get("message", ""), "paths": paths, "mod": mod})
    # fallback by scanning paths
    if not results:
        path_pat = re.compile(r"(?i)([A-Za-z]:\\[^\n\r]*?\\Mods\\[^\n\r\"']+|/[^\n\r]*?/Mods/[^\n\r\"']+)")
        hits = []
        for ln in lines:
            hits += [m.group(1) for m in path_pat.finditer(ln)]
        if hits:
            mods = []
            for h in hits:
                m = re.search(r"(?i)/Mods/([^/]+)/", ("/" + h.replace("\\", "/").lstrip("/")))
                if m:
                    mods.append(m.group(1))
            mods = list(dict.fromkeys(mods))
            for name in mods[:10]:
                results.append({"type": "MCCC", "message": "", "paths": [h for h in hits if f"/Mods/{name}/" in h.replace('\\','/')], "mod": name})
    return {"sims_version": sims_version, "results": results}

def analyze_generic_log_text(content: str) -> List[Dict[str, object]]:
    lines = content.splitlines()
    results = []
    for blk in _extract_exception_blocks(lines):
        paths = blk.get("paths") or []
        mod = ""
        for p in paths:
            m = re.search(r"(?i)/Mods/([^/]+)/", ("/" + p.replace("\\", "/").lstrip("/")))
            if m:
                mod = m.group(1)
                break
        results.append({"type": blk.get("type", ""), "message": blk.get("message", ""), "paths": paths, "mod": mod})
    return results

def scan_directory(directory, progress_callback=None, *, recursive=True):
    _log = logging.getLogger("Sims4ModTool")
    start_time = datetime.utcnow()
    try:
        _log.info("Scan start: %s (recursive=%s)", directory, bool(recursive))
    except Exception:
        pass
    package_files = {}
    ts4script_files = {}
    snapshot_entries = []
    normalized_root = os.path.abspath(directory)
    relevant_files = []
    if recursive:
        for root, dirs, files in os.walk(directory):
            for file in files:
                lower_name = file.lower()
                if lower_name.endswith((".package", ".ts4script")):
                    full_path = os.path.join(root, file)
                    relevant_files.append((file, lower_name, full_path))
    else:
        try:
            for file in os.listdir(directory):
                full_path = os.path.join(directory, file)
                if not os.path.isfile(full_path):
                    continue
                lower_name = file.lower()
                if lower_name.endswith((".package", ".ts4script")):
                    relevant_files.append((file, lower_name, full_path))
        except OSError:
            pass

    total_files = len(relevant_files)
    if progress_callback is not None:
        try:
            progress_callback(0, total_files, "")
        except Exception:
            pass

    def _stat_safe(path):
        try:
            return os.stat(path)
        except OSError:
            return None

    # Populate extension maps quickly (order-preserving)
    for file, lower_name, full_path in relevant_files:
        if lower_name.endswith(".package"):
            package_files[file] = full_path
        else:
            ts4script_files[file] = full_path

    # I/O-bound: parallelize stat collection for faster scans
    max_workers = min(32, max(4, (os.cpu_count() or 4) * 2))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        for index, ((file, lower_name, full_path), stat_result) in enumerate(zip(relevant_files, executor.map(_stat_safe, [t[2] for t in relevant_files])), start=1):
            if not stat_result:
                continue
            relative_path = os.path.relpath(full_path, normalized_root)
            snapshot_entries.append({
                "path": relative_path.replace("\\", "/"),
                "mtime": int(stat_result.st_mtime),
                "size": int(stat_result.st_size),
                "type": "package" if lower_name.endswith(".package") else "ts4script",
            })
            if progress_callback is not None:
                try:
                    progress_callback(index, total_files, full_path)
                except Exception:
                    pass
    snapshot_entries.sort(key=lambda item: item["path"].casefold())
    snapshot = {
        "root": normalized_root.replace("\\", "/"),
        "generated_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "entries": snapshot_entries,
    }
    try:
        elapsed_ms = (datetime.utcnow() - start_time).total_seconds() * 1000.0
        _log.info(
            "Scan complete: relevant_files=%d, entries=%d, elapsed=%.1fms",
            len(relevant_files), len(snapshot_entries), elapsed_ms
        )
    except Exception:
        pass
    return package_files, ts4script_files, snapshot


def normalize_mod_basename(name):
    if not name:
        return ""
    base_name = os.path.splitext(os.path.basename(name))[0]
    # Strip trailing version suffixes like _v1.2.3 or -V1.23
    base_low = base_name
    try:
        _ver_suffix = re.compile(r"(?i)[ _-]v\d+(?:\.\d+){1,3}$")
        while True:
            new_low = _ver_suffix.sub("", base_low)
            if new_low == base_low:
                break
            base_low = new_low
        base_low = base_low.strip()
    except Exception:
        pass
    normalized = MOD_NAME_SANITIZE_RE.sub("", base_low.casefold())
    return normalized


def stable_mod_name_hash(normalized_name):
    if not normalized_name:
        return 0
    digest = hashlib.sha1(normalized_name.encode("utf-8", errors="ignore")).hexdigest()
    return int(digest, 16)


def similarity_confidence_label(ratio):
    if ratio >= 0.9:
        return "Élevée"
    if ratio >= 0.75:
        return "Moyenne"
    return "Faible"


def extract_version_from_name(name: str) -> str:
    """Extract an explicit version from a filename if present.
    Supports patterns like _v1.2.3, -V1.2, or full Sims version 1.118.257.1020.
    """
    if not name:
        return ""
    base = os.path.splitext(os.path.basename(name))[0]
    # Prefer Sims full version-like first
    m = re.search(r"\b(\d+\.\d+\.\d+\.\d+)\b", base)
    if m:
        return m.group(1)
    # Fallback: vX.Y[.Z]
    m = re.search(r"(?i)[ _-]v(\d+(?:\.\d+){1,3})$", base)
    if m:
        return m.group(1)
    return ""

def generate_data_rows(directory, settings, version_releases, progress_callback=None, yield_callback=None, notify_callback=None, *, recursive=True):
    t0 = datetime.utcnow()
    package_files, ts4script_files, snapshot = scan_directory(directory, progress_callback=progress_callback, recursive=recursive)
    try:
        logging.getLogger("Sims4ModTool").info(
            "Pairing start: packages=%d, scripts=%d",
            len(package_files), len(ts4script_files)
        )
    except Exception:
        pass
    # Build path->mod group mapping from installed_mods.json if available
    group_by_path = {}
    disabled_by_path = {}
    try:
        installed = load_installed_mods()
    except Exception:
        installed = []
    atf_by_name = set()
    for entry in installed:
        target_folder = entry.get("target_folder") or ""
        name = entry.get("name") or os.path.basename(target_folder) or ""
        if not target_folder:
            continue
        if entry.get("atf") and name:
            atf_by_name.add(name)
        root = os.path.normcase(os.path.abspath(target_folder))
        # direct files list
        disabled_flag = bool(entry.get("disabled", False))
        for rel in (entry.get("files") or []):
            rel_norm = str(rel).replace("\\", "/").strip().lstrip("/")
            if not rel_norm:
                continue
            abs_path = os.path.join(root, *rel_norm.split("/"))
            key = os.path.normcase(os.path.abspath(abs_path))
            group_by_path[key] = name
            disabled_by_path[key] = disabled_flag
        # addons paths
        for addon in (entry.get("addons") or []):
            for rel in (addon.get("paths") or []):
                rel_norm = str(rel).replace("\\", "/").strip().lstrip("/")
                if not rel_norm:
                    continue
                abs_path = os.path.join(root, *rel_norm.split("/"))
                key = os.path.normcase(os.path.abspath(abs_path))
                group_by_path[key] = name
                disabled_by_path[key] = disabled_flag
    previous_snapshot = load_mod_scan_cache()
    snapshot_changed = previous_snapshot is not None and not mod_scan_snapshots_equal(previous_snapshot, snapshot)
    if notify_callback:
        try:
            notify_callback("Écriture du cache…")
        except Exception:
            pass
    save_mod_scan_cache(snapshot)
    version_filters_enabled = settings.get("enable_version_filters", True)
    start_version = settings.get("version_filter_start") or ""
    end_version = settings.get("version_filter_end") or ""
    if not version_filters_enabled:
        start_version = ""
        end_version = ""
    start_date = version_releases.get(start_version)
    end_date = version_releases.get(end_version)
    start_limit = datetime.combine(start_date, time.min) if start_date else None
    latest_version_key = next(reversed(version_releases)) if version_releases else None
    if end_version and latest_version_key and end_version == latest_version_key:
        end_limit = datetime.combine(date.today(), time.max)
    else:
        end_limit = datetime.combine(end_date, time.max) if end_date else None
    if start_limit and end_limit and start_limit > end_limit:
        start_limit, end_limit = end_limit, start_limit

    data_rows = []
    throttle_counter = 0

    def _maybe_yield():
        nonlocal throttle_counter
        if yield_callback is None:
            return
        throttle_counter += 1
        if throttle_counter % 25 != 0:
            return
        try:
            yield_callback()
        except Exception:
            pass
    ignored_mods = set(settings.get("ignored_mods", []))
    show_ignored = settings.get("show_ignored", False)
    # derive file filter mode
    mode = (settings.get("file_filter_mode") or "both").strip().lower()
    if mode not in {"both", "package", "ts4script", "installer_only"}:
        mode = "both"
    show_packages = mode in {"both", "package", "installer_only"}
    show_scripts = mode in {"both", "ts4script", "installer_only"}
    hide_installer_mods = bool(settings.get("hide_installer_mods", False)) and mode != "installer_only"

    def _resolve_parent(path):
        if not path:
            return ""
        return os.path.normcase(os.path.abspath(os.path.dirname(path)))

    package_entries = {}
    for pkg, pkg_path in package_files.items():
        normalized_name = normalize_mod_basename(pkg)
        hash_source = normalized_name or os.path.splitext(pkg)[0].casefold()
        package_entries[pkg] = {
            "path": pkg_path,
            "base": os.path.splitext(pkg)[0],
            "normalized": normalized_name,
            "parent": _resolve_parent(pkg_path),
            "hash": stable_mod_name_hash(hash_source),
        }

    script_entries = {}
    for script, script_path in ts4script_files.items():
        normalized_name = normalize_mod_basename(script)
        hash_source = normalized_name or os.path.splitext(script)[0].casefold()
        script_entries[script] = {
            "path": script_path,
            "base": os.path.splitext(script)[0],
            "normalized": normalized_name,
            "parent": _resolve_parent(script_path),
            "hash": stable_mod_name_hash(hash_source),
        }

    unpaired_packages = set(package_entries.keys())
    unpaired_scripts = set(script_entries.keys())
    matches: Dict[str, Dict[str, str]] = {}

    # Pass 1 – même base normalisée
    scripts_by_norm = defaultdict(list)
    for script_name in unpaired_scripts:
        scripts_by_norm[script_entries[script_name]["normalized"]].append(script_name)
    for script_list in scripts_by_norm.values():
        script_list.sort(key=str.casefold)

    for pkg_name in list(unpaired_packages):
        pkg_info = package_entries[pkg_name]
        norm_name = pkg_info["normalized"]
        if not norm_name:
            continue
        candidates = scripts_by_norm.get(norm_name)
        if not candidates:
            continue
        script_name = next((candidate for candidate in candidates if candidate in unpaired_scripts), None)
        if not script_name:
            continue
        normalized_display = norm_name or pkg_info["base"].casefold()
        matches[pkg_name] = {
            "script": script_name,
            "confidence": "Élevée",
            "tooltip": (
                f"Appariement basé sur un nom normalisé identique ({normalized_display})."
            ),
        }
        unpaired_packages.remove(pkg_name)
        unpaired_scripts.remove(script_name)

    # Pass 2 – même dossier parent (require similarity >= threshold)
    scripts_by_parent = defaultdict(list)
    for script_name in unpaired_scripts:
        scripts_by_parent[script_entries[script_name]["parent"]].append(script_name)
    for script_list in scripts_by_parent.values():
        script_list.sort(key=str.casefold)

    for pkg_name in list(unpaired_packages):
        pkg_info = package_entries[pkg_name]
        parent = pkg_info["parent"]
        candidates = scripts_by_parent.get(parent)
        if not candidates:
            continue
        # Choose best candidate by normalized name similarity
        norm_pkg = pkg_info.get("normalized") or ""
        best_name = None
        best_ratio = 0.0
        for candidate in candidates:
            if candidate not in unpaired_scripts:
                continue
            norm_scr = script_entries[candidate].get("normalized") or ""
            if len(norm_pkg) < MIN_NAME_LENGTH or len(norm_scr) < MIN_NAME_LENGTH:
                continue
            r = SequenceMatcher(None, norm_pkg, norm_scr).ratio()
            if r >= MIN_SIMILARITY_RATIO and r > best_ratio:
                best_ratio = r
                best_name = candidate
        script_name = best_name
        if not script_name:
            continue
        try:
            rel_parent = os.path.relpath(parent, directory) if parent else "."
        except ValueError:
            rel_parent = parent
        if rel_parent in (".", ""):
            folder_display = "(racine du dossier mods)"
        else:
            folder_display = rel_parent.replace("\\", "/")
        tooltip = (
            f"Appariement basé sur le même dossier parent : {folder_display}."
        )
        matches[pkg_name] = {
            "script": script_name,
            "confidence": "Moyenne",
            "tooltip": tooltip,
        }
        unpaired_packages.remove(pkg_name)
        unpaired_scripts.remove(script_name)

    # Pass 3 – similarité + fallback hash
    scripts_by_prefix = defaultdict(list)
    scripts_by_length = defaultdict(list)
    for script_name in unpaired_scripts:
        info = script_entries[script_name]
        normalized = info["normalized"] or info["base"].casefold()
        if normalized:
            if len(normalized) >= 2:
                scripts_by_prefix[(2, normalized[:2])].append(script_name)
            scripts_by_prefix[(1, normalized[:1])].append(script_name)
        else:
            scripts_by_prefix[(0, "")].append(script_name)
        scripts_by_length[len(normalized)].append(script_name)
    for script_list in scripts_by_prefix.values():
        script_list.sort(key=str.casefold)

    similarity_candidates = []
    for pkg_name in list(unpaired_packages):
        pkg_info = package_entries[pkg_name]
        normalized = pkg_info["normalized"] or pkg_info["base"].casefold()
        if not normalized and not pkg_info["base"]:
            continue
        prefixes = []
        if normalized:
            if len(normalized) >= 2:
                prefixes.append((2, normalized[:2]))
            prefixes.append((1, normalized[:1]))
        else:
            base_lower = pkg_info["base"].casefold()
            if len(base_lower) >= 2:
                prefixes.append((2, base_lower[:2]))
            if base_lower:
                prefixes.append((1, base_lower[:1]))
        if not prefixes:
            prefixes.append((0, ""))

        seen_scripts = set()
        for prefix in prefixes:
            for script_name in scripts_by_prefix.get(prefix, []):
                if script_name not in unpaired_scripts or script_name in seen_scripts:
                    continue
                seen_scripts.add(script_name)
                script_info = script_entries[script_name]
                script_norm = script_info["normalized"] or script_info["base"].casefold()
                if len((normalized or "")) < MIN_NAME_LENGTH or len((script_norm or "")) < MIN_NAME_LENGTH:
                    continue
                ratio = SequenceMatcher(None, normalized, script_norm).ratio()
                if ratio < MIN_SIMILARITY_RATIO:
                    continue
                hash_gap = abs(pkg_info["hash"] - script_info["hash"])
                similarity_candidates.append((ratio, hash_gap, pkg_name, script_name))
                if len(seen_scripts) >= MAX_SIMILARITY_CANDIDATES:
                    break
            if len(seen_scripts) >= MAX_SIMILARITY_CANDIDATES:
                break

        if not seen_scripts:
            target_length = len(normalized)
            length_offsets = [0, 1, -1, 2, -2]
            for offset in length_offsets:
                candidate_length = target_length + offset
                if candidate_length < 0:
                    continue
                for script_name in scripts_by_length.get(candidate_length, []):
                    if script_name not in unpaired_scripts or script_name in seen_scripts:
                        continue
                    seen_scripts.add(script_name)
                    script_info = script_entries[script_name]
                    script_norm = script_info["normalized"] or script_info["base"].casefold()
                    if len((normalized or "")) < MIN_NAME_LENGTH or len((script_norm or "")) < MIN_NAME_LENGTH:
                        continue
                    ratio = SequenceMatcher(None, normalized, script_norm).ratio()
                    if ratio < MIN_SIMILARITY_RATIO:
                        continue
                    hash_gap = abs(pkg_info["hash"] - script_info["hash"])
                    similarity_candidates.append((ratio, hash_gap, pkg_name, script_name))
                    if len(seen_scripts) >= MAX_SIMILARITY_CANDIDATES:
                        break
                if len(seen_scripts) >= MAX_SIMILARITY_CANDIDATES:
                    break

    similarity_candidates.sort(key=lambda item: (-item[0], item[1], item[2], item[3]))

    for ratio, hash_gap, pkg_name, script_name in similarity_candidates:
        if pkg_name not in unpaired_packages or script_name not in unpaired_scripts:
            continue
        confidence = similarity_confidence_label(ratio)
        hash_note = f" (écart hash {hash_gap})" if hash_gap else ""
        tooltip = (
            f"Appariement basé sur la similarité des noms (SequenceMatcher {ratio:.2f}). "
            f"Départage assuré par un hash stable pour les égalités{hash_note}."
        )
        matches[pkg_name] = {
            "script": script_name,
            "confidence": confidence,
            "tooltip": tooltip,
        }
        unpaired_packages.remove(pkg_name)
        unpaired_scripts.remove(script_name)

    # Log pairing summary before rendering
    try:
        logging.getLogger("Sims4ModTool").info(
            "Pairing summary: exact+parent+similar matches=%d, remaining: packages=%d, scripts=%d",
            len(matches), len(unpaired_packages), len(unpaired_scripts)
        )
    except Exception:
        pass

    # Lignes finales
    for pkg, pkg_info in package_entries.items():
        pkg_path = pkg_info["path"]
        pkg_date = get_file_date(pkg_path)
        match_info = matches.get(pkg)
        script_name = match_info["script"] if match_info else ""
        script_path = ts4script_files.get(script_name) if script_name else None
        script_date = get_file_date(script_path) if script_path else None

        mod_latest_date = max((dt for dt in (pkg_date, script_date) if dt is not None), default=None)

        if end_limit and mod_latest_date and mod_latest_date > end_limit:
            continue
        if start_limit and mod_latest_date and mod_latest_date < start_limit:
            continue

        has_package = True
        has_script = script_path is not None
        if not ((has_package and show_packages) or (has_script and show_scripts)):
            continue

        candidates = [name for name in (pkg, script_name if script_path else None) if name]
        ignored = any(name in ignored_mods for name in candidates)
        if ignored and not show_ignored:
            continue

        status = "X" if script_path else "MS"
        # Prefer explicit version extracted from filenames, fallback to estimate
        version = extract_version_from_name(pkg_info.get("base") or "") or (
            extract_version_from_name(os.path.splitext(script_name)[0]) if script_name else ""
        ) or estimate_version_from_dates(pkg_date, script_date, version_releases)
        confidence_value = match_info["confidence"] if match_info else "—"
        confidence_tooltip = match_info["tooltip"] if match_info else "Aucun appariement détecté."

        # Resolve group and disabled via tracked paths
        group_value = ""
        disabled_value = False
        is_installer_item = False
        for p in (pkg_path, script_path):
            if not p:
                continue
            key = os.path.normcase(os.path.abspath(p))
            if not group_value:
                group_value = group_by_path.get(key, "")
            if not disabled_value:
                disabled_value = bool(disabled_by_path.get(key, False))
                if key in group_by_path:
                    is_installer_item = True

        # If still missing, resolve deterministically from paths (first-level folder)
        if not group_value:
            try:
                name_from_paths, ok = ai_resolve_group_from_paths(directory, [x for x in (pkg_path, script_path) if x])
                if ok and name_from_paths:
                    group_value = name_from_paths
            except Exception:
                pass

        # AI group overrides when group missing
        if not group_value:
            try:
                overrides = settings.get("ai_group_overrides", {}) or {}
                n_pkg = normalize_mod_basename(pkg_info.get("base") or "")
                n_scr = normalize_mod_basename(os.path.splitext(script_name)[0]) if script_name else ""
                for cand in (n_pkg, n_scr):
                    if cand and cand in overrides:
                        group_value = overrides.get(cand) or group_value
                        if group_value:
                            break
            except Exception:
                pass

        # Final fallback for non‑installer items: group by filename stem (more granular than folder)
        if not group_value:
            try:
                base = pkg_info.get("base") or (os.path.splitext(script_name)[0] if script_name else "")
                group_value = normalize_mod_basename(base or "") or (base or "")
            except Exception:
                pass

        if mode == "installer_only" and not is_installer_item:
            _maybe_yield()
            continue
        if hide_installer_mods and is_installer_item:
            _maybe_yield()
            continue
        if is_installer_item:
            base_tip = "Installé via Mod Installer."
            confidence_value = "100%"
            if confidence_tooltip and confidence_tooltip != "Aucun appariement détecté.":
                confidence_tooltip = f"{base_tip} {confidence_tooltip}"
            else:
                confidence_tooltip = base_tip
        data_rows.append({
            "status": status,
            "group": group_value,
            "disabled": disabled_value,
            "installer": is_installer_item,
            "package": pkg,
            "package_date": format_datetime(pkg_date),
            "script": script_name if script_path else "",
            "script_date": format_datetime(script_date),
            "version": version,
            "confidence": confidence_value,
            "confidence_tooltip": confidence_tooltip,
            "ignored": ignored,
            "ignore_candidates": candidates or [pkg],
            "paths": [path for path in (pkg_path, script_path) if path],
            "atf": bool(group_value and group_value in atf_by_name),
        })
        _maybe_yield()

    for script_name in sorted(unpaired_scripts, key=str.casefold):
        script_path = ts4script_files.get(script_name)
        if not script_path:
            continue
        script_date = get_file_date(script_path)

        if end_limit and script_date and script_date > end_limit:
            continue
        if start_limit and script_date and script_date < start_limit:
            continue
        if not show_scripts:
            continue

        candidates = [script_name]
        ignored = any(name in ignored_mods for name in candidates)
        if ignored and not show_ignored:
            continue

        status = "MP"
        version = extract_version_from_name(os.path.splitext(script_name)[0]) or estimate_version_from_dates(None, script_date, version_releases)

        group_value = ""
        key = os.path.normcase(os.path.abspath(script_path)) if script_path else None
        disabled_value = False
        is_installer_item = False
        if key:
            group_value = group_by_path.get(key, "")
            disabled_value = bool(disabled_by_path.get(key, False))
            if key in group_by_path:
                is_installer_item = True
        if not group_value:
            try:
                name_from_paths, ok = ai_resolve_group_from_paths(directory, [script_path] if script_path else [])
                if ok and name_from_paths:
                    group_value = name_from_paths
            except Exception:
                pass
        if not group_value:
            try:
                base = os.path.splitext(script_name)[0] if script_name else ""
                group_value = normalize_mod_basename(base or "") or (base or "")
            except Exception:
                pass
        if mode == "installer_only" and not is_installer_item:
            _maybe_yield()
            continue
        if hide_installer_mods and is_installer_item:
            _maybe_yield()
            continue
        confidence_value = "—"
        confidence_tooltip = "Aucun package correspondant trouvé."
        if is_installer_item:
            base_tip = "Installé via Mod Installer."
            confidence_value = "100%"
            confidence_tooltip = base_tip
        data_rows.append({
            "status": status,
            "group": group_value,
            "disabled": disabled_value,
            "installer": is_installer_item,
            "package": "",
            "package_date": "",
            "script": script_name,
            "script_date": format_datetime(script_date),
            "version": version,
            "confidence": confidence_value,
            "confidence_tooltip": confidence_tooltip,
            "ignored": ignored,
            "ignore_candidates": candidates,
            "paths": [script_path],
            "atf": bool(group_value and group_value in atf_by_name),
        })
        _maybe_yield()

    if yield_callback is not None:
        try:
            yield_callback()
        except Exception:
            pass

    try:
        elapsed = (datetime.utcnow() - t0).total_seconds() * 1000.0
        logging.getLogger("Sims4ModTool").info(
            "Pairing complete: rows=%d, snapshot_changed=%s, elapsed=%.1fms",
            len(data_rows), bool(snapshot_changed), elapsed
        )
    except Exception:
        pass
    return data_rows, snapshot_changed

def export_to_excel(save_path, data_rows, headers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mods"

    headers = list(headers or [])
    for idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=h)

    for row_idx, row in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(save_path)


class AddUpdateDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ajouter une mise à jour")
        self.setModal(True)
        try:
            self.setWindowFlags(self.windowFlags() | QtCore.Qt.Window)
            self.setSizeGripEnabled(True)
        except Exception:
            pass

        layout = QtWidgets.QFormLayout(self)
        self.version_edit = QtWidgets.QLineEdit(self)
        self.version_edit.setPlaceholderText("1.118.257.1020")
        self.date_edit = QtWidgets.QDateEdit(self)
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDisplayFormat("dd/MM/yyyy")
        self.date_edit.setDate(QtCore.QDate.currentDate())

        layout.addRow("Numéro de version :", self.version_edit)
        layout.addRow("Date de sortie :", self.date_edit)

        button_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, self)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def accept(self):
        if not self.version_edit.text().strip():
            QtWidgets.QMessageBox.warning(self, "Informations manquantes", "Indique un numéro de version valide.")
            return
        super().accept()

    def get_values(self):
        return self.version_edit.text().strip(), self.date_edit.date()


class ConfigurationDialog(QtWidgets.QDialog):
    def __init__(self, parent, settings):
        super().__init__(parent)
        self.setWindowTitle("Configuration")
        self.setModal(True)
        try:
            self.setWindowFlags(self.windowFlags() | QtCore.Qt.Window)
            self.setSizeGripEnabled(True)
        except Exception:
            pass
        self.resize(780, 680)
        self._parent = parent

        layout = QtWidgets.QVBoxLayout()

        # Section: Chemins
        section_paths = QtWidgets.QLabel("Chemins")
        f = section_paths.font()
        f.setBold(True)
        section_paths.setFont(f)
        layout.addWidget(section_paths)

        self.mod_directory_edit = QtWidgets.QLineEdit(self)
        self.mod_directory_edit.setText(settings.get("mod_directory", ""))
        mod_dir_browse = QtWidgets.QPushButton("Parcourir...")
        mod_dir_browse.clicked.connect(lambda: self._browse_directory(self.mod_directory_edit))

        mod_dir_layout = QtWidgets.QHBoxLayout()
        mod_dir_layout.addWidget(QtWidgets.QLabel("Dossier des mods :"))
        mod_dir_layout.addWidget(self.mod_directory_edit)
        mod_dir_layout.addWidget(mod_dir_browse)
        layout.addLayout(mod_dir_layout)

        self.cache_directory_edit = QtWidgets.QLineEdit(self)
        self.cache_directory_edit.setText(settings.get("sims_cache_directory", ""))
        cache_dir_browse = QtWidgets.QPushButton("Parcourir...")
        cache_dir_browse.clicked.connect(lambda: self._browse_directory(self.cache_directory_edit))

        cache_dir_layout = QtWidgets.QHBoxLayout()
        cache_dir_layout.addWidget(QtWidgets.QLabel("Dossier caches sims :"))
        cache_dir_layout.addWidget(self.cache_directory_edit)
        cache_dir_layout.addWidget(cache_dir_browse)
        layout.addLayout(cache_dir_layout)

        self.backups_directory_edit = QtWidgets.QLineEdit(self)
        self.backups_directory_edit.setText(settings.get("backups_directory", ""))
        backups_dir_browse = QtWidgets.QPushButton("Parcourir...")
        backups_dir_browse.clicked.connect(lambda: self._browse_directory(self.backups_directory_edit))

        backups_dir_layout = QtWidgets.QHBoxLayout()
        backups_dir_layout.addWidget(QtWidgets.QLabel("Dossier Backups :"))
        backups_dir_layout.addWidget(self.backups_directory_edit)
        backups_dir_layout.addWidget(backups_dir_browse)
        layout.addLayout(backups_dir_layout)

        layout.addSpacing(6)
        section_exec = QtWidgets.QLabel("Sims 4")
        section_exec.setFont(f)
        layout.addWidget(section_exec)

        self.sims_executable_edit = QtWidgets.QLineEdit(self)
        self.sims_executable_edit.setText(settings.get("sims_executable_path", ""))
        sims_exec_browse = QtWidgets.QPushButton("Parcourir...")
        sims_exec_browse.clicked.connect(self._browse_executable)

        sims_exec_layout = QtWidgets.QHBoxLayout()
        sims_exec_layout.addWidget(QtWidgets.QLabel("Exécutable TS4_X64.exe :"))
        sims_exec_layout.addWidget(self.sims_executable_edit)
        sims_exec_layout.addWidget(sims_exec_browse)
        layout.addLayout(sims_exec_layout)

        self.sims_arguments_edit = QtWidgets.QLineEdit(self)
        self.sims_arguments_edit.setText(settings.get("sims_executable_arguments", ""))
        self.sims_arguments_edit.setPlaceholderText("Arguments supplémentaires (ex: -w)")

        sims_args_layout = QtWidgets.QHBoxLayout()
        sims_args_layout.addWidget(QtWidgets.QLabel("Arguments TS4_X64.exe :"))
        sims_args_layout.addWidget(self.sims_arguments_edit)
        layout.addLayout(sims_args_layout)

        layout.addSpacing(6)
        section_logs = QtWidgets.QLabel("Logs")
        section_logs.setFont(f)
        layout.addWidget(section_logs)

        self.log_extensions_edit = QtWidgets.QLineEdit(self)
        extra_extensions = ", ".join(settings.get("log_extra_extensions", []))
        self.log_extensions_edit.setText(extra_extensions)
        self.log_extensions_edit.setPlaceholderText("Extensions supplémentaires (.mdmp, .html, ...)")

        logs_ext_layout = QtWidgets.QHBoxLayout()
        logs_ext_layout.addWidget(QtWidgets.QLabel("Extensions de logs (supplémentaires) :"))
        logs_ext_layout.addWidget(self.log_extensions_edit)
        layout.addLayout(logs_ext_layout)

        self.grab_logs_ignore_edit = QtWidgets.QPlainTextEdit(self)
        self.grab_logs_ignore_edit.setPlaceholderText("last_crash.txt\nExceptionLog.txt")
        ignore_lines = "\n".join(settings.get("grab_logs_ignore_files", []))
        self.grab_logs_ignore_edit.setPlainText(ignore_lines)
        self.grab_logs_ignore_edit.setFixedHeight(100)

        ignore_layout = QtWidgets.QVBoxLayout()
        ignore_layout.addWidget(QtWidgets.QLabel("Fichiers de logs à ignorer (un par ligne ou séparés par des virgules) :"))
        ignore_layout.addWidget(self.grab_logs_ignore_edit)
        layout.addLayout(ignore_layout)

        # Niveau de logs
        self.log_level_combo = QtWidgets.QComboBox(self)
        self.log_level_combo.addItems(["DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"])
        current_level = str(settings.get("log_level", "DEBUG")).upper()
        idx = self.log_level_combo.findText(current_level)
        self.log_level_combo.setCurrentIndex(idx if idx != -1 else 0)

        log_level_layout = QtWidgets.QHBoxLayout()
        log_level_layout.addWidget(QtWidgets.QLabel("Niveau de logs :"))
        log_level_layout.addWidget(self.log_level_combo)
        layout.addLayout(log_level_layout)

        # Mod Installer (archives)
        installer_group = QtWidgets.QGroupBox("Mod Installer (archives)", self)
        installer_layout = QtWidgets.QVBoxLayout(installer_group)
        self.installer_mod_root_checkbox = QtWidgets.QCheckBox(
            "Utiliser la logique Mod Root (ZIP/7z/rar)", installer_group
        )
        self.installer_mod_root_checkbox.setChecked(bool(settings.get("installer_use_mod_root", True)))
        installer_layout.addWidget(self.installer_mod_root_checkbox)
        self.installer_include_extras_checkbox = QtWidgets.QCheckBox(
            "Inclure fichiers non essentiels (docs/images)", installer_group
        )
        self.installer_include_extras_checkbox.setChecked(bool(settings.get("installer_include_extras", False)))
        installer_layout.addWidget(self.installer_include_extras_checkbox)
        layout.addWidget(installer_group)

        # Intelligence Artificielle
        ai_group = QtWidgets.QGroupBox("Intelligence Artificielle", self)
        ai_layout = QtWidgets.QVBoxLayout(ai_group)
        self.ai_enabled_cb = QtWidgets.QCheckBox("Activer l'IA (expérimental)", ai_group)
        self.ai_enabled_cb.setChecked(bool(settings.get("ai_enabled", False)))
        ai_layout.addWidget(self.ai_enabled_cb)

        self.ai_auto_train_cb = QtWidgets.QCheckBox("Auto-train au démarrage", ai_group)
        self.ai_auto_train_cb.setChecked(bool(settings.get("ai_auto_train", True)))
        ai_layout.addWidget(self.ai_auto_train_cb)

        ai_path_row = QtWidgets.QHBoxLayout()
        ai_path_row.addWidget(QtWidgets.QLabel("Fichier modèle (JSON):", ai_group))
        self.ai_model_path_edit = QtWidgets.QLineEdit(ai_group)
        self.ai_model_path_edit.setText(settings.get("ai_model_path", "mod_ai.json"))
        ai_path_row.addWidget(self.ai_model_path_edit, 1)
        ai_browse = QtWidgets.QPushButton("Parcourir…", ai_group)
        def _browse_ai_model():
            path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Choisir le modèle IA", self.ai_model_path_edit.text() or "mod_ai.json", "JSON (*.json)")
            if path:
                self.ai_model_path_edit.setText(path)
        ai_browse.clicked.connect(_browse_ai_model)
        ai_path_row.addWidget(ai_browse)
        ai_layout.addLayout(ai_path_row)
        help_ai = QtWidgets.QLabel("Le classifieur TF‑IDF (si entraîné) est automatiquement préféré pour la prédiction.", ai_group)
        try:
            help_ai.setStyleSheet("color: #90a4ae;")
        except Exception:
            pass
        ai_layout.addWidget(help_ai)
        layout.addWidget(ai_group)

        # Options de démarrage
        self.auto_scan_checkbox = QtWidgets.QCheckBox("Scan automatique au démarrage", self)
        self.auto_scan_checkbox.setChecked(bool(settings.get("auto_scan_on_start", True)))
        layout.addWidget(self.auto_scan_checkbox)

        version_group = QtWidgets.QGroupBox("Gestion des versions de patch", self)
        version_layout = QtWidgets.QVBoxLayout(version_group)
        self.version_list_widget = QtWidgets.QListWidget(version_group)
        self.version_list_widget.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.version_list_widget.setMinimumHeight(150)
        version_layout.addWidget(self.version_list_widget)
        self.add_version_button = QtWidgets.QPushButton("Add update info", version_group)
        self.add_version_button.clicked.connect(self._open_add_version_dialog)
        version_layout.addWidget(self.add_version_button)
        layout.addWidget(version_group)

        self._refresh_version_list()

        # Apparence
        section_appearance = QtWidgets.QLabel("Apparence")
        section_appearance.setFont(f)
        layout.addWidget(section_appearance)

        # Language selection
        lang_row = QtWidgets.QHBoxLayout()
        lang_row.addWidget(QtWidgets.QLabel("Langue :", self))
        self.language_combo = QtWidgets.QComboBox(self)
        self.language_combo.addItem("Français (fr-FR)", "fr-fr")
        self.language_combo.addItem("English (en-US)", "en-us")
        cur_lang = str(settings.get("language", "fr-fr")).lower()
        idx = self.language_combo.findData(cur_lang)
        self.language_combo.setCurrentIndex(idx if idx != -1 else 0)
        lang_row.addWidget(self.language_combo)
        lang_row.addStretch(1)
        layout.addLayout(lang_row)

        self.app_bg_edit = QtWidgets.QLineEdit(self)
        self.app_bg_edit.setText(settings.get("background_image_path", ""))
        app_bg_browse = QtWidgets.QPushButton("Parcourir…")
        app_bg_browse.clicked.connect(lambda: self._browse_image(self.app_bg_edit))
        app_bg_layout = QtWidgets.QHBoxLayout()
        app_bg_layout.addWidget(QtWidgets.QLabel("Fond d'écran (app) :"))
        app_bg_layout.addWidget(self.app_bg_edit)
        app_bg_layout.addWidget(app_bg_browse)
        layout.addLayout(app_bg_layout)

        self.splash_bg_edit = QtWidgets.QLineEdit(self)
        self.splash_bg_edit.setText(settings.get("splash_background_image_path", ""))
        splash_bg_browse = QtWidgets.QPushButton("Parcourir…")
        splash_bg_browse.clicked.connect(lambda: self._browse_image(self.splash_bg_edit))
        splash_bg_layout = QtWidgets.QHBoxLayout()
        splash_bg_layout.addWidget(QtWidgets.QLabel("Fond d'écran (splash) :"))
        splash_bg_layout.addWidget(self.splash_bg_edit)
        splash_bg_layout.addWidget(splash_bg_browse)
        layout.addLayout(splash_bg_layout)

        # Opacité des cadres
        opacity_layout = QtWidgets.QHBoxLayout()
        opacity_layout.addWidget(QtWidgets.QLabel("Opacité des cadres :", self))
        self.opacity_slider = QtWidgets.QSlider(QtCore.Qt.Horizontal, self)
        self.opacity_slider.setRange(0, 100)
        self.opacity_slider.setSingleStep(1)
        try:
            cur_opacity = int(settings.get("ui_frame_opacity", 100))
        except Exception:
            cur_opacity = 100
        self.opacity_slider.setValue(max(0, min(100, cur_opacity)))
        self.opacity_value = QtWidgets.QLabel(f"{self.opacity_slider.value()}%", self)
        self.opacity_slider.valueChanged.connect(lambda v: self.opacity_value.setText(f"{int(v)}%"))
        opacity_layout.addWidget(self.opacity_slider, stretch=1)
        opacity_layout.addWidget(self.opacity_value)
        layout.addLayout(opacity_layout)

        button_box = QtWidgets.QDialogButtonBox()
        save_button = button_box.addButton("Sauvegarder", QtWidgets.QDialogButtonBox.AcceptRole)
        reset_button = button_box.addButton("Reset config", QtWidgets.QDialogButtonBox.DestructiveRole)
        cancel_button = button_box.addButton(QtWidgets.QDialogButtonBox.Cancel)
        save_button.clicked.connect(self._save_configuration)
        reset_button.clicked.connect(self._reset_configuration)
        cancel_button.clicked.connect(self.reject)
        layout.addWidget(button_box)

        self.setLayout(layout)

    def _refresh_version_list(self):
        if not hasattr(self, "version_list_widget"):
            return
        self.version_list_widget.clear()
        if self._parent is None:
            return
        for version, release_date in self._parent.get_version_entries():
            self.version_list_widget.addItem(f"{version} – {format_release_date(release_date)}")

    def _open_add_version_dialog(self):
        if self._parent is None:
            return
        dialog = AddUpdateDialog(self)
        if dialog.exec_() != QtWidgets.QDialog.Accepted:
            return
        version, release_date = dialog.get_values()
        success, message = self._parent.add_version_release(version, release_date)
        if success:
            QtWidgets.QMessageBox.information(self, "Version ajoutée", message)
            self._refresh_version_list()
        else:
            QtWidgets.QMessageBox.warning(self, "Ajout impossible", message)

    def _browse_directory(self, target_edit):
        folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Choisir un dossier")
        if folder:
            target_edit.setText(folder)

    def _browse_executable(self):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Sélectionner TS4_X64.exe",
            "",
            "Executable Sims 4 (TS4_X64.exe);;Tous les fichiers (*)"
        )
        if file_path:
            self.sims_executable_edit.setText(file_path)

    def _browse_image(self, target_edit):
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "Choisir une image",
            "",
            "Images (*.png *.jpg *.jpeg *.bmp *.gif);;Tous les fichiers (*)"
        )
        if file_path:
            target_edit.setText(file_path)

    def _save_configuration(self):
        mod_directory = self.mod_directory_edit.text().strip()
        cache_directory = self.cache_directory_edit.text().strip()
        backups_directory = self.backups_directory_edit.text().strip()
        sims_executable_path = self.sims_executable_edit.text().strip()
        sims_executable_arguments = self.sims_arguments_edit.text().strip()
        log_extensions_text = self.log_extensions_edit.text().strip()
        ignore_text = self.grab_logs_ignore_edit.toPlainText()
        log_level_value = self.log_level_combo.currentText().strip().upper()
        auto_scan_on_start = bool(self.auto_scan_checkbox.isChecked())
        installer_use_mod_root = bool(self.installer_mod_root_checkbox.isChecked())
        installer_include_extras = bool(self.installer_include_extras_checkbox.isChecked())

        extra_extensions = []
        if log_extensions_text:
            for part in re.split(r"[,;\s]+", log_extensions_text):
                cleaned = part.strip()
                if not cleaned:
                    continue
                if not cleaned.startswith("."):
                    cleaned = f".{cleaned}"
                extra_extensions.append(cleaned.lower())

        ignore_files = []
        seen_ignore = set()
        if ignore_text:
            for part in re.split(r"[,;\n]+", ignore_text):
                cleaned = part.strip()
                if not cleaned:
                    continue
                key = cleaned.lower()
                if key in seen_ignore:
                    continue
                seen_ignore.add(key)
                ignore_files.append(cleaned)

        # Persist appearance immediately
        if self._parent is not None:
            try:
                self._parent.settings["background_image_path"] = self.app_bg_edit.text().strip()
                self._parent.settings["splash_background_image_path"] = self.splash_bg_edit.text().strip()
                self._parent.settings["language"] = str(self.language_combo.currentData() or "fr-fr").lower()
                # Persist installer toggles
                self._parent.settings["installer_use_mod_root"] = installer_use_mod_root
                self._parent.settings["installer_include_extras"] = installer_include_extras
                # Persist AI settings
                self._parent.settings["ai_enabled"] = bool(self.ai_enabled_cb.isChecked())
                self._parent.settings["ai_auto_train"] = bool(self.ai_auto_train_cb.isChecked())
                self._parent.settings["ai_model_path"] = self.ai_model_path_edit.text().strip() or "mod_ai.json"
                save_settings(self._parent.settings)
            except Exception:
                pass
        if self._parent is not None:
            self._parent.apply_configuration(
                mod_directory,
                cache_directory,
                backups_directory,
                sims_executable_path,
                sims_executable_arguments,
                sorted(set(extra_extensions)),
                ignore_files,
                log_level_value,
                auto_scan_on_start,
            )
            try:
                # Persist and appliquer l'opacité avant le fond pour aperçu
                if hasattr(self, 'opacity_slider'):
                    try:
                        self._parent.settings["ui_frame_opacity"] = int(self.opacity_slider.value())
                    except Exception:
                        self._parent.settings["ui_frame_opacity"] = 100
                    save_settings(self._parent.settings)
                self._parent._apply_background()
                # Apply AI mode immediately: load/unload model and update label
                try:
                    if bool(self._parent.settings.get("ai_enabled", False)):
                        self._parent.mod_ai = ModAI.load(str(self._parent.settings.get("ai_model_path", "mod_ai.json")))
                    else:
                        self._parent.mod_ai = None
                    self._parent._update_ai_mode_label()
                except Exception:
                    pass
            except Exception:
                pass
        self.accept()

    def _reset_configuration(self):
        reply = QtWidgets.QMessageBox.question(
            self,
            "Reset configuration",
            (
                "Cette action va réinitialiser tous les paramètres à leurs valeurs par défaut,\n"
                "en conservant uniquement les chemins de fichiers/dossiers. Continuer ?"
            ),
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if reply != QtWidgets.QMessageBox.Yes:
            return

        # Conserver seulement les chemins saisis dans l'UI + xls_file_path existant
        kept = {
            "mod_directory": self.mod_directory_edit.text().strip(),
            "sims_cache_directory": self.cache_directory_edit.text().strip(),
            "backups_directory": self.backups_directory_edit.text().strip(),
            "sims_executable_path": self.sims_executable_edit.text().strip(),
            "sims_executable_arguments": self.sims_arguments_edit.text().strip(),
        }
        if self._parent is not None and hasattr(self._parent, "settings"):
            xls_path = self._parent.settings.get("xls_file_path", "")
            if xls_path:
                kept["xls_file_path"] = xls_path

        # Écrase le fichier de configuration avec uniquement les chemins conservés
        save_settings(kept)
        # Recharge dans le parent pour appliquer les valeurs par défaut manquantes
        if self._parent is not None:
            self._parent.settings = load_settings()

        QtWidgets.QMessageBox.information(self, "Reset", "Configuration réinitialisée.")


class FileDropDialog(QtWidgets.QDialog):
    def __init__(self, title, instruction, drop_handler, parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        try:
            self.setWindowFlags(self.windowFlags() | QtCore.Qt.Window)
            self.setSizeGripEnabled(True)
        except Exception:
            pass
        self.resize(520, 300)
        self.setAcceptDrops(True)
        self._drop_handler = drop_handler

        layout = QtWidgets.QVBoxLayout(self)

        info_label = QtWidgets.QLabel(instruction, self)
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        self.drop_label = QtWidgets.QLabel("Déposez vos fichiers ici", self)
        self.drop_label.setAlignment(QtCore.Qt.AlignCenter)
        self.drop_label.setMinimumHeight(100)
        self._drop_idle_style = "QLabel { border: 2px dashed #aaaaaa; padding: 24px; background-color: #3a3a3a; }"
        self._drop_active_style = "QLabel { border: 2px solid #00aa88; padding: 24px; background-color: #2a2a2a; }"
        self.drop_label.setStyleSheet(self._drop_idle_style)
        layout.addWidget(self.drop_label)

        close_button = QtWidgets.QPushButton("Fermer", self)
        close_button.clicked.connect(self.reject)
        layout.addWidget(close_button, alignment=QtCore.Qt.AlignRight)

    def dragEnterEvent(self, event):
        if self._contains_supported_files(event):
            event.acceptProposedAction()
            self._set_drop_active(True)
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        super().dragLeaveEvent(event)
        self._set_drop_active(False)

    def dropEvent(self, event):
        if not event.mimeData().hasUrls():
            event.ignore()
            self._set_drop_active(False)
            return

        file_paths = [url.toLocalFile() for url in event.mimeData().urls() if url.isLocalFile()]
        success_messages, error_messages = self._drop_handler(file_paths)

        if success_messages:
            QtWidgets.QMessageBox.information(
                self,
                "Opération terminée",
                "\n".join(success_messages),
            )
        if error_messages:
            QtWidgets.QMessageBox.warning(
                self,
                "Certaines opérations ont échoué",
                "\n".join(error_messages),
            )

        if success_messages and not error_messages:
            self.accept()
        elif success_messages:
            # Laisser la fenêtre ouverte pour d'éventuels ajouts complémentaires
            self._set_drop_active(False)
        else:
            self._set_drop_active(False)

        event.acceptProposedAction()

    def _contains_supported_files(self, event):
        if not event.mimeData().hasUrls():
            return False
        for url in event.mimeData().urls():
            if not url.isLocalFile():
                continue
            if os.path.splitext(url.toLocalFile())[1].lower() in SUPPORTED_INSTALL_EXTENSIONS:
                return True
        return False

    def _set_drop_active(self, active):
        self.drop_label.setStyleSheet(self._drop_active_style if active else self._drop_idle_style)


class ModInstallerDialog(QtWidgets.QDialog):
    def __init__(self, parent=None, mod_directory=""):
        super().__init__(parent)
        self.setWindowTitle("Mod Installer")
        self.setModal(True)
        self.resize(720, 420)
        try:
            # Allow maximizing and manual resize
            self.setSizeGripEnabled(True)
            self.setMinimumSize(720, 420)
            self.setWindowFlag(QtCore.Qt.WindowMaximizeButtonHint, True)
            self.setWindowFlag(QtCore.Qt.WindowMinimizeButtonHint, True)
        except Exception:
            pass
        self.setAcceptDrops(True)

        # Attach settings from parent if available, otherwise load
        try:
            self.settings = dict(getattr(parent, "settings", load_settings()))
        except Exception:
            self.settings = load_settings()

        self.mod_directory = mod_directory
        self.installations_performed = False
        self.installed_mods = load_installed_mods()

        layout = QtWidgets.QVBoxLayout(self)

        info_label = QtWidgets.QLabel(
            "Glissez-déposez un fichier .package, .ts4script ou .zip pour l'installer dans le dossier des mods configuré."
        )
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        mod_dir_display = mod_directory if mod_directory else "(dossier non défini)"
        self.target_directory_label = QtWidgets.QLabel(f"Dossier cible : {mod_dir_display}")
        self.target_directory_label.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        layout.addWidget(self.target_directory_label)

        self.drop_label = QtWidgets.QLabel("Déposez vos fichiers ici")
        self.drop_label.setAlignment(QtCore.Qt.AlignCenter)
        self.drop_label.setMinimumHeight(80)
        self._drop_idle_style = "QLabel { border: 2px dashed #aaaaaa; padding: 24px; background-color: #3a3a3a; }"
        self._drop_active_style = "QLabel { border: 2px solid #00aa88; padding: 24px; background-color: #2a2a2a; }"
        self.drop_label.setStyleSheet(self._drop_idle_style)
        layout.addWidget(self.drop_label)

        self.table = QtWidgets.QTableWidget(self)
        self.table.setColumnCount(8)
        # URL moved to last; columns are interactive/resizable
        self._installer_headers = [
            "Mod",
            "Type",
            "Installé le",
            "Version",
            "Dossier",
            "Addons",
            "Statut",
            "URL",
        ]
        self.table.setHorizontalHeaderLabels(self._installer_headers)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        try:
            header.setStretchLastSection(False)
        except Exception:
            pass
        self.table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        # Header: right-click to toggle visible columns
        header = self.table.horizontalHeader()
        try:
            header.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
            header.customContextMenuRequested.connect(self._show_installer_header_menu)
        except Exception:
            pass
        self.column_filters: Dict[int, str] = {}
        try:
            self.table.horizontalHeader().sectionClicked.connect(self._on_header_section_clicked)
        except Exception:
            pass
        layout.addWidget(self.table, stretch=1)

        footer_layout = QtWidgets.QHBoxLayout()
        self.settings_button = QtWidgets.QPushButton("Settings", self)
        self.settings_button.clicked.connect(self._open_installer_settings)
        footer_layout.addWidget(self.settings_button)
        self.recovery_button = QtWidgets.QPushButton("Recovery list", self)
        self.recovery_button.clicked.connect(self._recover_from_markers)
        footer_layout.addWidget(self.recovery_button)

        footer_layout.addStretch(1)
        close_button = QtWidgets.QPushButton("Fermer", self)
        close_button.clicked.connect(self.accept)
        footer_layout.addWidget(close_button)
        layout.addLayout(footer_layout)

        self.refresh_table()

    def dragEnterEvent(self, event):
        if self._contains_supported_files(event):
            event.acceptProposedAction()
            self._set_drop_active(True)
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        super().dragLeaveEvent(event)
        self._set_drop_active(False)

    def dropEvent(self, event):
        if not event.mimeData().hasUrls():
            event.ignore()
            self._set_drop_active(False)
            return

        file_paths = [url.toLocalFile() for url in event.mimeData().urls() if url.isLocalFile()]
        success_entries = []
        error_messages = []

        for path in file_paths:
            success, message = self.install_mod_from_path(path)
            if success and message:
                success_entries.append(message)
            elif (not success) and message:
                error_messages.append(message)

        if success_entries:
            QtWidgets.QMessageBox.information(
                self,
                "Installation terminée",
                "\n".join(success_entries),
            )
        if error_messages:
            QtWidgets.QMessageBox.warning(
                self,
                "Installation incomplète",
                "\n".join(error_messages),
            )

        event.acceptProposedAction()
        self._set_drop_active(False)

    def _set_drop_active(self, active):
        self.drop_label.setStyleSheet(self._drop_active_style if active else self._drop_idle_style)

    def _compute_zip_candidates(self, zip_path):
        candidates = []
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                top_dirs = {}
                root_useful = False
                for info in zf.infolist():
                    if info.is_dir():
                        continue
                    name = info.filename.replace('\\', '/').lstrip('/')
                    base = os.path.basename(name).lower()
                    if not (base.endswith('.package') or base.endswith('.ts4script')):
                        continue
                    parts = name.split('/')
                    if len(parts) == 1:
                        root_useful = True
                    else:
                        top = parts[0]
                        top_dirs[top] = top_dirs.get(top, 0) + 1
                if root_useful:
                    candidates.append(("", "Racine (/)", 0))
                for d, cnt in sorted(top_dirs.items(), key=lambda x: (-x[1], x[0].lower())):
                    candidates.append((d + "/", d, cnt))
        except Exception:
            pass
        return candidates

    def _compute_dir_candidates(self, root_dir):
        candidates = []
        root_useful = False
        top_dirs = {}
        try:
            for cur, dirs, files in os.walk(root_dir):
                rel = os.path.relpath(cur, root_dir).replace('\\', '/')
                depth = 0 if rel == '.' else len([p for p in rel.split('/') if p])
                for f in files:
                    low = f.lower()
                    if not (low.endswith('.package') or low.endswith('.ts4script')):
                        continue
                    if depth == 0:
                        root_useful = True
                    elif depth >= 1:
                        top = rel.split('/', 1)[0]
                        top_dirs[top] = top_dirs.get(top, 0) + 1
        except Exception:
            pass
        if root_useful:
            candidates.append(("", "Racine (/)", 0))
        for d, cnt in sorted(top_dirs.items(), key=lambda x: (-x[1], x[0].lower())):
            candidates.append((d + "/", d, cnt))
        return candidates

    def _prompt_install_plan(self, file_path, extension, temp_dir, mods_root, planned_root, planned_dest_name, include_extras_default):
        """Returns (override_root, dest_name_override, include_extras_op).
        override_root: "" for root, 'FolderName/' for a top-level dir, or None for default.
        dest_name_override: string or None.
        include_extras_op: bool.
        """
        dlg = QtWidgets.QDialog(self)
        try:
            dlg.setWindowFlags(dlg.windowFlags() | QtCore.Qt.Window)
            dlg.setSizeGripEnabled(True)
        except Exception:
            pass
        dlg.setWindowTitle("Plan d'installation")
        dlg.setModal(True)
        try:
            dlg.setSizeGripEnabled(True)
        except Exception:
            pass
        layout = QtWidgets.QFormLayout(dlg)

        # Candidates
        if extension == '.zip':
            candidates = self._compute_zip_candidates(file_path)
        else:
            candidates = self._compute_dir_candidates(temp_dir or '') if temp_dir else []

        root_combo = QtWidgets.QComboBox(dlg)
        items = []
        found_index = -1
        for idx, (key, label, count) in enumerate(candidates):
            display = label if count == 0 else f"{label}  (utiles: {count})"
            root_combo.addItem(display, key)
            items.append(key)
            if planned_root is not None and (key or "") == (planned_root or ""):
                found_index = idx
        if found_index >= 0:
            root_combo.setCurrentIndex(found_index)
        layout.addRow("Racine du mod :", root_combo)

        dest_edit = QtWidgets.QLineEdit(dlg)
        dest_edit.setText(planned_dest_name)
        layout.addRow("Dossier destination :", dest_edit)

        include_extras_cb = QtWidgets.QCheckBox("Inclure fichiers non essentiels", dlg)
        include_extras_cb.setChecked(bool(include_extras_default))
        layout.addRow(include_extras_cb)

        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dlg)
        layout.addRow(buttons)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        if dlg.exec_() != QtWidgets.QDialog.Accepted:
            return None, None, include_extras_default

        chosen_root = root_combo.currentData()
        dest_name = dest_edit.text().strip()
        if dest_name and dest_name != planned_dest_name:
            dest_override = dest_name
        else:
            dest_override = None
        if chosen_root == (planned_root or ""):
            root_override = None
        else:
            root_override = chosen_root
        return root_override, dest_override, bool(include_extras_cb.isChecked())

    def _contains_supported_files(self, event):
        if not event.mimeData().hasUrls():
            return False
        for url in event.mimeData().urls():
            if not url.isLocalFile():
                continue
            if self._is_supported_extension(url.toLocalFile()):
                return True
        return False

    @staticmethod
    def _is_supported_extension(file_path):
        return os.path.splitext(file_path)[1].lower() in SUPPORTED_INSTALL_EXTENSIONS

    def _collect_tracked_folders(self):
        tracked = set()
        for entry in self.installed_mods:
            folder = entry.get("target_folder")
            if not folder:
                continue
            normalized = os.path.normcase(os.path.abspath(folder))
            tracked.add(normalized)
        return tracked

    def _find_untracked_duplicates(self, file_path):
        if not self.mod_directory or not os.path.isdir(self.mod_directory):
            return []
        file_name = os.path.basename(file_path)
        if not file_name:
            return []
        tracked = self._collect_tracked_folders()
        duplicates = []
        for root, dirs, files in os.walk(self.mod_directory):
            for candidate in files:
                if candidate.lower() != file_name.lower():
                    continue
                candidate_path = os.path.join(root, candidate)
                try:
                    if os.path.samefile(candidate_path, file_path):
                        continue
                except OSError:
                    pass
                parent_dir = os.path.normcase(os.path.abspath(root))
                if parent_dir in tracked:
                    continue
                duplicates.append(os.path.abspath(candidate_path))
        return duplicates

    def install_mod_from_path(self, file_path):
        # Block installation while Sims 4 is running
        try:
            parent = self.parent()
            if parent is not None and hasattr(parent, "_is_sims_running") and parent._is_sims_running():
                return False, "Installation impossible: TS4_x64.exe est en cours d'exécution."
        except Exception:
            pass
        if not os.path.isfile(file_path):
            return False, f"Fichier introuvable : {file_path}"

        if not self._is_supported_extension(file_path):
            return False, f"Extension non supportée : {os.path.basename(file_path)}"

        if not self.mod_directory or not os.path.isdir(self.mod_directory):
            return False, "Définissez d'abord un dossier de mods valide dans la configuration."

        sanitized_name = sanitize_mod_folder_name(file_path)
        extension = os.path.splitext(file_path)[1].lower()
        zip_plan = None
        display_name = os.path.splitext(os.path.basename(file_path))[0]

        duplicates_to_replace: List[str] = []

        # New Mod Root installer logic for archives (ZIP/7z/rar)
        try:
            use_mod_root = bool(self.settings.get("installer_use_mod_root", True))
        except Exception:
            use_mod_root = True
        if use_mod_root and extension in {".zip", ".7z", ".rar"} and (mr_install_zip is not None):
            mods_root = self.mod_directory
            try:
                include_extras = bool(self.settings.get("installer_include_extras", False))
            except Exception:
                include_extras = False
            target_folder = None
            temp_dir = None
            try:
                if extension == ".zip":
                    try:
                        planned_dest, a_type, planned_root, just = mr_plan_zip(file_path, mods_root)
                    except Exception as exc:
                        return False, f"Impossible de préparer l'installation de l'archive: {exc}"
                    target_folder = planned_dest
                else:
                    # 7z/rar → extract to temp to plan
                    seven_zip = self._find_7z_executable()
                    if not seven_zip:
                        return False, "7-Zip (7z) est requis pour extraire ce format (7z/rar). Installez 7-Zip et ajoutez-le au PATH."
                    temp_dir = tempfile.mkdtemp(prefix="s4mt_mr_")
                    args = [seven_zip, 'x', '-y', f"-o{temp_dir}", file_path]
                    completed = subprocess.run(args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                    if completed.returncode != 0:
                        output = (completed.stderr or completed.stdout or "").strip()
                        shutil.rmtree(temp_dir, ignore_errors=True)
                        return False, f"Extraction 7z a échoué: {output}"
                    default_name = sanitize_mod_folder_name(os.path.splitext(os.path.basename(file_path))[0])
                    try:
                        planned_dest, a_type, planned_root, just = mr_plan_extracted_dir(temp_dir, mods_root, default_name)
                    except Exception as exc:
                        shutil.rmtree(temp_dir, ignore_errors=True)
                        return False, f"Impossible de préparer l'installation (extrait): {exc}"
                    target_folder = planned_dest

                # Let user adjust structure (destination name and include extras; for MIXED allow root override)
                override_root = None
                dest_name_override = None
                include_extras_op = include_extras
                try:
                    override_root, dest_name_override, include_extras_op = self._prompt_install_plan(
                        file_path,
                        extension,
                        temp_dir,
                        mods_root,
                        planned_root,
                        os.path.basename(target_folder.rstrip("/\\")),
                        include_extras,
                    )
                except Exception:
                    pass

                # Compute final target based on overrides
                final_target_folder = os.path.join(mods_root, dest_name_override) if dest_name_override else target_folder

                # Prompt if target exists (ATF guard + update confirmation)
                replace_existing = False
                if final_target_folder and os.path.exists(final_target_folder):
                    try:
                        existing_atf = False
                        for ent in self.installed_mods:
                            if os.path.normcase(os.path.abspath(ent.get("target_folder", ""))) == os.path.normcase(os.path.abspath(final_target_folder)):
                                existing_atf = bool(ent.get("atf", False))
                                break
                        if existing_atf:
                            resp = QtWidgets.QMessageBox.question(
                                self,
                "Confirmation Protected",
                ("Ce mod est marqué Protected.\n"
                                 "Confirmer que tu souhaites le mettre à jour ?"),
                                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                                QtWidgets.QMessageBox.No,
                            )
                            if resp != QtWidgets.QMessageBox.Yes:
                                return False, f"Mise à jour de '{os.path.basename(final_target_folder)}' annulée (ATF)."
                    except Exception:
                        pass
                    response = QtWidgets.QMessageBox.question(
                        self,
                        "Mod déjà installé",
                        (
                            f"Le mod '{os.path.basename(final_target_folder)}' existe déjà dans le dossier des mods.\n"
                            "Voulez-vous le mettre à jour avec le fichier sélectionné ?"
                        ),
                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                        QtWidgets.QMessageBox.Yes,
                    )
                    if response != QtWidgets.QMessageBox.Yes:
                        return False, f"Installation de '{os.path.basename(file_path)}' annulée."
                    replace_existing = True

                if replace_existing and final_target_folder:
                    try:
                        shutil.rmtree(final_target_folder)
                    except Exception:
                        pass

                # Perform installation
                try:
                    if extension == ".zip":
                        final_dest = mr_install_zip(
                            file_path,
                            mods_root,
                            include_extras=include_extras_op,
                            override_root=override_root,
                            dest_folder_name=dest_name_override,
                        )
                    else:
                        if temp_dir is None:
                            temp_dir = tempfile.mkdtemp(prefix="s4mt_mr_")
                            # Already extracted for planning in prior branch; if not, extract now
                            seven_zip = self._find_7z_executable()
                            if not seven_zip:
                                return False, "7-Zip (7z) est requis pour extraire ce format (7z/rar). Installez 7-Zip et ajoutez-le au PATH."
                            args = [seven_zip, 'x', '-y', f"-o{temp_dir}", file_path]
                            completed = subprocess.run(args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                            if completed.returncode != 0:
                                output = (completed.stderr or completed.stdout or "").strip()
                                shutil.rmtree(temp_dir, ignore_errors=True)
                                return False, f"Extraction 7z a échoué: {output}"
                        default_name = sanitize_mod_folder_name(os.path.splitext(os.path.basename(file_path))[0])
                        final_dest = mr_install_extracted_dir(
                            temp_dir,
                            mods_root,
                            default_name,
                            include_extras=include_extras_op,
                            override_root=override_root,
                            dest_folder_name=dest_name_override,
                        )
                finally:
                    if temp_dir is not None:
                        try:
                            shutil.rmtree(temp_dir, ignore_errors=True)
                        except Exception:
                            pass

                # Build installed entries list
                installed_entries = []
                try:
                    for p in Path(final_dest).rglob("*"):
                        if not p.is_file():
                            continue
                        ext = p.suffix.lower()
                        if include_extras or ext in {".package", ".ts4script"}:
                            rel = str(p.relative_to(final_dest)).replace("\\", "/")
                            installed_entries.append(rel)
                except Exception:
                    pass

                installed_at = datetime.utcnow().replace(microsecond=0).isoformat()
                display_name = os.path.basename(final_dest)
                entry = {
                    "name": display_name,
                    "type": self._describe_install_type([file_path]),
                    "installed_at": installed_at,
                    "target_folder": final_dest,
                    "source": os.path.basename(file_path),
                    "addons": [],
                    "files": list(installed_entries or []),
                }
                self._record_installation(entry)
                self._write_marker_file(final_dest, entry)
                try:
                    self._prompt_mod_metadata(entry)
                except Exception:
                    pass

                self.installations_performed = True
                return True, f"{os.path.basename(file_path)} installé dans '{os.path.basename(final_dest)}'."
            finally:
                pass

        if extension == ".zip":
            plan_result = build_zip_install_plan(
                file_path,
                mod_directory=self.mod_directory,
                default_mod_name=sanitized_name,
            )
            if not plan_result.success or plan_result.plan is None:
                return False, plan_result.message or "Impossible de préparer l'installation de l'archive."
            zip_plan = plan_result.plan
            sanitized_name = zip_plan.mod_folder_name
            target_folder = zip_plan.target_folder
            display_name = sanitized_name
        else:
            target_folder = os.path.join(self.mod_directory, sanitized_name)
            duplicates_to_replace = self._find_untracked_duplicates(file_path)
            if duplicates_to_replace:
                message_lines = [
                    "Le fichier existe déjà dans le dossier des mods en dehors du Mod Installer.",
                    "Chemins détectés :",
                ]
                message_lines.extend(duplicates_to_replace)
                message_lines.append("")
                message_lines.append("Souhaites-tu remplacer ces occurrences ?")
                response = QtWidgets.QMessageBox.question(
                    self,
                    "Fichier déjà présent",
                    "\n".join(message_lines),
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                    QtWidgets.QMessageBox.Yes,
                )
                if response != QtWidgets.QMessageBox.Yes:
                    duplicates_to_replace = []

        replace_existing = False
        if not duplicates_to_replace and os.path.exists(target_folder):
            # If target mod is ATF, request an extra confirmation
            try:
                existing_atf = False
                for ent in self.installed_mods:
                    if os.path.normcase(os.path.abspath(ent.get("target_folder", ""))) == os.path.normcase(os.path.abspath(target_folder)):
                        existing_atf = bool(ent.get("atf", False))
                        break
                if existing_atf:
                    resp = QtWidgets.QMessageBox.question(
                        self,
                "Confirmation Protected",
                ("Ce mod est marqué Protected.\n"
                         "Confirmer que tu souhaites le mettre à jour ?"),
                        QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                        QtWidgets.QMessageBox.No,
                    )
                    if resp != QtWidgets.QMessageBox.Yes:
                        return False, f"Mise à jour de '{display_name}' annulée (ATF)."
            except Exception:
                pass
            response = QtWidgets.QMessageBox.question(
                self,
                "Mod déjà installé",
                (
                    f"Le mod '{sanitized_name}' existe déjà dans le dossier des mods.\n"
                    "Voulez-vous le mettre à jour avec le fichier sélectionné ?"
                ),
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.Yes,
            )
            if response != QtWidgets.QMessageBox.Yes:
                return False, f"Installation de '{display_name}' annulée."
            replace_existing = True

        if duplicates_to_replace:
            parent_directories = []
            for duplicate_path in duplicates_to_replace:
                parent_dir = os.path.dirname(duplicate_path)
                if parent_dir not in parent_directories:
                    parent_directories.append(parent_dir)
            success_messages = []
            for parent_dir in parent_directories:
                success, install_message, installed_paths = self._install_file_to_target(
                    file_path,
                    parent_dir,
                    clean_before=False,
                    merge=True,
                    zip_plan=None,
                    skip_existing_prompt=True,
                )
                if not success:
                    return False, install_message
                success_messages.append(install_message)
                installed_at = datetime.utcnow().replace(microsecond=0).isoformat()
                entry = {
                    "name": display_name,
                    "type": self._describe_install_type([file_path]),
                    "installed_at": installed_at,
                    "target_folder": parent_dir,
                    "source": os.path.basename(file_path),
                    "addons": [],
                    "files": list(installed_paths or []),
                }
                self._record_installation(entry)
                self._write_marker_file(parent_dir, entry)
                # Prompt for optional metadata (version, URL)
                try:
                    self._prompt_mod_metadata(entry)
                except Exception:
                    pass
            self.installations_performed = True
            return True, "\n".join(success_messages)

        success, install_message, installed_paths = self._install_file_to_target(
            file_path,
            target_folder,
            clean_before=replace_existing,
            merge=False,
            zip_plan=zip_plan,
        )
        if not success:
            return False, install_message

        installed_at = datetime.utcnow().replace(microsecond=0).isoformat()
        entry = {
            "name": display_name,
            "type": self._describe_install_type([file_path]),
            "installed_at": installed_at,
            "target_folder": target_folder,
            "source": os.path.basename(file_path),
            "addons": [],
            "files": list(installed_paths or []),
        }
        self._record_installation(entry)
        self._write_marker_file(target_folder, entry)
        # Prompt for optional metadata (version, URL)
        try:
            self._prompt_mod_metadata(entry)
        except Exception:
            pass

        self.installations_performed = True
        final_message = install_message or f"'{display_name}' installé avec succès."
        return True, final_message

    def _install_file_to_target(
        self,
        file_path,
        target_folder,
        *,
        clean_before=False,
        merge=False,
        zip_plan=None,
        skip_existing_prompt=False,
    ):
        extension = os.path.splitext(file_path)[1].lower()
        installed_entries = []
        plan_warnings: List[str] = []

        # AI: warn on reinstall or recently disabled copy
        if not clean_before:
            try:
                marker_path = os.path.join(target_folder, MOD_MARKER_FILENAME)
                has_marker = os.path.isfile(marker_path)
            except Exception:
                has_marker = False
            disabled_hint = ""
            try:
                name_guess = os.path.basename(target_folder.rstrip("/\\"))
                for ent in load_installed_mods():
                    if bool(ent.get("disabled")) and (ent.get("name") == name_guess):
                        disabled_hint = ent.get("disabled_path") or ""
                        break
            except Exception:
                pass
            if has_marker or disabled_hint:
                extra = (f"\nUne version désactivée existe: {disabled_hint}" if disabled_hint else "")
                resp = QtWidgets.QMessageBox.question(
                    self,
                    "Réinstallation détectée",
                    (
                        "Un mod semble déjà installé dans ce dossier.\n"
                        "Souhaitez-vous remplacer entièrement le contenu (clean) ?\n"
                        "Oui = remplacer (clean) • Non = fusionner (garder les fichiers existants)."
                        f"{extra}"
                    ),
                    QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No | QtWidgets.QMessageBox.Cancel,
                    QtWidgets.QMessageBox.Yes,
                )
                if resp == QtWidgets.QMessageBox.Cancel:
                    return False, "Installation annulée par l'utilisateur.", []
                if resp == QtWidgets.QMessageBox.Yes:
                    clean_before = True

        if clean_before and os.path.exists(target_folder):
            try:
                shutil.rmtree(target_folder)
            except OSError as exc:
                return False, f"Impossible de nettoyer le dossier existant : {exc}", []

        if extension == ".zip" and not zipfile.is_zipfile(file_path):
            return False, f"Le fichier n'est pas une archive zip valide : {os.path.basename(file_path)}", []

        try:
            os.makedirs(target_folder, exist_ok=True)
        except OSError as exc:
            return False, f"Impossible de créer le dossier cible : {exc}", []

        try:
            if extension in {".package", ".ts4script"}:
                destination_path = os.path.join(target_folder, os.path.basename(file_path))
                if os.path.exists(destination_path) and not clean_before:
                    if skip_existing_prompt:
                        response = QtWidgets.QMessageBox.Yes
                    else:
                        response = QtWidgets.QMessageBox.question(
                            self,
                            "Fichier déjà présent",
                            (
                                f"Le fichier '{os.path.basename(file_path)}' existe déjà dans le dossier cible.\n"
                                "Souhaitez-vous le remplacer ?"
                            ),
                            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                            QtWidgets.QMessageBox.Yes,
                        )
                    if response != QtWidgets.QMessageBox.Yes:
                        return False, f"Copie de '{os.path.basename(file_path)}' annulée.", []
                shutil.copy2(file_path, destination_path)
                installed_entries.append(os.path.basename(destination_path))
            elif extension == ".zip":
                target_root = os.path.abspath(target_folder)
                if zip_plan:
                    plan_warnings.extend(zip_plan.warnings)
                with zipfile.ZipFile(file_path, "r") as archive:
                    if zip_plan and zip_plan.entries:
                        for entry in zip_plan.entries:
                            try:
                                info = archive.getinfo(entry.member_name)
                            except KeyError:
                                plan_warnings.append(f"Entrée introuvable dans l'archive: {entry.member_name}")
                                continue
                            if info.is_dir():
                                continue
                            dest_parts = [part for part in entry.relative_parts if part]
                            if not dest_parts:
                                continue
                            dest_path = os.path.abspath(os.path.join(target_root, *dest_parts))
                            if os.path.commonpath([target_root, dest_path]) != target_root:
                                plan_warnings.append(f"Chemin invalide ignoré: {'/'.join(dest_parts)}")
                                continue
                            os.makedirs(os.path.dirname(dest_path), exist_ok=True)
                            try:
                                with archive.open(info, "r") as source, open(dest_path, "wb") as target_file:
                                    shutil.copyfileobj(source, target_file)
                                try:
                                    dt = datetime(*info.date_time)
                                    os.utime(dest_path, (dt.timestamp(), dt.timestamp()))
                                except Exception:
                                    pass
                                rel_display = "/".join(dest_parts)
                                installed_entries.append(rel_display)
                            except OSError as exc:
                                plan_warnings.append(f"Écriture impossible: {'/'.join(dest_parts)} → {exc}")
                    else:
                        for info in archive.infolist():
                            member_name = str(info.filename)
                            norm = member_name.replace("\\", "/").lstrip("/").strip()
                            if not norm or norm.endswith("/"):
                                continue  # directory entries handled implicitly
                            parts = [p for p in norm.split("/") if p and p not in {".", ".."}]
                            if not parts:
                                continue
                            if _member_should_be_skipped(parts, parts[0]):
                                continue
                            _, ext = os.path.splitext(parts[-1])
                            if ext.lower() in DISALLOWED_ARCHIVE_EXTENSIONS:
                                plan_warnings.append(f"Fichier ignoré pour sécurité: {norm}")
                                continue
                            dest_path = os.path.abspath(os.path.join(target_root, *parts))
                            if os.path.commonpath([target_root, dest_path]) != target_root:
                                continue
                            os.makedirs(os.path.dirname(dest_path), exist_ok=True)
                            try:
                                with archive.open(info, "r") as source, open(dest_path, "wb") as target_file:
                                    shutil.copyfileobj(source, target_file)
                                try:
                                    dt = datetime(*info.date_time)
                                    os.utime(dest_path, (dt.timestamp(), dt.timestamp()))
                                except Exception:
                                    pass
                                rel_display = "/".join(parts)
                                installed_entries.append(rel_display)
                            except OSError as exc:
                                plan_warnings.append(f"Écriture impossible: {norm} → {exc}")

                if not installed_entries:
                    return False, "L'archive ne contient aucun fichier exploitable.", []
            else:
                # Extract via 7-Zip (7z/rar/others) into temp, normalize, then copy to target
                seven_zip = self._find_7z_executable()
                if not seven_zip:
                    return False, "7-Zip (7z) est requis pour extraire ce format (7z/rar). Installez 7-Zip et ajoutez-le au PATH.", []
                temp_dir = tempfile.mkdtemp(prefix="s4mt_")
                try:
                    args = [seven_zip, 'x', '-y', f"-o{temp_dir}", file_path]
                    try:
                        completed = subprocess.run(args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
                    except OSError as exc:
                        return False, f"Échec de l'extraction via 7z: {exc}", []
                    if completed.returncode != 0:
                        output = (completed.stderr or completed.stdout or "").strip()
                        return False, f"Extraction 7z a échoué: {output}", []

                    default_mod_name = sanitize_mod_folder_name(os.path.splitext(os.path.basename(file_path))[0])
                    plan_result = build_extracted_install_plan(
                        temp_dir,
                        mod_directory=self.mod_directory,
                        default_mod_name=default_mod_name,
                        existing_target=target_folder,
                    )
                    if not plan_result.success or plan_result.plan is None:
                        return False, plan_result.message or "Impossible de préparer l'installation du contenu extrait.", []
                    plan = plan_result.plan
                    plan_warnings.extend(plan.warnings)
                    target_root = os.path.abspath(plan.target_folder)
                    os.makedirs(target_root, exist_ok=True)
                    for entry in plan.entries:
                        # entry.member_name is the relative path inside temp_dir
                        member_rel = entry.member_name.replace("\\", "/").lstrip("/")
                        source_path = os.path.abspath(os.path.join(temp_dir, *member_rel.split("/")))
                        dest_parts = [part for part in entry.relative_parts if part]
                        if not dest_parts:
                            continue
                        dest_path = os.path.abspath(os.path.join(target_root, *dest_parts))
                        if os.path.commonpath([target_root, dest_path]) != target_root:
                            continue
                        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
                        try:
                            shutil.copy2(source_path, dest_path)
                            installed_entries.append("/".join(dest_parts))
                        except OSError as exc:
                            plan_warnings.append(f"Écriture impossible: {'/'.join(dest_parts)} → {exc}")
                finally:
                    try:
                        shutil.rmtree(temp_dir, ignore_errors=True)
                    except Exception:
                        pass
        except (OSError, zipfile.BadZipFile, RuntimeError) as exc:
            return False, f"Erreur lors de la copie : {exc}", []

        verb = "ajouté" if merge and not clean_before else "installé"
        message = f"{os.path.basename(file_path)} {verb} dans '{os.path.basename(target_folder)}'."
        if plan_warnings:
            message = message + "\n" + "\n".join(plan_warnings)
        return True, message, installed_entries

    def _find_7z_executable(self):
        candidates = [
            "7z",
            os.path.join(os.environ.get("ProgramFiles", r"C:\\Program Files"), "7-Zip", "7z.exe"),
            os.path.join(os.environ.get("ProgramFiles(x86)", r"C:\\Program Files (x86)"), "7-Zip", "7z.exe"),
        ]
        for path in candidates:
            if not path:
                continue
            try:
                completed = subprocess.run([path, "-h"], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                if completed.returncode == 0 or completed.stdout or completed.stderr:
                    return path
            except Exception:
                continue
        return None

    def _snapshot_relative_files(self, target_folder):
        root = os.path.abspath(target_folder)
        rels = set()
        for cur, _dirs, files in os.walk(root):
            for f in files:
                rel = os.path.relpath(os.path.join(cur, f), root).replace("\\", "/")
                rels.add(rel)
        return rels

    @staticmethod
    def _describe_install_type(file_paths):
        extensions = []
        for path in file_paths:
            extension = os.path.splitext(path)[1].lower()
            if extension in {".package", ".ts4script"}:
                extensions.append(f"fichier {extension}")
            elif extension in {".zip", ".7z", ".rar"}:
                extensions.append(f"archive {extension}")
        if not extensions:
            return ""
        if len(set(extensions)) == 1:
            return extensions[0]
        return ", ".join(sorted(set(extensions)))

    def _record_installation(self, entry):
        normalized_entry = dict(entry)
        normalized_entry["addons"] = normalize_addon_metadata(entry.get("addons", []))
        # Normalize and store file list if provided
        files = []
        for p in normalized_entry.get("files", []) or []:
            s = str(p).replace("\\", "/").strip()
            if s and s not in files:
                files.append(s)
        if files:
            normalized_entry["files"] = files

        target = normalized_entry.get("target_folder")
        if not target:
            return
        replaced = False
        for existing in self.installed_mods:
            if existing.get("target_folder") == target:
                # Merge dict but union addons and files lists
                for k, v in normalized_entry.items():
                    if k in {"addons", "files"}:
                        continue
                    existing[k] = v
                existing["addons"] = normalize_addon_metadata(existing.get("addons", []))
                merged_files = []
                for lst in (existing.get("files", []) or [], normalized_entry.get("files", []) or []):
                    for p in lst:
                        if p and p not in merged_files:
                            merged_files.append(p)
                if merged_files:
                    existing["files"] = merged_files
                replaced = True
                break
        if not replaced:
            normalized_entry.setdefault("addons", [])
            normalized_entry.setdefault("files", [])
            self.installed_mods.append(normalized_entry)
        self.installed_mods.sort(key=lambda item: item.get("installed_at", ""), reverse=True)
        save_installed_mods(self.installed_mods)
        # Apply persisted hidden columns and populate
        self._apply_installer_hidden_columns()
        self.refresh_table()

    def _open_installer_settings(self):
        dlg = QtWidgets.QDialog(self)
        try:
            dlg.setWindowFlags(dlg.windowFlags() | QtCore.Qt.Window)
            dlg.setSizeGripEnabled(True)
        except Exception:
            pass
        dlg.setWindowTitle("Mod Installer – Settings")
        dlg.setModal(True)
        layout = QtWidgets.QVBoxLayout(dlg)
        use_mod_root_cb = QtWidgets.QCheckBox("Utiliser la logique Mod Root (ZIP/7z/rar)", dlg)
        use_mod_root_cb.setChecked(bool(self.settings.get("installer_use_mod_root", True)))
        layout.addWidget(use_mod_root_cb)
        include_extras_cb = QtWidgets.QCheckBox("Inclure fichiers non essentiels (docs/images)", dlg)
        include_extras_cb.setChecked(bool(self.settings.get("installer_include_extras", False)))
        layout.addWidget(include_extras_cb)
        layout.addStretch(1)
        btns = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dlg)
        layout.addWidget(btns)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            self.settings["installer_use_mod_root"] = bool(use_mod_root_cb.isChecked())
            self.settings["installer_include_extras"] = bool(include_extras_cb.isChecked())
            # Persist to parent if possible
            parent = self.parent()
            if parent is not None and hasattr(parent, "settings"):
                parent.settings["installer_use_mod_root"] = self.settings["installer_use_mod_root"]
                parent.settings["installer_include_extras"] = self.settings["installer_include_extras"]
                save_settings(parent.settings)
            else:
                save_settings(self.settings)

    def _prompt_mod_metadata(self, entry):
        # Ask for optional version and URL; allow empty values
        dlg = QtWidgets.QDialog(self)
        try:
            dlg.setWindowFlags(dlg.windowFlags() | QtCore.Qt.Window)
            dlg.setSizeGripEnabled(True)
        except Exception:
            pass
        dlg.setWindowTitle("Informations du mod (optionnel)")
        dlg.setModal(True)
        try:
            dlg.setSizeGripEnabled(True)
        except Exception:
            pass
        form = QtWidgets.QFormLayout(dlg)
        version_edit = QtWidgets.QLineEdit(dlg)
        version_edit.setPlaceholderText("1.118.257.1020 (facultatif)")
        # Prefill from filename suffix (_vX.Y.Z) or fallback to latest game version <= file date
        prefill = entry.get("mod_version", "")
        try:
            # Try explicit in installed files or source
            candidates = list(entry.get("files") or [])
            src = entry.get("source", "")
            if src:
                candidates.append(src)
            for name in candidates:
                v = extract_version_from_name(name)
                if v:
                    prefill = v
                    break
            if not prefill:
                # Guess from file modified time
                paths = []
                target_folder = entry.get("target_folder") or ""
                for rel in entry.get("files") or []:
                    try:
                        ap = os.path.abspath(os.path.join(target_folder, rel))
                        if os.path.isfile(ap):
                            paths.append(ap)
                    except Exception:
                        pass
                if not paths and target_folder and os.path.isdir(target_folder):
                    # fallback to any file under target (best-effort)
                    for cur, _d, files in os.walk(target_folder):
                        for f in files:
                            paths.append(os.path.join(cur, f))
                dt = None
                for p in paths:
                    try:
                        ts = os.path.getmtime(p)
                        if not dt or ts > dt:
                            dt = ts
                    except Exception:
                        pass
                if dt:
                    latest_date = datetime.fromtimestamp(dt).date()
                    releases = {}
                    try:
                        parent = self.parent()
                        releases = getattr(parent, "version_releases", {}) if parent else {}
                    except Exception:
                        releases = {}
                    # pick the latest version with release_date <= latest_date
                    chosen = ""
                    try:
                        for ver, d in releases.items():
                            if d and isinstance(d, date) and d <= latest_date:
                                if not chosen:
                                    chosen = ver
                                else:
                                    # releases dict seems sorted asc already; keep iterating
                                    chosen = ver
                    except Exception:
                        pass
                    if chosen:
                        prefill = chosen
        except Exception:
            pass
        version_edit.setText(prefill)
        url_edit = QtWidgets.QLineEdit(dlg)
        url_edit.setPlaceholderText("https://... (facultatif)")
        url_edit.setText(entry.get("url", ""))
        form.addRow("Numéro de version :", version_edit)
        form.addRow("URL du mod :", url_edit)
        buttons = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel, dlg)
        form.addWidget(buttons)
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        if dlg.exec_() == QtWidgets.QDialog.Accepted:
            updated = dict(entry)
            updated["mod_version"] = version_edit.text().strip()
            updated["url"] = url_edit.text().strip()
            self._record_installation(updated)
            self._write_marker_file(updated.get("target_folder"), updated)

    def _write_marker_file(self, target_folder, entry):
        try:
            marker_path = os.path.join(target_folder, MOD_MARKER_FILENAME)
            data = {
                "name": entry.get("name", ""),
                "type": entry.get("type", ""),
                "installed_at": entry.get("installed_at", ""),
                "source": entry.get("source", ""),
                "app_version": APP_VERSION,
                "app_version_date": APP_VERSION_DATE,
                "files": list(entry.get("files", []) or []),
                "mod_version": entry.get("mod_version", ""),
                "url": entry.get("url", ""),
                "atf": bool(entry.get("atf", False)),
            }
            with open(marker_path, "w", encoding="utf-8") as fh:
                json.dump(data, fh, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _recover_from_markers(self):
        if not self.mod_directory or not os.path.isdir(self.mod_directory):
            QtWidgets.QMessageBox.warning(self, "Dossier des mods invalide", "Définis un dossier des mods valide dans la configuration.")
            return
        recovered = 0
        errors = []
        seen_targets = set()
        try:
            for root, dirs, files in os.walk(self.mod_directory):
                if MOD_MARKER_FILENAME in files:
                    marker_path = os.path.join(root, MOD_MARKER_FILENAME)
                    try:
                        with open(marker_path, "r", encoding="utf-8") as fh:
                            data = json.load(fh)
                        entry = {
                            "name": str(data.get("name") or os.path.basename(root)),
                            "type": str(data.get("type") or ""),
                            "installed_at": str(data.get("installed_at") or ""),
                            "target_folder": root,
                            "source": str(data.get("source") or ""),
                            "addons": [],
                            "files": list(data.get("files", []) or []),
                            "mod_version": str(data.get("mod_version") or ""),
                            "url": str(data.get("url") or ""),
                            "atf": bool(data.get("atf", False)),
                        }
                        target_key = os.path.normcase(os.path.abspath(root))
                        if target_key in seen_targets:
                            continue
                        seen_targets.add(target_key)
                        self._record_installation(entry)
                        recovered += 1
                    except (OSError, json.JSONDecodeError) as exc:
                        errors.append(f"{marker_path} → {exc}")
        finally:
            self.refresh_table()
        message = [f"{recovered} mod(s) récupéré(s) depuis les marqueurs."]
        if errors:
            message.append("Erreurs:\n" + "\n".join(errors))
        QtWidgets.QMessageBox.information(self, "Recovery list", "\n".join(message))

    def _show_context_menu(self, position):
        item = self.table.itemAt(position)
        if item is None:
            return
        row = item.row()
        if row < 0 or row >= len(self.installed_mods):
            return
        entry = self.installed_mods[row]

        menu = QtWidgets.QMenu(self)
        search_action = menu.addAction("Recherche Google")
        patreon_action = menu.addAction("Chercher sur Patreon")
        search_update_action = menu.addAction("Search for Update")
        repair_marker_action = menu.addAction("Repair definition…")
        menu.addSeparator()
        open_action = menu.addAction("Ouvrir dans l'explorateur")
        rename_action = menu.addAction("Renommer le mod")
        set_version_action = menu.addAction("Définir version…")
        set_url_action = menu.addAction("Définir URL…")
        addons_action = menu.addAction("Ajouter add-ons")
        remove_addons_action = menu.addAction("Supprimer add-ons")
        remove_addons_action.setEnabled(bool(entry.get("addons")))
        atf_action = menu.addAction("Retirer Protected" if entry.get("atf") else "Marquer Protected")
        disable_action = None
        if entry.get("disabled"):
            disable_action = menu.addAction("Réactiver le mod")
        else:
            disable_action = menu.addAction("Désactiver le mod")
        delete_action = menu.addAction("Supprimer le mod")
        update_action = menu.addAction("Mettre à jour le mod")

        chosen_action = menu.exec_(self.table.viewport().mapToGlobal(position))
        if chosen_action is None:
            return
        if chosen_action == search_action:
            self._open_google_search(entry)
        elif chosen_action == patreon_action:
            self._open_patreon_search(entry)
        elif chosen_action == search_update_action:
            self._search_for_update(entry)
        elif chosen_action == repair_marker_action:
            self._repair_definition(entry)
        elif chosen_action == open_action:
            self._open_in_file_manager(entry.get("target_folder"))
        elif chosen_action == rename_action:
            self._rename_mod(entry)
        elif chosen_action == set_version_action:
            self._set_mod_version(entry)
        elif chosen_action == set_url_action:
            self._set_mod_url(entry)
        elif chosen_action == addons_action:
            self._prompt_addons(entry)
        elif chosen_action == remove_addons_action:
            self._prompt_remove_addons(entry)
        elif chosen_action == atf_action:
            entry = dict(entry)
            entry["atf"] = not bool(entry.get("atf", False))
            self._record_installation(entry)
            self._write_marker_file(entry.get("target_folder"), entry)
        elif chosen_action == disable_action:
            self._toggle_disable_mod(entry)
        elif chosen_action == delete_action:
            self._delete_mod(entry)
        elif chosen_action == update_action:
            self._prompt_update_mod(entry)

    def _show_installer_header_menu(self, pos):
        header = self.table.horizontalHeader()
        global_pos = header.mapToGlobal(pos)
        menu = QtWidgets.QMenu(self)
        labels = list(self._installer_headers)
        parent = self.parent()
        settings = getattr(parent, "settings", {}) if parent else {}
        hidden = set(int(c) for c in settings.get("installer_hidden_columns", []))
        for col, label in enumerate(labels):
            action = QtWidgets.QAction(label, menu)
            action.setCheckable(True)
            action.setChecked(col not in hidden)
            action.triggered.connect(partial(self._toggle_installer_column_visibility, col))
            menu.addAction(action)
        menu.exec_(global_pos)

    def _toggle_installer_column_visibility(self, col, checked):
        parent = self.parent()
        if not parent or not hasattr(parent, "settings"):
            self.table.setColumnHidden(col, not checked)
            return
        settings = parent.settings
        hidden = set(int(c) for c in settings.get("installer_hidden_columns", []))
        if checked and col in hidden:
            hidden.remove(col)
        elif not checked:
            hidden.add(col)
        settings["installer_hidden_columns"] = sorted(int(c) for c in hidden)
        save_settings(settings)
        self.table.setColumnHidden(col, not checked)

    def _apply_installer_hidden_columns(self):
        parent = self.parent()
        hidden = set()
        if parent and hasattr(parent, "settings"):
            hidden = set(int(c) for c in parent.settings.get("installer_hidden_columns", []))
        for col in range(self.table.columnCount()):
            try:
                self.table.setColumnHidden(col, col in hidden)
            except Exception:
                pass

    def _on_header_section_clicked(self, section):
        if section < 0 or section >= len(self._installer_headers):
            return
        base_label = self._installer_headers[section]
        current_filter = self.column_filters.get(section, "")
        try:
            text, ok = QtWidgets.QInputDialog.getText(
                self,
                "Filtrer la colonne",
                (
                    f"Saisis un filtre pour la colonne '{base_label}'.\n"
                    "Laisse vide pour supprimer le filtre."
                ),
                QtWidgets.QLineEdit.Normal,
                current_filter,
            )
        except Exception:
            return
        if not ok:
            return
        value = text.strip()
        if value:
            self.column_filters[section] = value
        else:
            self.column_filters.pop(section, None)
        self._update_header_labels()
        self._apply_table_filters()

    def _update_header_labels(self):
        for index, label in enumerate(self._installer_headers):
            header_item = self.table.horizontalHeaderItem(index)
            if not header_item:
                continue
            if self.column_filters.get(index):
                header_item.setText(f"{label} (filtre)")
            else:
                header_item.setText(label)

    def _apply_table_filters(self):
        if not self.column_filters:
            for row in range(self.table.rowCount()):
                self.table.setRowHidden(row, False)
            return
        active_filters = {col: value.casefold() for col, value in self.column_filters.items() if value}
        for row in range(self.table.rowCount()):
            hide_row = False
            for column, needle in active_filters.items():
                item = self.table.item(row, column)
                cell_value = item.text().casefold() if item else ""
                if needle not in cell_value:
                    hide_row = True
                    break
            self.table.setRowHidden(row, hide_row)

    def _set_mod_version(self, entry):
        current = entry.get("mod_version", "")
        text, ok = QtWidgets.QInputDialog.getText(self, "Définir version", "Numéro de version :", QtWidgets.QLineEdit.Normal, current)
        if not ok:
            return
        entry["mod_version"] = text.strip()
        self._record_installation(entry)

    def _set_mod_url(self, entry):
        current = entry.get("url", "")
        text, ok = QtWidgets.QInputDialog.getText(self, "Définir URL", "URL du mod :", QtWidgets.QLineEdit.Normal, current)
        if not ok:
            return
        entry["url"] = text.strip()
        self._record_installation(entry)

    def _rename_mod(self, entry):
        old_name = entry.get("name", "")
        target_folder = entry.get("target_folder", "")
        new_name, ok = QtWidgets.QInputDialog.getText(self, "Renommer le mod", "Nouveau nom :", QtWidgets.QLineEdit.Normal, old_name)
        if not ok:
            return
        new_name = new_name.strip()
        if not new_name:
            return
        # Compute new folder path
        parent = os.path.dirname(target_folder)
        new_folder = os.path.join(parent, sanitize_mod_folder_name(new_name))
        if os.path.abspath(new_folder) == os.path.abspath(target_folder):
            entry["name"] = new_name
            self._record_installation(entry)
            return
        if os.path.exists(new_folder):
            QtWidgets.QMessageBox.warning(self, "Impossible de renommer", "Un dossier portant ce nom existe déjà.")
            return
        try:
            os.rename(target_folder, new_folder)
        except OSError as exc:
            QtWidgets.QMessageBox.critical(self, "Erreur", f"Impossible de renommer le dossier : {exc}")
            return
        entry["name"] = new_name
        entry["target_folder"] = new_folder
        self._record_installation(entry)

    def _toggle_disable_mod(self, entry):
        backups_dir = self.parent().settings.get("backups_directory", "") if self.parent() else ""
        if not backups_dir:
            QtWidgets.QMessageBox.warning(self, "Backups manquant", "Définis un dossier de backups dans la configuration.")
            return
        disabled_root = os.path.join(backups_dir, "Disabled Mod")
        os.makedirs(disabled_root, exist_ok=True)
        target_folder = entry.get("target_folder", "")
        if entry.get("disabled"):
            # Reactivate
            disabled_path = entry.get("disabled_path") or ""
            if not disabled_path or not os.path.exists(disabled_path):
                QtWidgets.QMessageBox.warning(self, "Réactivation impossible", "Le dossier désactivé est introuvable.")
                return
            destination = target_folder
            final = destination
            i = 1
            while os.path.exists(final):
                final = f"{destination}_{i}"
                i += 1
            try:
                shutil.move(disabled_path, final)
            except OSError as exc:
                QtWidgets.QMessageBox.critical(self, "Erreur", f"Impossible de réactiver : {exc}")
                return
            entry["disabled"] = False
            entry["disabled_path"] = ""
            entry["target_folder"] = final
            self._record_installation(entry)
            QtWidgets.QMessageBox.information(self, "Mod réactivé", "Le mod a été réactivé.")
        else:
            # Disable (move to backups)
            if not target_folder or not os.path.isdir(target_folder):
                QtWidgets.QMessageBox.warning(self, "Dossier introuvable", "Le dossier du mod est introuvable.")
                return
            dest = os.path.join(disabled_root, os.path.basename(target_folder))
            final = dest
            i = 1
            while os.path.exists(final):
                final = f"{dest}_{i}"
                i += 1
            try:
                shutil.move(target_folder, final)
            except OSError as exc:
                QtWidgets.QMessageBox.critical(self, "Erreur", f"Impossible de désactiver : {exc}")
                return
            entry["disabled"] = True
            entry["disabled_path"] = final
            # Keep original target_folder to restore later
            self._record_installation(entry)
            QtWidgets.QMessageBox.information(self, "Mod désactivé", "Le mod a été déplacé dans Backups/Disabled Mod.")

    def _on_installer_item_changed(self, item):
        try:
            row = item.row()
            col = item.column()
            if row < 0 or row >= len(self.installed_mods):
                return
            entry = dict(self.installed_mods[row])
            text = item.text().strip()
            if col == 3:  # Version
                entry["mod_version"] = text
                self._record_installation(entry)
            elif col == 4:  # URL
                entry["url"] = text
                self._record_installation(entry)
        except Exception:
            pass

    def _open_in_file_manager(self, target_path):
        if not target_path:
            return
        if not os.path.exists(target_path):
            QtWidgets.QMessageBox.warning(
                self,
                "Dossier introuvable",
                "Le dossier du mod est introuvable. Vérifiez qu'il n'a pas été supprimé.",
            )
            return

        if os.path.isfile(target_path):
            target_path = os.path.dirname(target_path) or target_path

        if sys.platform.startswith("win"):
            try:
                os.startfile(target_path)
            except OSError:
                QtWidgets.QMessageBox.warning(
                    self,
                    "Erreur",
                    "Impossible d'ouvrir l'explorateur de fichiers.",
                )
        elif sys.platform == "darwin":
            QtCore.QProcess.startDetached("open", [target_path])
        else:
            QtCore.QProcess.startDetached("xdg-open", [target_path])

    def _open_google_search(self, entry):
        mod_name = entry.get("name") or os.path.basename(entry.get("target_folder", ""))
        if not mod_name:
            return
        query = quote_plus(mod_name)
        webbrowser.open(f"https://www.google.com/search?q={query}")

    def _open_patreon_search(self, entry):
        mod_name = entry.get("name") or os.path.basename(entry.get("target_folder", ""))
        if not mod_name:
            return
        q = quote_plus(f"site:patreon.com {mod_name}")
        webbrowser.open(f"https://www.google.com/search?q={q}")

    def _prompt_update_mod(self, entry):
        target_folder = entry.get("target_folder")
        if not target_folder or not os.path.isdir(target_folder):
            QtWidgets.QMessageBox.warning(
                self,
                "Dossier introuvable",
                "Le dossier du mod est introuvable. Vérifiez qu'il n'a pas été supprimé.",
            )
            return

        def handle_drop(file_paths):
            return self._perform_update(entry, file_paths)

        instruction = (
            "Glissez-déposez un fichier .package, .ts4script ou .zip pour remplacer le contenu du dossier du mod."
        )
        dialog = FileDropDialog("Mettre à jour le mod", instruction, handle_drop, self)
        dialog.exec_()

    def _search_for_update(self, entry):
        # Ouvre l'Updates Checker, filtre par nom et tente une correspondance
        name = entry.get("name") or os.path.basename(entry.get("target_folder", ""))
        if not name:
            return
        parent = self.parent()
        try:
            dlg = UpdatesCheckerDialog(parent if parent is not None else self)
        except Exception as exc:
            QtWidgets.QMessageBox.warning(self, "Updates Checker", str(exc))
            return
        dlg.show()
        QtWidgets.QApplication.processEvents()
        try:
            dlg.search_edit.setText(name)
        except Exception:
            pass
        # Find the row and run Try
        try:
            for r in range(dlg.table.rowCount()):
                if dlg.table.isRowHidden(r):
                    continue
                n = dlg.table.item(r, 1).text() if dlg.table.item(r, 1) else ""
                if n.lower() == name.lower():
                    dlg._try_fetch(name, r)
                    try:
                        dlg.table.selectRow(r)
                    except Exception:
                        pass
                    break
        except Exception:
            pass

    def _repair_definition(self, entry):
        # Permet de re-vérifier un dossier et de recréer un marker propre
        start_dir = entry.get("target_folder", "") or self.mod_directory
        folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Choisir le dossier du mod", start_dir)
        if not folder:
            return
        if not os.path.isdir(folder):
            QtWidgets.QMessageBox.warning(self, "Dossier invalide", "Sélectionne un dossier existant.")
            return
        # Collecte fichiers pertinents
        rel_files = []
        for cur, _dirs, files in os.walk(folder):
            for f in files:
                ext = os.path.splitext(f)[1].lower()
                if ext not in {'.package', '.ts4script'}:
                    continue
                full = os.path.join(cur, f)
                rel = os.path.relpath(full, folder).replace('\\', '/')
                rel_files.append(rel)
        rel_files.sort()
        # Reconstruit l'entrée
        new_entry = dict(entry)
        new_entry["name"] = os.path.basename(folder)
        new_entry["target_folder"] = folder
        # décrit le type à partir du contenu
        types = set(os.path.splitext(f)[1].lower() for f in rel_files)
        if not types:
            mod_type = ""
        elif types == {'.package'}:
            mod_type = "fichier .package"
        elif types == {'.ts4script'}:
            mod_type = "fichier .ts4script"
        else:
            mod_type = ", ".join(sorted({
                ("fichier .package" if t == '.package' else "fichier .ts4script") for t in types
            }))
        new_entry["type"] = mod_type
        new_entry["files"] = rel_files
        # Écrit marker et met à jour la table
        self._write_marker_file(folder, new_entry)
        self._record_installation(new_entry)
        QtWidgets.QMessageBox.information(self, "Repair", "Le fichier .s4mt_mod_marker.json a été reconstruit.")

    def _prompt_addons(self, entry):
        target_folder = entry.get("target_folder")
        if not target_folder or not os.path.isdir(target_folder):
            QtWidgets.QMessageBox.warning(
                self,
                "Dossier introuvable",
                "Le dossier du mod est introuvable. Vérifiez qu'il n'a pas été supprimé.",
            )
            return

        def handle_drop(file_paths):
            return self._perform_addons(entry, file_paths)

        folder_display = target_folder
        instruction = (
            f"Dossier du mod maître :\n{folder_display}\n\n"
            "Glissez-déposez des fichiers .package, .ts4script, .zip, .7z ou .rar pour les ajouter au dossier du mod."
        )
        title = f"Ajouter des add-ons – {os.path.basename(target_folder)}"
        dialog = FileDropDialog(title, instruction, handle_drop, self)
        dialog.exec_()

    def _prompt_remove_addons(self, entry):
        entry_addons = normalize_addon_metadata(entry.get("addons", []))
        if not entry_addons:
            QtWidgets.QMessageBox.information(
                self,
                "Aucun add-on",
                "Aucun add-on n'est enregistré pour ce mod.",
            )
            return

        entry["addons"] = entry_addons

        dialog = QtWidgets.QDialog(self)
        try:
            dialog.setWindowFlags(dialog.windowFlags() | QtCore.Qt.Window)
            dialog.setSizeGripEnabled(True)
        except Exception:
            pass
        dialog.setWindowTitle("Supprimer des add-ons")
        dialog.setModal(True)

        layout = QtWidgets.QVBoxLayout(dialog)
        info_label = QtWidgets.QLabel("Sélectionnez les add-ons à supprimer :")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)

        list_widget = QtWidgets.QListWidget(dialog)
        list_widget.setSelectionMode(QtWidgets.QAbstractItemView.MultiSelection)
        for addon in entry_addons:
            label = addon.get("label") or "Add-on"
            added_at_display = format_installation_display(addon.get("added_at", ""))
            if added_at_display:
                item_text = f"{label} – {added_at_display}"
            else:
                item_text = label
            list_item = QtWidgets.QListWidgetItem(item_text)
            list_widget.addItem(list_item)
        layout.addWidget(list_widget)

        buttons = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
            QtCore.Qt.Horizontal,
            dialog,
        )
        remove_button = buttons.button(QtWidgets.QDialogButtonBox.Ok)
        remove_button.setText("Supprimer")
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)

        if dialog.exec_() != QtWidgets.QDialog.Accepted:
            return

        selected_rows = sorted({index.row() for index in list_widget.selectedIndexes()}, reverse=True)
        if not selected_rows:
            QtWidgets.QMessageBox.information(
                self,
                "Aucune sélection",
                "Sélectionnez au moins un add-on à supprimer.",
            )
            return

        success_messages = []
        error_messages = []
        removed_count = 0

        for row in selected_rows:
            if row < 0 or row >= len(entry_addons):
                continue
            addon = entry_addons[row]
            label = addon.get("label") or f"Add-on {row + 1}"
            removed, missing, errors = self._remove_addon_files(entry, addon)
            if errors:
                for message in errors:
                    error_messages.append(f"{label} : {message}")
                continue

            del entry_addons[row]
            removed_count += 1
            details = []
            if removed:
                details.append(f"{len(removed)} élément(s) supprimé(s)")
            if missing:
                details.append(f"{len(missing)} élément(s) introuvable(s)")
            if details:
                success_messages.append(f"{label} – {', '.join(details)}")
            else:
                success_messages.append(f"{label} supprimé.")

        if removed_count:
            updated_entry = dict(entry)
            updated_entry["addons"] = normalize_addon_metadata(entry_addons)
            updated_entry["installed_at"] = datetime.utcnow().replace(microsecond=0).isoformat()
            self._record_installation(updated_entry)
            self.installations_performed = True

        if success_messages:
            QtWidgets.QMessageBox.information(
                self,
                "Suppression terminée",
                "\n".join(success_messages),
            )
        if error_messages:
            QtWidgets.QMessageBox.warning(
                self,
                "Suppression partielle",
                "\n".join(error_messages),
            )

    def _remove_addon_files(self, entry, addon):
        target_folder = entry.get("target_folder")
        removed = []
        missing = []
        errors = []

        if not target_folder or not os.path.isdir(target_folder):
            return removed, missing, ["Dossier du mod introuvable."]

        raw_paths = []
        if isinstance(addon, dict):
            raw_paths = list(addon.get("paths", []))
            if not raw_paths and addon.get("label"):
                raw_paths = [addon.get("label")]
        else:
            raw_paths = [str(addon)]

        entries = []
        seen = set()
        for raw_path in raw_paths:
            path_str = str(raw_path).replace("\\", "/").strip()
            if not path_str:
                continue
            is_directory = path_str.endswith("/")
            trimmed = path_str[:-1] if is_directory else path_str
            normalized = os.path.normpath(trimmed).replace("\\", "/")
            if not normalized or normalized in {".", ""}:
                continue
            if normalized.startswith("..") or "/../" in normalized:
                errors.append(f"Chemin invalide ignoré : {path_str}")
                continue
            key = (normalized, is_directory)
            if key in seen:
                continue
            seen.add(key)
            entries.append((normalized, is_directory))

        if not entries:
            return removed, missing, errors

        target_root = os.path.realpath(target_folder)

        def _sorted_by_depth(items):
            return sorted(items, key=lambda value: (value.count("/"), value), reverse=True)

        file_entries = [value for value, is_dir in entries if not is_dir]
        dir_entries = [value for value, is_dir in entries if is_dir]

        for rel_path in _sorted_by_depth(file_entries):
            absolute_path = os.path.realpath(os.path.join(target_root, rel_path))
            if os.path.commonpath([target_root, absolute_path]) != target_root:
                errors.append(f"Chemin hors du mod : {rel_path}")
                continue
            if os.path.isfile(absolute_path):
                try:
                    os.remove(absolute_path)
                    removed.append(rel_path)
                except OSError as exc:
                    errors.append(f"{rel_path} : {exc}")
            elif os.path.isdir(absolute_path):
                try:
                    shutil.rmtree(absolute_path)
                    removed.append(f"{rel_path}/")
                except OSError as exc:
                    errors.append(f"{rel_path} : {exc}")
            else:
                missing.append(rel_path)

        for rel_path in _sorted_by_depth(dir_entries):
            absolute_path = os.path.realpath(os.path.join(target_root, rel_path))
            if os.path.commonpath([target_root, absolute_path]) != target_root:
                errors.append(f"Chemin hors du mod : {rel_path}/")
                continue
            if os.path.isdir(absolute_path):
                try:
                    shutil.rmtree(absolute_path)
                    removed.append(f"{rel_path}/")
                except OSError as exc:
                    errors.append(f"{rel_path}/ : {exc}")
            elif os.path.isfile(absolute_path):
                try:
                    os.remove(absolute_path)
                    removed.append(rel_path)
                except OSError as exc:
                    errors.append(f"{rel_path}/ : {exc}")
            else:
                missing.append(f"{rel_path}/")

        return removed, missing, errors

    def _perform_update(self, entry, file_paths):
        target_folder = entry.get("target_folder")
        if not target_folder:
            return [], ["Dossier cible invalide."]

        success_messages = []
        error_messages = []
        processed_files = []

        for index, path in enumerate(file_paths):
            if not os.path.isfile(path):
                error_messages.append(f"Fichier introuvable : {path}")
                continue
            if not self._is_supported_extension(path):
                error_messages.append(f"Extension non supportée : {os.path.basename(path)}")
                continue
            success, message, _ = self._install_file_to_target(
                path,
                target_folder,
                clean_before=(index == 0),
                merge=False,
            )
            if success:
                success_messages.append(message)
                processed_files.append(path)
            elif message:
                error_messages.append(message)

        if processed_files:
            updated_entry = dict(entry)
            updated_entry["installed_at"] = datetime.utcnow().replace(microsecond=0).isoformat()
            updated_entry["type"] = self._describe_install_type(processed_files)
            updated_entry["source"] = ", ".join(os.path.basename(path) for path in processed_files)
            updated_entry["addons"] = []
            self._record_installation(updated_entry)
            self._write_marker_file(target_folder, updated_entry)
            # Prompt optional metadata after update too
            try:
                self._prompt_mod_metadata(updated_entry)
            except Exception:
                pass
            self.installations_performed = True

        return success_messages, error_messages

    def _perform_addons(self, entry, file_paths):
        # Block while Sims 4 is running
        try:
            if hasattr(self.parent(), "_is_sims_running") and self.parent()._is_sims_running():
                return [], ["Ajout d'add-ons impossible: TS4_x64.exe est en cours d'exécution." ]
        except Exception:
            pass
        target_folder = entry.get("target_folder")
        if not target_folder:
            return [], ["Dossier cible invalide."]

        success_messages = []
        error_messages = []
        added_sources = []
        new_addons = []

        for path in file_paths:
            if not os.path.isfile(path):
                error_messages.append(f"Fichier introuvable : {path}")
                continue
            if not self._is_supported_extension(path):
                error_messages.append(f"Extension non supportée : {os.path.basename(path)}")
                continue
            success, message, installed_paths = self._install_file_to_target(
                path,
                target_folder,
                clean_before=False,
                merge=True,
            )
            if success:
                success_messages.append(message)
                label = os.path.basename(path)
                added_sources.append(label)
                new_addons.append(
                    {
                        "label": label,
                        "paths": installed_paths,
                        "added_at": datetime.utcnow().replace(microsecond=0).isoformat(),
                    }
                )
            elif message:
                error_messages.append(message)

        if added_sources:
            updated_entry = dict(entry)
            existing_addons = normalize_addon_metadata(updated_entry.get("addons", []))
            existing_addons.extend(new_addons)
            updated_entry["addons"] = normalize_addon_metadata(existing_addons)
            updated_entry["installed_at"] = datetime.utcnow().replace(microsecond=0).isoformat()
            self._record_installation(updated_entry)
            self._write_marker_file(target_folder, updated_entry)
            self.installations_performed = True

        return success_messages, error_messages

    def _delete_mod(self, entry):
        # Prevent deletion if Sims is running
        try:
            if hasattr(self.parent(), "_is_sims_running") and self.parent()._is_sims_running():
                QtWidgets.QMessageBox.warning(self, "Suppression impossible", "TS4_x64.exe est en cours d'exécution. Fermez le jeu avant de supprimer des mods.")
                return
        except Exception:
            pass
        target_folder = entry.get("target_folder")
        mod_name = entry.get("name") or os.path.basename(target_folder or "")
        response = QtWidgets.QMessageBox.question(
            self,
            "Supprimer le mod",
            (
                f"Voulez-vous supprimer le mod '{mod_name}' ?\n"
                "Le dossier correspondant sera supprimé du disque."
            ),
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if response != QtWidgets.QMessageBox.Yes:
            return

        if target_folder and os.path.exists(target_folder):
            try:
                shutil.rmtree(target_folder)
            except OSError as exc:
                QtWidgets.QMessageBox.critical(
                    self,
                    "Suppression impossible",
                    f"Le dossier n'a pas pu être supprimé : {exc}",
                )
                return

        self.installed_mods = [item for item in self.installed_mods if item.get("target_folder") != target_folder]
        save_installed_mods(self.installed_mods)
        self.refresh_table()
        self.installations_performed = True

    def refresh_table(self):
        self.table.blockSignals(True)
        self.table.setRowCount(len(self.installed_mods))
        for row, entry in enumerate(self.installed_mods):
            mod_name = entry.get("name", "")
            mod_type = entry.get("type", "")
            installed_at = format_installation_display(entry.get("installed_at", ""))
            mod_version = entry.get("mod_version", "")
            folder_name = os.path.basename(entry.get("target_folder", ""))
            status = "Désactivé" if entry.get("disabled") else ""
            url = entry.get("url", "")

            # Fill columns in new order: Mod, Type, Installé le, Version, Dossier, Addons, Statut, URL
            columns_values = [mod_name, mod_type, installed_at, mod_version, folder_name, "", status, url]
            for column, value in enumerate(columns_values):
                if column == 5:
                    # Addons flag as read-only checkbox
                    item = QtWidgets.QTableWidgetItem("")
                    try:
                        item.setFlags((item.flags() | QtCore.Qt.ItemIsUserCheckable) & ~QtCore.Qt.ItemIsEditable)
                        item.setCheckState(QtCore.Qt.Checked if entry.get("addons") else QtCore.Qt.Unchecked)
                    except Exception:
                        pass
                else:
                    item = QtWidgets.QTableWidgetItem(value)
                    # Only version (3) and URL (7) are editable
                    if column in (3, 7):
                        try:
                            item.setFlags(item.flags() | QtCore.Qt.ItemIsEditable)
                        except Exception:
                            pass
                    else:
                        item.setFlags(item.flags() ^ QtCore.Qt.ItemIsEditable)
                # visual cue for disabled
                if status:
                    try:
                        item.setForeground(QtGui.QBrush(QtGui.QColor("#aaaaaa")))
                    except Exception:
                        pass
                self.table.setItem(row, column, item)
        # Protected highlight overrides status coloring
        if entry.get("atf"):
            bg = QtGui.QBrush(QtGui.QColor("#ffc0cb"))
            fg = QtGui.QBrush(QtGui.QColor("#000000"))
            for c in range(self.table.columnCount()):
                it = self.table.item(row, c)
                if it is not None:
                    it.setBackground(bg)
                    it.setForeground(fg)
        self.table.blockSignals(False)
        # Ensure change handler connected once
        try:
            if not getattr(self, "_item_changed_connected", False):
                self.table.itemChanged.connect(self._on_installer_item_changed)
                self._item_changed_connected = True
        except Exception:
            pass
        self._update_header_labels()
        self._apply_table_filters()
        # Ensure column visibility persists
        try:
            self._apply_installer_hidden_columns()
        except Exception:
            pass


class DuplicateFinderDialog(QtWidgets.QDialog):
    def __init__(self, parent, root_directory):
        super().__init__(parent)
        self.setWindowTitle("Find duplicates")
        self.setModal(True)
        try:
            self.setWindowFlags(self.windowFlags() | QtCore.Qt.Window)
            self.setSizeGripEnabled(True)
        except Exception:
            pass
        self.resize(800, 520)
        self.root_directory = os.path.abspath(root_directory)

        layout = QtWidgets.QVBoxLayout(self)

        # Controls
        path_layout = QtWidgets.QHBoxLayout()
        path_layout.addWidget(QtWidgets.QLabel("Dossier scanné :", self))
        self.root_edit = QtWidgets.QLineEdit(self)
        self.root_edit.setReadOnly(True)
        self.root_edit.setText(self.root_directory)
        path_layout.addWidget(self.root_edit)
        layout.addLayout(path_layout)

        options_layout = QtWidgets.QHBoxLayout()
        self.strict_checkbox = QtWidgets.QCheckBox("Scan avancé (même taille)", self)
        self.strict_checkbox.setChecked(False)
        options_layout.addWidget(self.strict_checkbox)

        self.package_checkbox = QtWidgets.QCheckBox("Inclure .package", self)
        self.package_checkbox.setChecked(True)
        options_layout.addWidget(self.package_checkbox)

        self.ts4_checkbox = QtWidgets.QCheckBox("Inclure .ts4script", self)
        self.ts4_checkbox.setChecked(True)
        options_layout.addWidget(self.ts4_checkbox)

        options_layout.addStretch(1)
        self.scan_button = QtWidgets.QPushButton("Analyser", self)
        self.scan_button.clicked.connect(self._run_scan)
        options_layout.addWidget(self.scan_button)
        layout.addLayout(options_layout)

        # Results tree
        self.tree = QtWidgets.QTreeWidget(self)
        self.tree.setColumnCount(4)
        self.tree.setHeaderLabels(["Fichier", "Chemin", "Taille (o)", "Modifié le"])
        self.tree.setAlternatingRowColors(True)
        self.tree.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.tree.itemDoubleClicked.connect(self._open_item_in_explorer)
        self.tree.itemSelectionChanged.connect(self._update_delete_button_enabled)
        layout.addWidget(self.tree, stretch=1)

        # Footer
        footer_layout = QtWidgets.QHBoxLayout()
        self.summary_label = QtWidgets.QLabel("", self)
        self.progress_bar = QtWidgets.QProgressBar(self)
        self.progress_bar.setVisible(False)
        self.progress_bar.setMinimum(0)
        self.progress_bar.setMaximum(1)
        self.progress_bar.setValue(0)
        footer_layout.addWidget(self.summary_label)
        footer_layout.addWidget(self.progress_bar, stretch=1)
        footer_layout.addStretch(1)
        self.delete_button = QtWidgets.QPushButton("Supprimer sélection", self)
        self.delete_button.setEnabled(False)
        self.delete_button.clicked.connect(self._delete_selected)
        footer_layout.addWidget(self.delete_button)
        close_button = QtWidgets.QPushButton("Fermer", self)
        close_button.clicked.connect(self.accept)
        footer_layout.addWidget(close_button)
        layout.addLayout(footer_layout)

        QtCore.QTimer.singleShot(0, self._run_scan)

    def _open_item_in_explorer(self, item, _column):
        path = item.text(1).strip()
        if not path:
            return
        directory = os.path.dirname(path) or path
        parent = self.parent()
        if hasattr(parent, "_open_in_file_manager"):
            parent._open_in_file_manager(directory)

    def _iter_files(self, include_package=True, include_ts4=True):
        allowed = set()
        if include_package:
            allowed.add(".package")
        if include_ts4:
            allowed.add(".ts4script")
        backups_directory = ""
        try:
            if hasattr(self.parent(), "settings"):
                backups_directory = os.path.normpath(self.parent().settings.get("backups_directory", ""))
        except Exception:
            backups_directory = ""

        for current_root, dirs, files in os.walk(self.root_directory):
            # skip backups directory subtree
            if backups_directory:
                dirs[:] = [d for d in dirs if os.path.normpath(os.path.join(current_root, d)) != backups_directory]
            for file_name in files:
                ext = os.path.splitext(file_name)[1].lower()
                if ext not in allowed:
                    continue
                full_path = os.path.join(current_root, file_name)
                yield full_path, file_name

    def _run_scan(self):
        # Block while Sims is running
        try:
            if hasattr(self.parent(), "_is_sims_running") and self.parent()._is_sims_running():
                QtWidgets.QMessageBox.warning(self, "Opération bloquée", "TS4_x64.exe est en cours d'exécution. Fermez le jeu pour analyser un dossier.")
                return
        except Exception:
            pass
        self.tree.clear()
        self.scan_button.setEnabled(False)
        self.delete_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(0)  # busy while enumerating
        self.progress_bar.setValue(0)
        self._yield()
        include_package = self.package_checkbox.isChecked()
        include_ts4 = self.ts4_checkbox.isChecked()
        strict = self.strict_checkbox.isChecked()
        # enumerate files first
        files = list(self._iter_files(include_package, include_ts4))
        total = len(files)
        self.progress_bar.setMaximum(max(1, total))
        self.progress_bar.setValue(0)

        records_by_key = defaultdict(list)
        for idx, (full_path, file_name) in enumerate(files, start=1):
            try:
                stat_result = os.stat(full_path)
                size = int(stat_result.st_size)
                mtime = int(stat_result.st_mtime)
            except OSError:
                size = -1
                mtime = -1
            name_key = file_name.casefold()
            if strict:
                key = (name_key, size)  # date removed from advanced matching
            else:
                key = name_key
            records_by_key[key].append((file_name, full_path, size, mtime))

            if idx % 25 == 0 or idx == total:
                self.progress_bar.setValue(min(idx, total))
                self._yield()

        duplicate_groups = []
        for key, items in records_by_key.items():
            if len(items) > 1:
                items_sorted = sorted(items, key=lambda t: t[1].casefold())
                filename = items_sorted[0][0]
                duplicate_groups.append((filename, items_sorted))

        duplicate_groups.sort(key=lambda g: g[0].casefold())

        for filename, items in duplicate_groups:
            top = QtWidgets.QTreeWidgetItem([self._format_group_title(filename, len(items)), "", "", ""])
            self.tree.addTopLevelItem(top)
            for _, path, size, mtime in items:
                dt = datetime.fromtimestamp(mtime) if mtime and mtime > 0 else None
                top.addChild(QtWidgets.QTreeWidgetItem([
                    "",
                    os.path.abspath(path),
                    str(size if size >= 0 else ""),
                    format_datetime(dt) if dt else "",
                ]))
            top.setExpanded(True)

        self.progress_bar.setVisible(False)
        self.scan_button.setEnabled(True)
        self._update_summary()
        self._update_delete_button_enabled()
        if hasattr(self.parent(), "logger"):
            try:
                self.parent().logger.info("Duplicate scan in %s: %d groups", self.root_directory, len(duplicate_groups))
            except Exception:
                pass

    def _yield(self):
        app = QtWidgets.QApplication.instance()
        if app is not None:
            try:
                app.processEvents()
            except Exception:
                pass

    def _format_group_title(self, filename, count):
        return f"{filename} ({count})"

    def _update_group_title(self, group_item):
        if group_item is None:
            return
        # Count only children that are still present
        count = group_item.childCount()
        # Extract original filename from current title (before space + '(')
        title = group_item.text(0)
        base = title.split(' (', 1)[0]
        group_item.setText(0, self._format_group_title(base, count))

    def _update_summary(self):
        self.summary_label.setText(f"Groupes de doublons : {self.tree.topLevelItemCount()}")

    def _update_delete_button_enabled(self):
        items = self.tree.selectedItems()
        deletable = False
        for it in items:
            if it and it.parent() is not None and it.text(1).strip():
                deletable = True
                break
        self.delete_button.setEnabled(deletable)

    def _delete_selected(self):
        selected = [it for it in self.tree.selectedItems() if it.parent() is not None and it.text(1).strip()]
        if not selected:
            return
        # Unique paths only
        to_delete = []
        seen = set()
        for it in selected:
            path = it.text(1).strip()
            if path and path not in seen:
                seen.add(path)
                to_delete.append((it, path))

        confirm = QtWidgets.QMessageBox.question(
            self,
            "Confirmer la suppression",
            f"Supprimer définitivement {len(to_delete)} fichier(s) ?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if confirm != QtWidgets.QMessageBox.Yes:
            return

        errors = []
        deleted = 0
        # Delete files and remove items from tree
        for item, path in to_delete:
            try:
                if os.path.isfile(path):
                    os.remove(path)
                else:
                    raise OSError("fichier introuvable")
                parent = item.parent()
                idx = parent.indexOfChild(item)
                parent.takeChild(idx)
                self._update_group_title(parent)
                # Remove group if less than 2 remain
                if parent.childCount() < 2:
                    top_index = self.tree.indexOfTopLevelItem(parent)
                    self.tree.takeTopLevelItem(top_index)
                deleted += 1
            except OSError as exc:
                errors.append(f"{path} → {exc}")

        self._update_summary()
        self._update_delete_button_enabled()
        if deleted:
            QtWidgets.QMessageBox.information(self, "Suppression terminée", f"{deleted} fichier(s) supprimé(s).")
        if errors:
            QtWidgets.QMessageBox.warning(self, "Erreurs de suppression", "\n".join(errors))


class ModComparatorDialog(QtWidgets.QDialog):
    def __init__(self, parent, start_directory):
        super().__init__(parent)
        self.setWindowTitle("Comparateur de mods")
        self.setModal(True)
        try:
            self.setWindowFlags(self.windowFlags() | QtCore.Qt.Window)
            self.setSizeGripEnabled(True)
        except Exception:
            pass
        self.resize(980, 620)
        self.parent_app = parent
        self.left_dir = os.path.abspath(start_directory) if start_directory else ""
        self.right_dir = os.path.abspath(start_directory) if start_directory else ""

        layout = QtWidgets.QVBoxLayout(self)

        # Pickers
        pick_row = QtWidgets.QHBoxLayout()
        pick_row.addWidget(QtWidgets.QLabel("Mod A :", self))
        self.left_edit = QtWidgets.QLineEdit(self)
        self.left_edit.setText(self.left_dir)
        pick_row.addWidget(self.left_edit, stretch=1)
        left_browse = QtWidgets.QPushButton("Parcourir…", self)
        left_browse.clicked.connect(lambda: self._browse(self.left_edit))
        pick_row.addWidget(left_browse)

        pick_row.addSpacing(10)
        pick_row.addWidget(QtWidgets.QLabel("Mod B :", self))
        self.right_edit = QtWidgets.QLineEdit(self)
        self.right_edit.setText(self.right_dir)
        pick_row.addWidget(self.right_edit, stretch=1)
        right_browse = QtWidgets.QPushButton("Parcourir…", self)
        right_browse.clicked.connect(lambda: self._browse(self.right_edit))
        pick_row.addWidget(right_browse)

        layout.addLayout(pick_row)

        # Options
        opt_row = QtWidgets.QHBoxLayout()
        self.hash_checkbox = QtWidgets.QCheckBox("Comparer le contenu (hash)", self)
        self.hash_checkbox.setChecked(True)
        opt_row.addWidget(self.hash_checkbox)
        self.open_archives_checkbox = QtWidgets.QCheckBox("Comparer l'intérieur des .ts4script", self)
        self.open_archives_checkbox.setChecked(True)
        opt_row.addWidget(self.open_archives_checkbox)
        opt_row.addStretch(1)
        self.compare_button = QtWidgets.QPushButton("Comparer", self)
        self.compare_button.clicked.connect(self._run_compare)
        opt_row.addWidget(self.compare_button)
        layout.addLayout(opt_row)

        # Results
        self.tree = QtWidgets.QTreeWidget(self)
        self.tree.setColumnCount(4)
        self.tree.setHeaderLabels(["Élément", "Mod A", "Mod B", "Détail"])
        self.tree.setAlternatingRowColors(True)
        self.tree.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        layout.addWidget(self.tree, stretch=1)

        # Footer summary
        self.summary_label = QtWidgets.QLabel("", self)
        layout.addWidget(self.summary_label)

    def _browse(self, line_edit):
        folder = QtWidgets.QFileDialog.getExistingDirectory(self, "Choisir un dossier de mod")
        if folder:
            line_edit.setText(folder)

    def _hash_file(self, path):
        h = hashlib.sha1()
        try:
            with open(path, "rb") as fh:
                while True:
                    chunk = fh.read(1024 * 1024)
                    if not chunk:
                        break
                    h.update(chunk)
            return h.hexdigest()
        except OSError:
            return ""

    def _scan_mod(self, root, *, hash_content=True, open_archives=True):
        root = os.path.abspath(root)
        result = {
            "packages": {},  # rel -> {size, hash}
            "scripts": {},   # rel -> {size, hash, inner: {name->crc/size}}
            "totals": {"packages": 0, "scripts": 0, "script_entries": 0},
            "addons_count": 0,
        }
        marker_path = os.path.join(root, MOD_MARKER_FILENAME)
        if os.path.isfile(marker_path):
            try:
                with open(marker_path, "r", encoding="utf-8") as fh:
                    data = json.load(fh)
                result["addons_count"] = len(data.get("addons") or [])
            except Exception:
                pass
        for cur, _dirs, files in os.walk(root):
            for fname in files:
                full = os.path.join(cur, fname)
                rel = os.path.relpath(full, root).replace("\\", "/")
                low = fname.lower()
                try:
                    st = os.stat(full)
                except OSError:
                    continue
                if low.endswith(".package"):
                    info = {"size": int(st.st_size)}
                    if hash_content:
                        info["hash"] = self._hash_file(full)
                    result["packages"][rel] = info
                elif low.endswith(".ts4script"):
                    info = {"size": int(st.st_size)}
                    if hash_content:
                        info["hash"] = self._hash_file(full)
                    if open_archives:
                        inner = {}
                        try:
                            with zipfile.ZipFile(full, "r") as zf:
                                for zi in zf.infolist():
                                    if zi.is_dir():
                                        continue
                                    inner[zi.filename.replace("\\", "/")] = {"size": zi.file_size, "crc": zi.CRC}
                        except (OSError, zipfile.BadZipFile):
                            pass
                        info["inner"] = inner
                    result["scripts"][rel] = info
        result["totals"]["packages"] = len(result["packages"])
        result["totals"]["scripts"] = len(result["scripts"])
        result["totals"]["script_entries"] = sum(len(v.get("inner", {})) for v in result["scripts"].values())
        return result

    def _diff_dicts(self, left, right):
        left_keys = set(left.keys())
        right_keys = set(right.keys())
        added = sorted(right_keys - left_keys, key=str.casefold)
        removed = sorted(left_keys - right_keys, key=str.casefold)
        common = sorted(left_keys & right_keys, key=str.casefold)
        changed = []
        for k in common:
            if left[k] != right[k]:
                changed.append(k)
        return added, removed, sorted(changed, key=str.casefold)

    def _run_compare(self):
        self.tree.clear()
        left = self.left_edit.text().strip()
        right = self.right_edit.text().strip()
        if not left or not right or not os.path.isdir(left) or not os.path.isdir(right):
            QtWidgets.QMessageBox.warning(self, "Entrées invalides", "Sélectionnez deux dossiers de mods valides.")
            return
        hash_content = self.hash_checkbox.isChecked()
        open_arch = self.open_archives_checkbox.isChecked()
        try:
            left_res = self._scan_mod(left, hash_content=hash_content, open_archives=open_arch)
            right_res = self._scan_mod(right, hash_content=hash_content, open_archives=open_arch)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Erreur", str(exc))
            return

        root = QtWidgets.QTreeWidgetItem(self.tree, ["Comparaison", os.path.basename(left), os.path.basename(right), ""])

        # Totals and addons
        totals = QtWidgets.QTreeWidgetItem(root, ["Résumé", "", "", ""]) 
        QtWidgets.QTreeWidgetItem(totals, ["Packages", str(left_res['totals']['packages']), str(right_res['totals']['packages']), ""]) 
        QtWidgets.QTreeWidgetItem(totals, ["TS4Scripts", str(left_res['totals']['scripts']), str(right_res['totals']['scripts']), ""]) 
        QtWidgets.QTreeWidgetItem(totals, ["Entrées TS4Scripts", str(left_res['totals']['script_entries']), str(right_res['totals']['script_entries']), ""]) 
        QtWidgets.QTreeWidgetItem(totals, ["Add-ons", str(left_res['addons_count']), str(right_res['addons_count']), ""]) 
        totals.setExpanded(True)

        # Packages diff
        pkg_node = QtWidgets.QTreeWidgetItem(root, ["Packages", "", "", ""])
        la, lr, lc = self._diff_dicts(left_res["packages"], right_res["packages"])
        if la:
            node = QtWidgets.QTreeWidgetItem(pkg_node, ["Ajoutés", "", "", ""])
            for k in la:
                QtWidgets.QTreeWidgetItem(node, [k, "—", "présent", ""])
        if lr:
            node = QtWidgets.QTreeWidgetItem(pkg_node, ["Supprimés", "", "", ""])
            for k in lr:
                QtWidgets.QTreeWidgetItem(node, [k, "présent", "—", ""])
        if lc:
            node = QtWidgets.QTreeWidgetItem(pkg_node, ["Modifiés", "", "", ""])
            for k in lc:
                l = left_res["packages"][k].get("hash") or left_res["packages"][k].get("size")
                r = right_res["packages"][k].get("hash") or right_res["packages"][k].get("size")
                QtWidgets.QTreeWidgetItem(node, [k, str(l), str(r), "hash/taille diff"])
        pkg_node.setExpanded(True)

        # Scripts diff
        sc_node = QtWidgets.QTreeWidgetItem(root, ["TS4Script", "", "", ""])
        sa, sr, sc = self._diff_dicts(left_res["scripts"], right_res["scripts"])
        if sa:
            node = QtWidgets.QTreeWidgetItem(sc_node, ["Ajoutés", "", "", ""])
            for k in sa:
                QtWidgets.QTreeWidgetItem(node, [k, "—", "présent", ""])
        if sr:
            node = QtWidgets.QTreeWidgetItem(sc_node, ["Supprimés", "", "", ""])  # keep consistent
            for k in sr:
                QtWidgets.QTreeWidgetItem(node, [k, "présent", "—", ""])
        if sc:
            node = QtWidgets.QTreeWidgetItem(sc_node, ["Modifiés", "", "", ""])
            for k in sc:
                l = left_res["scripts"][k].get("hash") or left_res["scripts"][k].get("size")
                r = right_res["scripts"][k].get("hash") or right_res["scripts"][k].get("size")
                QtWidgets.QTreeWidgetItem(node, [k, str(l), str(r), "hash/taille diff ou contenu"])
        sc_node.setExpanded(True)

        # Inner scripts diff when requested
        if open_arch:
            inner_node = QtWidgets.QTreeWidgetItem(root, ["Contenu .ts4script", "", "", ""])
            common_scripts = sorted(set(left_res["scripts"].keys()) & set(right_res["scripts"].keys()), key=str.casefold)
            for k in common_scripts:
                l_inner = left_res["scripts"][k].get("inner", {})
                r_inner = right_res["scripts"][k].get("inner", {})
                ia, ir, ic = self._diff_dicts(l_inner, r_inner)
                if ia or ir or ic:
                    script_node = QtWidgets.QTreeWidgetItem(inner_node, [k, "", "", ""])
                    if ia:
                        n = QtWidgets.QTreeWidgetItem(script_node, ["Ajoutés", "", "", ""])
                        for name in ia:
                            QtWidgets.QTreeWidgetItem(n, [name, "—", "présent", ""])
                    if ir:
                        n = QtWidgets.QTreeWidgetItem(script_node, ["Supprimés", "", "", ""])
                        for name in ir:
                            QtWidgets.QTreeWidgetItem(n, [name, "présent", "—", ""])
                    if ic:
                        n = QtWidgets.QTreeWidgetItem(script_node, ["Modifiés", "", "", ""])
                        for name in ic:
                            l = l_inner[name]
                            r = r_inner[name]
                            detail = "crc/size diff"
                            QtWidgets.QTreeWidgetItem(n, [name, f"{l.get('size')}/{l.get('crc')}", f"{r.get('size')}/{r.get('crc')}", detail])
            inner_node.setExpanded(True)

        self.summary_label.setText("Comparaison terminée.")

    def _open_item_in_explorer(self, item, _column):
        # If it is a child (has a path), attempt to open its folder
        path = item.text(1).strip()
        if not path:
            return
        directory = os.path.dirname(path) or path
        parent = self.parent()
        if hasattr(parent, "_open_in_file_manager"):
            parent._open_in_file_manager(directory)

    def _iter_files(self, include_package=True, include_ts4=True):
        allowed = set()
        if include_package:
            allowed.add(".package")
        if include_ts4:
            allowed.add(".ts4script")
        backups_directory = ""
        try:
            if hasattr(self.parent(), "settings"):
                backups_directory = os.path.normpath(self.parent().settings.get("backups_directory", ""))
        except Exception:
            backups_directory = ""

        for current_root, dirs, files in os.walk(self.root_directory):
            # skip backups directory subtree
            if backups_directory:
                dirs[:] = [d for d in dirs if os.path.normpath(os.path.join(current_root, d)) != backups_directory]
            for file_name in files:
                ext = os.path.splitext(file_name)[1].lower()
                if ext not in allowed:
                    continue
                full_path = os.path.join(current_root, file_name)
                yield full_path, file_name

    def _run_scan(self):
        # Block while Sims is running
        try:
            if hasattr(self.parent(), "_is_sims_running") and self.parent()._is_sims_running():
                QtWidgets.QMessageBox.warning(self, "Opération bloquée", "TS4_x64.exe est en cours d'exécution. Fermez le jeu pour analyser un dossier.")
                return
        except Exception:
            pass
        self.tree.clear()
        self.scan_button.setEnabled(False)
        self.delete_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setMaximum(0)  # busy while enumerating
        self.progress_bar.setValue(0)
        self._yield()
        include_package = self.package_checkbox.isChecked()
        include_ts4 = self.ts4_checkbox.isChecked()
        strict = self.strict_checkbox.isChecked()
        # enumerate files first
        files = list(self._iter_files(include_package, include_ts4))
        total = len(files)
        self.progress_bar.setMaximum(max(1, total))
        self.progress_bar.setValue(0)

        records_by_key = defaultdict(list)
        for idx, (full_path, file_name) in enumerate(files, start=1):
            try:
                stat_result = os.stat(full_path)
                size = int(stat_result.st_size)
                mtime = int(stat_result.st_mtime)
            except OSError:
                size = -1
                mtime = -1
            name_key = file_name.casefold()
            if strict:
                key = (name_key, size)  # date removed from advanced matching
            else:
                key = name_key
            records_by_key[key].append((file_name, full_path, size, mtime))

            if idx % 25 == 0 or idx == total:
                self.progress_bar.setValue(min(idx, total))
                self._yield()

        duplicate_groups = []
        for key, items in records_by_key.items():
            if len(items) > 1:
                items_sorted = sorted(items, key=lambda t: t[1].casefold())
                filename = items_sorted[0][0]
                duplicate_groups.append((filename, items_sorted))

        duplicate_groups.sort(key=lambda g: g[0].casefold())

        for filename, items in duplicate_groups:
            top = QtWidgets.QTreeWidgetItem([self._format_group_title(filename, len(items)), "", "", ""])
            self.tree.addTopLevelItem(top)
            for _, path, size, mtime in items:
                dt = datetime.fromtimestamp(mtime) if mtime and mtime > 0 else None
                top.addChild(QtWidgets.QTreeWidgetItem([
                    "",
                    os.path.abspath(path),
                    str(size if size >= 0 else ""),
                    format_datetime(dt) if dt else "",
                ]))
            top.setExpanded(True)

        self.progress_bar.setVisible(False)
        self.scan_button.setEnabled(True)
        self._update_summary()
        self._update_delete_button_enabled()
        if hasattr(self.parent(), "logger"):
            try:
                self.parent().logger.info("Duplicate scan in %s: %d groups", self.root_directory, len(duplicate_groups))
            except Exception:
                pass

    def _yield(self):
        app = QtWidgets.QApplication.instance()
        if app is not None:
            try:
                app.processEvents()
            except Exception:
                pass

    def _format_group_title(self, filename, count):
        return f"{filename} ({count})"

    def _update_group_title(self, group_item):
        if group_item is None:
            return
        # Count only children that are still present
        count = group_item.childCount()
        # Extract original filename from current title (before space + '(')
        title = group_item.text(0)
        base = title.split(' (', 1)[0]
        group_item.setText(0, self._format_group_title(base, count))

    def _update_summary(self):
        self.summary_label.setText(f"Groupes de doublons : {self.tree.topLevelItemCount()}")

    def _update_delete_button_enabled(self):
        items = self.tree.selectedItems()
        deletable = False
        for it in items:
            if it and it.parent() is not None and it.text(1).strip():
                deletable = True
                break
        self.delete_button.setEnabled(deletable)

    def _delete_selected(self):
        selected = [it for it in self.tree.selectedItems() if it.parent() is not None and it.text(1).strip()]
        if not selected:
            return
        # Unique paths only
        to_delete = []
        seen = set()
        for it in selected:
            path = it.text(1).strip()
            if path and path not in seen:
                seen.add(path)
                to_delete.append((it, path))

        confirm = QtWidgets.QMessageBox.question(
            self,
            "Confirmer la suppression",
            f"Supprimer définitivement {len(to_delete)} fichier(s) ?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if confirm != QtWidgets.QMessageBox.Yes:
            return

        errors = []
        deleted = 0
        # Delete files and remove items from tree
        for item, path in to_delete:
            try:
                if os.path.isfile(path):
                    os.remove(path)
                else:
                    raise OSError("fichier introuvable")
                parent = item.parent()
                idx = parent.indexOfChild(item)
                parent.takeChild(idx)
                self._update_group_title(parent)
                # Remove group if less than 2 remain
                if parent.childCount() < 2:
                    top_index = self.tree.indexOfTopLevelItem(parent)
                    self.tree.takeTopLevelItem(top_index)
                deleted += 1
            except OSError as exc:
                errors.append(f"{path} → {exc}")

        self._update_summary()
        self._update_delete_button_enabled()
        if deleted:
            QtWidgets.QMessageBox.information(self, "Suppression terminée", f"{deleted} fichier(s) supprimé(s).")
        if errors:
            QtWidgets.QMessageBox.warning(self, "Erreurs de suppression", "\n".join(errors))


class ConflictCheckerDialog(QtWidgets.QDialog):
    def __init__(self, parent, root_directory):
        super().__init__(parent)
        self.setWindowTitle("Conflict Checker")
        self.setModal(True)
        try:
            self.setWindowFlags(self.windowFlags() | QtCore.Qt.Window)
            self.setSizeGripEnabled(True)
        except Exception:
            pass
        self.resize(860, 560)
        self.root = os.path.abspath(root_directory)
        layout = QtWidgets.QVBoxLayout(self)
        info = QtWidgets.QLabel("Détecte plusieurs versions d'un même mod (basé sur les noms et suffixes de version).", self)
        info.setWordWrap(True)
        layout.addWidget(info)
        self.tree = QtWidgets.QTreeWidget(self)
        self.tree.setHeaderLabels(["Mod", "Version", "Fichier", "Date"])
        # Améliore la lisibilité
        self.tree.setAlternatingRowColors(False)
        try:
            self.tree.setUniformRowHeights(True)
        except Exception:
            pass
        layout.addWidget(self.tree, 1)
        footer = QtWidgets.QHBoxLayout()
        self.delete_btn = QtWidgets.QPushButton("Supprimer sélection", self)
        self.delete_btn.clicked.connect(self._delete_selected)
        footer.addStretch(1)
        footer.addWidget(self.delete_btn)
        close_btn = QtWidgets.QPushButton("Fermer", self)
        close_btn.clicked.connect(self.accept)
        footer.addWidget(close_btn)
        layout.addLayout(footer)
        QtCore.QTimer.singleShot(0, self._run_scan)

    def _iter_mod_files(self, root):
        for cur, _dirs, files in os.walk(root):
            for f in files:
                low = f.lower()
                if low.endswith('.package') or low.endswith('.ts4script'):
                    yield os.path.join(cur, f)

    def _run_scan(self):
        # Regroupe par nom normalisé ET type d'extension pour éviter les faux positifs
        groups = {}
        for path in self._iter_mod_files(self.root):
            base = os.path.basename(path)
            low = base.lower()
            ext_group = 'package' if low.endswith('.package') else ('ts4script' if low.endswith('.ts4script') else 'other')
            norm = normalize_mod_basename(base)
            ver = extract_version_from_name(base)
            key = (norm, ext_group)
            groups.setdefault(key, []).append((ver, path))
        self.tree.clear()
        for (norm, ext_group), items in sorted(groups.items(), key=lambda kv: (kv[0][0], kv[0][1])):
            if len(items) < 2:
                continue
            # pick newest by mtime as keeper; mark others for deletion
            with_dates = []
            for ver, path in items:
                try:
                    ts = os.path.getmtime(path)
                except Exception:
                    ts = 0
                with_dates.append((ver or "", path, ts))
            with_dates.sort(key=lambda x: x[2], reverse=True)
            # Affiche le type pour clarté
            top = QtWidgets.QTreeWidgetItem([f"{norm} ({ext_group})", "", "", ""])
            self.tree.addTopLevelItem(top)
            for idx, (v, p, ts) in enumerate(with_dates):
                dt = format_datetime(datetime.fromtimestamp(ts)) if ts else ""
                # Toujours afficher le nom du mod pour chaque ligne pour éviter les lignes vides
                child = QtWidgets.QTreeWidgetItem([norm, v, os.path.basename(p), dt])
                child.setData(0, QtCore.Qt.UserRole, p)
                # checkbox: check older ones by default
                child.setCheckState(0, QtCore.Qt.Unchecked if idx == 0 else QtCore.Qt.Checked)
                top.addChild(child)
            top.setExpanded(True)

    def _delete_selected(self):
        to_delete = []
        for i in range(self.tree.topLevelItemCount()):
            top = self.tree.topLevelItem(i)
            for j in range(top.childCount()):
                ch = top.child(j)
                if ch.checkState(0) == QtCore.Qt.Checked:
                    p = ch.data(0, QtCore.Qt.UserRole)
                    if p:
                        to_delete.append(str(p))
        if not to_delete:
            QtWidgets.QMessageBox.information(self, "Aucune sélection", "Sélectionnez au moins un fichier à supprimer.")
            return
        confirm = QtWidgets.QMessageBox.question(self, "Confirmer", f"Supprimer définitivement {len(to_delete)} fichier(s) ?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
        if confirm != QtWidgets.QMessageBox.Yes:
            return
        errors = []
        deleted = 0
        for p in to_delete:
            try:
                if os.path.exists(p):
                    os.remove(p)
                    deleted += 1
            except OSError as exc:
                errors.append(str(exc))
        self._run_scan()
        if deleted:
            QtWidgets.QMessageBox.information(self, "Supprimés", f"{deleted} fichier(s) supprimé(s).")
        if errors:
            QtWidgets.QMessageBox.warning(self, "Erreurs", "\n".join(errors))
class FolderScannerDialog(QtWidgets.QDialog):
    def __init__(self, parent, start_directory):
        super().__init__(parent)
        self.setWindowTitle("Scan dossier de mods")
        self.setModal(True)
        try:
            self.setWindowFlags(self.windowFlags() | QtCore.Qt.Window)
        except Exception:
            pass
        self.resize(900, 580)
        try:
            self.setSizeGripEnabled(True)
            self.setWindowFlag(QtCore.Qt.WindowMaximizeButtonHint, True)
            self.setWindowFlag(QtCore.Qt.WindowMinimizeButtonHint, True)
        except Exception:
            pass
        self.parent_app = parent
        self.scan_directory = os.path.abspath(start_directory) if start_directory else ""

        layout = QtWidgets.QVBoxLayout(self)

        # Path controls
        path_layout = QtWidgets.QHBoxLayout()
        path_layout.addWidget(QtWidgets.QLabel("Dossier à scanner :", self))
        self.path_edit = QtWidgets.QLineEdit(self)
        self.path_edit.setText(self.scan_directory)
        path_layout.addWidget(self.path_edit, stretch=1)
        browse_btn = QtWidgets.QPushButton("Parcourir…", self)
        browse_btn.clicked.connect(self._browse)
        path_layout.addWidget(browse_btn)
        self.recursive_checkbox = QtWidgets.QCheckBox("Récursif", self)
        self.recursive_checkbox.setChecked(True)
        path_layout.addWidget(self.recursive_checkbox)
        scan_btn = QtWidgets.QPushButton("Analyser", self)
        scan_btn.clicked.connect(self._run_folder_scan)
        path_layout.addWidget(scan_btn)
        layout.addLayout(path_layout)

        # Table
        self.table = QtWidgets.QTableWidget(self)
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            "État",
            "Mod (groupe)",
            "Fichier .package",
            "Date .package",
            "Fichier .ts4script",
            "Date .ts4script",
            "Version",
            "Confiance",
        ])
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        for col in range(2, self.table.columnCount()):
            header.setSectionResizeMode(col, QtWidgets.QHeaderView.ResizeToContents)
        # Header and row context menus
        try:
            header.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
            header.customContextMenuRequested.connect(self._show_header_menu)
            self.table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
            self.table.customContextMenuRequested.connect(self.show_context_menu)
        except Exception:
            pass
        layout.addWidget(self.table, stretch=1)

        # Footer
        footer = QtWidgets.QHBoxLayout()
        self.status_label = QtWidgets.QLabel("", self)
        footer.addWidget(self.status_label)
        footer.addStretch(1)
        close_btn = QtWidgets.QPushButton("Fermer", self)
        close_btn.clicked.connect(self.accept)
        footer.addWidget(close_btn)
        layout.addLayout(footer)

        # No auto-run. User must click "Analyser".

    def _browse(self):
        selected = QtWidgets.QFileDialog.getExistingDirectory(self, "Choisir un dossier à scanner")
        if selected:
            self.scan_directory = selected
            self.path_edit.setText(selected)

    def _run_folder_scan(self):
        # Block scan while Sims 4 is running
        try:
            if self.parent_app and hasattr(self.parent_app, "_is_sims_running") and self.parent_app._is_sims_running():
                QtWidgets.QMessageBox.warning(self, "Opération bloquée", "TS4_x64.exe est en cours d'exécution. Fermez le jeu pour analyser un dossier.")
                return
        except Exception:
            pass
        directory = self.path_edit.text().strip()
        if not directory or not os.path.isdir(directory):
            QtWidgets.QMessageBox.warning(self, "Dossier invalide", "Sélectionne un dossier existant à analyser.")
            return
        self.status_label.setText("Analyse en cours…")
        QtWidgets.QApplication.processEvents()
        settings = getattr(self.parent_app, "settings", {}) if self.parent_app else {}
        version_releases = getattr(self.parent_app, "version_releases", {}) if self.parent_app else {}

        rows = []
        try:
            # Force unfiltered behavior like "Show both"
            settings_for_scan = dict(settings)
            settings_for_scan["enable_version_filters"] = False
            settings_for_scan["file_filter_mode"] = "both"
            settings_for_scan["show_ignored"] = True
            rows, _changed = generate_data_rows(
                directory,
                settings_for_scan,
                version_releases,
                recursive=bool(self.recursive_checkbox.isChecked()),
            )
            if hasattr(self.parent_app, "logger"):
                try:
                    self.parent_app.logger.info("Folder scan complete: %d rows for %s", len(rows), directory)
                except Exception:
                    pass
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Erreur", f"L'analyse a échoué : {exc}")
            if hasattr(self.parent_app, "logger"):
                try:
                    self.parent_app.logger.error("Folder scan error: %s", exc)
                except Exception:
                    pass
            return
        self._render(rows)
        # Remember last folder scanned
        try:
            if self.parent_app and hasattr(self.parent_app, "settings"):
                self.parent_app.settings["last_folder_scan_directory"] = directory
                save_settings(self.parent_app.settings)
        except Exception:
            pass

    def _resolve_row_paths(self, row_index):
        status_item = self.table.item(row_index, 0)
        if status_item is None:
            return []
        paths = status_item.data(QtCore.Qt.UserRole + 1)
        return list(paths) if paths else []

    def show_context_menu(self, position):
        index = self.table.indexAt(position)
        if not index.isValid():
            return
        row = index.row()
        status_item = self.table.item(row, 0)
        candidates = []
        if status_item is not None:
            stored_candidates = status_item.data(QtCore.Qt.UserRole)
            if stored_candidates:
                candidates = list(stored_candidates)
        menu = QtWidgets.QMenu(self)
        ignore_action = menu.addAction("Ignorer")
        show_in_explorer_action = menu.addAction("Afficher dans l'explorateur")
        delete_action = menu.addAction("Supprimer le mod")
        google_action = menu.addAction("Recherche Google")
        # Manquait : action Patreon
        patreon_action = menu.addAction("Chercher sur Patreon")
        patreon_action = menu.addAction("Chercher sur Patreon")
        patreon_action = menu.addAction("Chercher sur Patreon")
        # Protected toggle if grouped
        group_item = self.table.item(row, 1)
        group_name = group_item.text().strip() if group_item else ""
        # Toggle Protected
        atf_action = None
        current_atf = False
        if group_name:
            try:
                for ent in load_installed_mods():
                    if str(ent.get("name", "")).strip().casefold() == group_name.casefold():
                        current_atf = bool(ent.get("atf", False))
                        break
            except Exception:
                current_atf = False
        label = "Retirer Protected" if current_atf else "Marquer Protected"
        atf_action = menu.addAction(label)

        selected = menu.exec_(self.table.viewport().mapToGlobal(position))
        if selected == ignore_action:
            if candidates and self.parent_app:
                ignored = set(self.parent_app.settings.get("ignored_mods", []))
                head = candidates[0]
                new_state = QtCore.Qt.Unchecked if head in ignored else QtCore.Qt.Checked
                try:
                    self.parent_app.update_ignore_mod(tuple(candidates), new_state)
                except Exception:
                    if new_state == QtCore.Qt.Checked:
                        ignored.add(head)
                    else:
                        for c in candidates:
                            ignored.discard(c)
                    self.parent_app.settings["ignored_mods"] = sorted(ignored)
                    save_settings(self.parent_app.settings)
        elif selected == show_in_explorer_action:
            paths = self._resolve_row_paths(row)
            if paths:
                target = os.path.dirname(paths[0]) or paths[0]
                if self.parent_app and hasattr(self.parent_app, "_open_in_file_manager"):
                    self.parent_app._open_in_file_manager(target)
        elif selected == delete_action:
            paths = self._resolve_row_paths(row)
            if not paths:
                return
            confirm = QtWidgets.QMessageBox.question(
                self,
                "Confirmer la suppression",
                "Supprimer ce mod supprimera définitivement les fichiers associés. Continuer ?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No,
            )
            if confirm != QtWidgets.QMessageBox.Yes:
                return
            errors = []
            for p in paths:
                try:
                    if os.path.exists(p):
                        os.remove(p)
                except OSError as exc:
                    errors.append(str(exc))
            if errors:
                QtWidgets.QMessageBox.warning(self, "Erreur lors de la suppression", "\n".join(errors))
            else:
                self.table.removeRow(row)
        elif selected == google_action:
            file_name = ""
            # Prefer ts4script when present, then fallback to package
            for col in (4, 2):
                item = self.table.item(row, col)
                if item:
                    t = item.text().strip()
                    if t:
                        file_name = t
                        break
            if file_name:
                base, _ = os.path.splitext(file_name)
                if base:
                    QtGui.QDesktopServices.openUrl(QtCore.QUrl(f"https://www.google.com/search?q={quote_plus(base)}"))
        elif selected == patreon_action:
            file_name = ""
            for col in (4, 2):
                item = self.table.item(row, col)
                if item:
                    t = item.text().strip()
                    if t:
                        file_name = t
                        break
            if file_name:
                base, _ = os.path.splitext(file_name)
                if base:
                    q = quote_plus(f"site:patreon.com {base}")
                    QtGui.QDesktopServices.openUrl(QtCore.QUrl(f"https://www.google.com/search?q={q}"))
        elif atf_action is not None and selected == atf_action:
            if self.parent_app and group_name:
                self.parent_app._toggle_atf_group(group_name)

    def _show_header_menu(self, pos):
        header = self.table.horizontalHeader()
        global_pos = header.mapToGlobal(pos)
        menu = QtWidgets.QMenu(self)
        labels = [
            "État",
            "Mod (groupe)",
            "Fichier .package",
            "Date .package",
            "Fichier .ts4script",
            "Date .ts4script",
            "Version",
            "Confiance",
        ]
        for col, label in enumerate(labels):
            action = QtWidgets.QAction(label, menu)
            action.setCheckable(True)
            action.setChecked(not self.table.isColumnHidden(col))
            # toggle visibility to match action state
            action.triggered.connect(lambda checked, c=col: self.table.setColumnHidden(c, not checked))
            menu.addAction(action)
        menu.exec_(global_pos)

    # Backward-compat: if any signal calls _run_scan, delegate
    def _run_scan(self):
        self._run_folder_scan()

    def _render(self, data_rows):
        self.table.setRowCount(len(data_rows))
        for r, row in enumerate(data_rows):
            status = row.get("status", "")
            group = row.get("group", "")
            pkg = row.get("package", "")
            pkg_date = row.get("package_date", "")
            script = row.get("script", "")
            script_date = row.get("script_date", "")
            version = row.get("version", "")
            confidence = row.get("confidence", "")
            values = [status, group, pkg, pkg_date, script, script_date, version, confidence]
            for c, value in enumerate(values):
                item = QtWidgets.QTableWidgetItem(value)
                if c == 0:
                    # Store candidates and absolute paths to support context menu actions
                    item.setData(QtCore.Qt.UserRole, tuple(row.get("ignore_candidates") or []))
                    item.setData(QtCore.Qt.UserRole + 1, tuple(row.get("paths") or []))
                if c == 7:  # confidence tooltip
                    tip = row.get("confidence_tooltip", "")
                    if tip:
                        item.setToolTip(tip)
                self.table.setItem(r, c, item)
        self.status_label.setText(f"{len(data_rows)} élément(s)")



class GroupViewDialog(QtWidgets.QDialog):
    def __init__(self, parent, rows):
        super().__init__(parent)
        self.setWindowTitle("Group View")
        self.setModal(True)
        try:
            self.setWindowFlags(self.windowFlags() | QtCore.Qt.Window)
            self.setSizeGripEnabled(True)
        except Exception:
            pass
        self.resize(860, 540)

        layout = QtWidgets.QVBoxLayout(self)
        toolbar = QtWidgets.QHBoxLayout()
        expand_btn = QtWidgets.QPushButton("Expand All", self)
        collapse_btn = QtWidgets.QPushButton("Collapse All", self)
        rebuild_ai_btn = QtWidgets.QPushButton("Reconstruire Groupes (AI)", self)
        rebuild_ai_btn.setToolTip("Analyse les logs sous Mods et complète les associations de groupe via l'IA.")
        rebuild_ai_btn.clicked.connect(self._rebuild_groups_from_ai)
        toolbar.addWidget(expand_btn)
        toolbar.addWidget(collapse_btn)
        toolbar.addSpacing(8)
        toolbar.addWidget(rebuild_ai_btn)
        toolbar.addStretch(1)
        layout.addLayout(toolbar)

        self.tree = QtWidgets.QTreeWidget(self)
        self.tree.setColumnCount(7)
        self.tree.setHeaderLabels([
            "Mod (groupe)", "État", "Fichier .package", "Date .package", "Fichier .ts4script", "Date .ts4script", "Confiance"
        ])
        self.tree.setAlternatingRowColors(True)
        self.tree.setStyleSheet(
            "QTreeWidget {background-color: #1f1f1f; alternate-background-color: #2b2b2b; color: #f0f0f0;}"
            "QTreeWidget::item:selected {background-color: #4b636e;}"
        )
        layout.addWidget(self.tree, stretch=1)

        expand_btn.clicked.connect(self.tree.expandAll)
        collapse_btn.clicked.connect(self.tree.collapseAll)

        # Group rows
        grouped = defaultdict(list)
        for row in rows:
            key = row.get("group") or "(ungrouped)"
            grouped[key].append(row)
        for group_name in sorted(grouped.keys(), key=str.casefold):
            items = grouped[group_name]
            top = QtWidgets.QTreeWidgetItem([group_name, "", "", "", "", "", ""])
            if group_name and group_name != "(ungrouped)":
                top.setBackground(0, QtGui.QBrush(QtGui.QColor("#2e7d32")))
                top.setForeground(0, QtGui.QBrush(QtGui.QColor("#ffffff")))
            self.tree.addTopLevelItem(top)
            for row in items:
                child = QtWidgets.QTreeWidgetItem([
                    "",
                    str(row.get("status", "")),
                    str(row.get("package", "")),
                    str(row.get("package_date", "")),
                    str(row.get("script", "")),
                    str(row.get("script_date", "")),
                    str(row.get("confidence", "")),
                ])
                # Highlight installed via Mod Installer
                if row.get("group"):
                    for col in range(self.tree.columnCount()):
                        child.setBackground(col, QtGui.QBrush(QtGui.QColor("#2e7d32")))
                        child.setForeground(col, QtGui.QBrush(QtGui.QColor("#ffffff")))
                top.addChild(child)
            top.setExpanded(True)

    def _rebuild_groups_from_ai(self):
        parent_app = self.parent() if hasattr(self, 'parent') else None
        if parent_app is None or not hasattr(parent_app, 'settings'):
            return
        settings = parent_app.settings
        mod_dir = str(settings.get('mod_directory', '') or '')
        if not (mod_dir and os.path.isdir(mod_dir)):
            QtWidgets.QMessageBox.information(self, "Dossier invalide", "Configure un dossier de mods valide dans la Configuration.")
            return

        def _preferred_label(candidates):
            ordered = sorted((c for c in candidates if c), key=lambda x: (len(x), x.casefold()))
            return ordered[0] if ordered else ""

        def _extract_prefix(name):
            if not name:
                return ""
            m = re.match(r"\[([^\]]+)\]", name)
            if m:
                return f"[{m.group(1).strip()}]"
            m2 = re.match(r"([A-Za-z0-9]{3,})[_-]", name)
            if m2:
                return m2.group(1)
            return ""

        entries = []
        prefix_map = defaultdict(set)
        for rootd, _dirs, files in os.walk(mod_dir):
            for fn in files:
                ext = os.path.splitext(fn)[1].lower()
                if ext not in {'.package', '.ts4script'}:
                    continue
                abs_path = os.path.join(rootd, fn)
                base = os.path.splitext(fn)[0]
                norm = normalize_mod_basename(base)
                entries.append({
                    "normalized": norm,
                    "base": base,
                    "ext": ext,
                    "path": abs_path,
                })
                if ext == '.package':
                    prefix = _extract_prefix(base)
                    if prefix:
                        prefix_map[prefix].add(norm)
        if not entries:
            QtWidgets.QMessageBox.information(self, "Analyse IA", "Aucun fichier .package ou .ts4script trouvé dans le dossier sélectionné.")
            return

        by_normalized = {}
        for info in entries:
            norm = info["normalized"]
            base = info["base"]
            data = by_normalized.setdefault(norm, {"bases": set(), "exts": set(), "paths": []})
            data["bases"].add(base)
            data["exts"].add(info["ext"])
            data["paths"].append(info["path"])

        overrides = dict(settings.get('ai_group_overrides', {}) or {})
        updates = {}

        # Heuristic 1: paired package + script -> use shared stem
        for norm, data in by_normalized.items():
            if not norm:
                continue
            if '.package' in data["exts"] and '.ts4script' in data["exts"]:
                label = _preferred_label(data["bases"])
                if label:
                    updates[norm] = label

        # Heuristic 2: script without package -> standalone label
        for norm, data in by_normalized.items():
            if not norm or norm in updates:
                continue
            if '.ts4script' in data["exts"] and '.package' not in data["exts"]:
                label = _preferred_label(data["bases"])
                if label:
                    updates[norm] = label

        # Heuristic 3: common prefixes for package collections
        MIN_PREFIX_GROUP = 3
        for prefix, norms in prefix_map.items():
            valid_norms = [n for n in norms if n]
            if len(valid_norms) < MIN_PREFIX_GROUP:
                continue
            for norm in valid_norms:
                if norm in updates:
                    continue
                updates[norm] = prefix

        applied = 0
        for norm, label in updates.items():
            if not label:
                continue
            current = overrides.get(norm)
            if current == label:
                continue
            overrides[norm] = label
            applied += 1

        if applied:
            settings['ai_group_overrides'] = overrides
            save_settings(settings)
            try:
                parent_app.refresh_table_only()
            except Exception:
                pass
            QtWidgets.QMessageBox.information(
                self,
                "Groupes (AI)",
                f"{applied} associations mises à jour à partir des fichiers .package et .ts4script."
            )
        else:
            QtWidgets.QMessageBox.information(self, "Groupes (AI)", "Aucune nouvelle association n'a été trouvée.")


###############################
# ID Conflict Viewer (DBPF)   #
###############################

def _read_le_u32(b, off):
    try:
        return int.from_bytes(b[off:off+4], 'little', signed=False)
    except Exception:
        return 0

def _read_tgi_entries_from_dbpf(path, logger=None, allow_tail_fallback=True, cancel_event=None):
    """
    Robust, best-effort DBPF index reader for Sims 4 .package.
    Returns a list of (type_id, group_id, instance_id) tuples.

    Strategy:
    1) Validate DBPF magic; read header.
    2) Try multiple header layouts to fetch (index_count, index_offset, index_size).
    3) Decode index with candidate entry sizes [16, 24, 28, 32, 36, 40]; choose the size
       that yields the largest number of non-zero unique TGIs.
    4) If header-based lookup fails, fallback to parsing the tail of the file as table
       with the same candidate entry sizes and keep the best result.
    """
    def _parse_table(table_bytes, count_hint=None, cancel_event=None):
        best = []
        best_len = 0
        candidates = (16, 24, 28, 32, 36, 40)
        tgis_local = []
        # If count hinted, use that; otherwise, deduce from length/entry_size
        for es in candidates:
            if cancel_event is not None and getattr(cancel_event, 'is_set', lambda: False)():
                break
            tgis_local = []
            if count_hint and count_hint > 0:
                n = min(count_hint, len(table_bytes) // es)
            else:
                n = len(table_bytes) // es
            if n <= 0:
                continue
            ok = 0
            for i in range(n):
                if cancel_event is not None and getattr(cancel_event, 'is_set', lambda: False)():
                    break
                off = i * es
                seg = table_bytes[off:off+16]
                if len(seg) < 16:
                    break
                t = int.from_bytes(seg[0:4], 'little', signed=False)
                g = int.from_bytes(seg[4:8], 'little', signed=False)
                ih = int.from_bytes(seg[8:12], 'little', signed=False)
                il = int.from_bytes(seg[12:16], 'little', signed=False)
                inst = (ih << 32) | il
                if t == 0 and g == 0 and inst == 0:
                    continue
                ok += 1
                tgis_local.append((t, g, inst))
            # Prefer the parse with the most entries
            if ok > best_len:
                best_len = ok
                best = tgis_local
        return best

    def _scan_tail_index(file_path, file_size, cancel_event=None):
        # Heuristic scan of tail for patterns: [T(4) G(4) I(8) Offset(4) Size(4) ...]
        # Validate offset/size within file to reduce false positives.
        max_tail = min(8 * 1024 * 1024, file_size)
        try:
            with open(file_path, 'rb') as fh:
                fh.seek(file_size - max_tail)
                buf = fh.read(max_tail)
        except Exception:
            return []
        out = []
        # Step by 4 for alignment
        limit = max(0, len(buf) - (16 + 8))
        for pos in range(0, limit, 4):
            if cancel_event is not None and getattr(cancel_event, 'is_set', lambda: False)():
                break
            seg = buf[pos:pos+16]
            t = int.from_bytes(seg[0:4], 'little', signed=False)
            g = int.from_bytes(seg[4:8], 'little', signed=False)
            ih = int.from_bytes(seg[8:12], 'little', signed=False)
            il = int.from_bytes(seg[12:16], 'little', signed=False)
            inst = (ih << 32) | il
            if (t | g | inst) == 0:
                continue
            off = int.from_bytes(buf[pos+16:pos+20], 'little', signed=False)
            size = int.from_bytes(buf[pos+20:pos+24], 'little', signed=False)
            if size <= 0:
                continue
            if off <= 0 or off >= file_size:
                continue
            if off + size > file_size:
                continue
            out.append((t, g, inst))
        # Deduplicate while preserving order
        seen = set()
        uniq = []
        for tgi in out:
            if tgi in seen:
                continue
            seen.add(tgi)
            uniq.append(tgi)
        return uniq

    try:
        if logger:
            try:
                logger.debug("DBPF: begin parse %s (fallback=%s)", os.path.basename(path), bool(allow_tail_fallback))
            except Exception:
                pass
        with open(path, 'rb') as fh:
            head = fh.read(128)
        if len(head) < 36 or head[:4] != b'DBPF':
            return []

        # Try several header layouts (offsets gathered from various DBPF docs/implementations)
        header_candidates = [
            # (count_off, index_off_off, size_off)
            (0x1C, 0x20, 0x24),
            (0x20, 0x24, 0x28),
            (0x24, 0x28, 0x2C),
            (0x28, 0x2C, 0x30),
            (0x2C, 0x30, 0x34),
            (0x30, 0x34, 0x38),
            (0x34, 0x38, 0x3C),
        ]

        with open(path, 'rb') as fh:
            fh.seek(0, os.SEEK_END)
            file_size = fh.tell()

        collected = []
        for c_off, o_off, s_off in header_candidates:
            count = _read_le_u32(head, c_off)
            off = _read_le_u32(head, o_off)
            size = _read_le_u32(head, s_off)
            if not (0 < off < file_size and 0 < size <= file_size - off):
                continue
            # Read table
            with open(path, 'rb') as fh:
                fh.seek(off)
                table = fh.read(size)
            if not table:
                continue
            tgis_local = _parse_table(table, count_hint=count, cancel_event=cancel_event)
            if tgis_local:
                collected.append((len(tgis_local), tgis_local))
        if collected:
            collected.sort(key=lambda x: x[0], reverse=True)
            out = collected[0][1]
            if logger:
                try:
                    logger.debug("DBPF: header index parsed -> %d entries", len(out))
                except Exception:
                    pass
            return out

        if not allow_tail_fallback:
            return []
        # Fallback A: parse tail region as generic table
        tail_size = min(2 * 1024 * 1024, file_size)
        with open(path, 'rb') as fh:
            fh.seek(file_size - tail_size)
            tail = fh.read(tail_size)
        best = _parse_table(tail, count_hint=None, cancel_event=cancel_event)
        if best:
            if logger:
                try:
                    logger.debug("DBPF: tail table parsed -> %d entries", len(best))
                except Exception:
                    pass
            return best
        # Fallback B: heuristic scan using offset/size plausibility
        out = _scan_tail_index(path, file_size, cancel_event=cancel_event)
        if logger:
            try:
                logger.debug("DBPF: tail heuristic parsed -> %d entries", len(out))
            except Exception:
                pass
        return out
    except Exception as exc:
        if logger:
            try:
                logger.debug("DBPF parse failed for %s: %s", path, exc)
            except Exception:
                pass
        return []


class IDConflictViewerDialog(QtWidgets.QDialog):
    def __init__(self, parent, mods_root):
        super().__init__(parent)
        self.setWindowTitle("ID Conflict Viewer")
        self.setModal(True)
        try:
            self.setWindowFlags(self.windowFlags() | QtCore.Qt.Window)
            self.setSizeGripEnabled(True)
        except Exception:
            pass
        self.resize(900, 600)
        self.parent_app = parent
        self.mods_root = os.path.abspath(mods_root)
        self.logger = logging.getLogger("Sims4ModTool")
        try:
            self.logger.debug("IDConflictViewerDialog.__init__(mods_root=%s)", self.mods_root)
        except Exception:
            pass

        layout = QtWidgets.QVBoxLayout(self)
        # Source folder selection
        src_row = QtWidgets.QHBoxLayout()
        src_row.addWidget(QtWidgets.QLabel("Dossier à analyser:", self))
        self.path_edit = QtWidgets.QLineEdit(self)
        try:
            self.path_edit.setText(str(getattr(self.parent_app, 'settings', {}).get('mod_directory', '') or ''))
        except Exception:
            pass
        src_row.addWidget(self.path_edit, 1)
        btn_browse_src = QtWidgets.QPushButton("Parcourir…", self)
        def _browse_src():
            d = QtWidgets.QFileDialog.getExistingDirectory(self, "Choisir un dossier", self.path_edit.text() or os.getcwd())
            if d:
                self.path_edit.setText(d)
        btn_browse_src.clicked.connect(_browse_src)
        src_row.addWidget(btn_browse_src)
        layout.addLayout(src_row)
        info = QtWidgets.QLabel(
            "Analyse les .package et détecte les conflits d'ID de ressource (Type-Group-Instance).\n"
            "Remarque: l'analyse DBPF est expérimentale; les erreurs sont ignorées et journalisées.",
            self,
        )
        info.setWordWrap(True)
        layout.addWidget(info)

        # Controls
        ctl = QtWidgets.QHBoxLayout()
        ctl.addWidget(QtWidgets.QLabel("Dossier:", self))
        self.path_edit = QtWidgets.QLineEdit(self)
        self.path_edit.setText(self.mods_root)
        self.path_edit.setReadOnly(True)
        ctl.addWidget(self.path_edit, stretch=1)
        self.recursive_checkbox = QtWidgets.QCheckBox("Récursif", self)
        self.recursive_checkbox.setChecked(True)
        ctl.addWidget(self.recursive_checkbox)
        self.use_scan_cache_checkbox = QtWidgets.QCheckBox("Utiliser cache fichiers", self)
        self.use_scan_cache_checkbox.setChecked(True)
        ctl.addWidget(self.use_scan_cache_checkbox)
        self.fast_mode_checkbox = QtWidgets.QCheckBox("Mode rapide (sans fallback)", self)
        self.fast_mode_checkbox.setChecked(False)
        ctl.addWidget(self.fast_mode_checkbox)
        self.scan_btn = QtWidgets.QPushButton("Analyser", self)
        self.scan_btn.clicked.connect(self._run_scan)
        ctl.addWidget(self.scan_btn)
        self.export_btn = QtWidgets.QPushButton("Exporter vers Excel", self)
        self.export_btn.clicked.connect(self._export_excel)
        ctl.addWidget(self.export_btn)
        layout.addLayout(ctl)

        self.progress = QtWidgets.QProgressBar(self)
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        # Results tree
        self.tree = QtWidgets.QTreeWidget(self)
        self.tree.setHeaderLabels(["Ressource (T:G:I)", "Conflits", "Type hex", "Group hex", "Instance hex"])
        try:
            hdr = self.tree.header()
            hdr.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
            hdr.setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
            hdr.setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
            hdr.setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
            hdr.setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)
        except Exception:
            pass
        self.tree.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self._context_menu)
        layout.addWidget(self.tree, stretch=1)

        # Footer
        footer = QtWidgets.QHBoxLayout()
        self.status_label = QtWidgets.QLabel("", self)
        footer.addWidget(self.status_label)
        footer.addStretch(1)
        self.stop_btn = QtWidgets.QPushButton("Stop", self)
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self._request_cancel)
        footer.addWidget(self.stop_btn)
        close_btn = QtWidgets.QPushButton("Fermer", self)
        close_btn.clicked.connect(self.accept)
        footer.addWidget(close_btn)
        layout.addLayout(footer)

        # Cancellation flag
        import threading as _threading
        self._cancel_event = _threading.Event()

        QtCore.QTimer.singleShot(0, self._run_scan)

    def _request_cancel(self):
        try:
            self.logger.info("IDConflictViewer: cancellation requested by user")
        except Exception:
            pass
        try:
            self.status_label.setText("Annulation demandée…")
        except Exception:
            pass
        try:
            self._cancel_event.set()
        except Exception:
            pass

    def _load_id_index_cache(self):
        try:
            with open(ID_INDEX_CACHE_PATH, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
            if not isinstance(data, dict):
                return {}
            return data
        except Exception:
            return {}

    def _save_id_index_cache(self, cache):
        try:
            with open(ID_INDEX_CACHE_PATH, 'w', encoding='utf-8') as fh:
                json.dump(cache, fh, ensure_ascii=False)
        except Exception:
            pass

    def _iter_packages(self, root, recursive=True):
        if not recursive:
            try:
                for f in os.listdir(root):
                    full = os.path.join(root, f)
                    if os.path.isfile(full) and f.lower().endswith('.package'):
                        yield full
            except OSError:
                return
            return
        for cur, _dirs, files in os.walk(root):
            for f in files:
                if f.lower().endswith('.package'):
                    yield os.path.join(cur, f)

    def _list_packages_from_scan_cache(self, root):
        # Use mod_scan_cache.json if it matches the same root
        try:
            with open(MOD_SCAN_CACHE_PATH, 'r', encoding='utf-8') as fh:
                cache = json.load(fh)
            cache_root = str(cache.get('root') or '')
            if not cache_root:
                try:
                    self.logger.debug("IDConflictViewer: scan cache missing 'root'")
                except Exception:
                    pass
                return None
            # Normalize compare
            if os.path.normcase(os.path.normpath(cache_root)) != os.path.normcase(os.path.normpath(root)):
                try:
                    self.logger.debug("IDConflictViewer: scan cache root mismatch: cache=%s current=%s", cache_root, root)
                except Exception:
                    pass
                return None
            entries = cache.get('entries') or []
            out = []
            for e in entries:
                try:
                    if str(e.get('type') or '').lower() != 'package':
                        continue
                    rel = e.get('path') or ''
                    if not rel:
                        continue
                    full = os.path.join(root, rel)
                    if os.path.isfile(full):
                        out.append(full)
                except Exception:
                    continue
            try:
                self.logger.debug("IDConflictViewer: scan cache yielded %d package files", len(out))
            except Exception:
                pass
            return out
        except Exception:
            try:
                self.logger.debug("IDConflictViewer: unable to read mod_scan_cache.json; fallback to os.walk")
            except Exception:
                pass
            return None

    def _run_scan(self):
        # Prepare UI state
        try:
            self._cancel_event.clear()
            self.scan_btn.setEnabled(False)
            self.stop_btn.setEnabled(True)
        except Exception:
            pass
        directory = self.mods_root
        if not directory or not os.path.isdir(directory):
            QtWidgets.QMessageBox.warning(self, "Dossier invalide", "Définis un dossier des mods valide.")
            return
        recursive = bool(self.recursive_checkbox.isChecked())
        self.tree.clear()
        self.status_label.setText("Analyse en cours…")
        self.progress.setVisible(True)
        files = []
        if bool(self.use_scan_cache_checkbox.isChecked()):
            files = self._list_packages_from_scan_cache(directory) or []
        if not files:
            files = list(self._iter_packages(directory, recursive=recursive))
        if self._cancel_event.is_set():
            self._end_scan(cancelled=True, stats=None)
            return
        try:
            self.logger.debug("IDConflictViewer: starting scan; files=%d, recursive=%s, use_scan_cache=%s, fast_mode=%s", len(files), recursive, bool(self.use_scan_cache_checkbox.isChecked()), bool(self.fast_mode_checkbox.isChecked()))
        except Exception:
            pass
        total = len(files)
        self.progress.setMinimum(0)
        self.progress.setMaximum(max(1, total))
        self.progress.setValue(0)

        conflicts = {}  # (t,g,i) -> set(paths)
        total_entries = 0
        parsed_files = 0
        cache = self._load_id_index_cache()
        start_time = datetime.now()

        # Worker for parallel parse without writing shared state
        def _parse_one(p):
            result = {
                'file': p,
                'tgis': [],
                'elapsed': 0.0,
                'cache_used': False,
                'allow_tail': not bool(self.fast_mode_checkbox.isChecked()),
                'cache_key': None,
            }
            try:
                if self._cancel_event.is_set():
                    return result
                try:
                    st = os.stat(p)
                    key = f"{p}|{st.st_size}|{int(st.st_mtime)}"
                except Exception:
                    key = p
                result['cache_key'] = key
                entry = cache.get(key) if isinstance(cache, dict) else None
                if entry and isinstance(entry, dict) and isinstance(entry.get('tgis'), list):
                    result['tgis'] = [(int(t), int(g), int(i)) for t, g, i in entry['tgis']]
                    result['cache_used'] = True
                    return result
                t0 = datetime.now()
                tgis = _read_tgi_entries_from_dbpf(p, logger=self.logger, allow_tail_fallback=result['allow_tail'], cancel_event=self._cancel_event)
                if (not tgis) and (result['allow_tail'] is False):
                    # retry once with fallback
                    if not self._cancel_event.is_set():
                        tgis = _read_tgi_entries_from_dbpf(p, logger=self.logger, allow_tail_fallback=True, cancel_event=self._cancel_event)
                result['elapsed'] = (datetime.now() - t0).total_seconds()
                result['tgis'] = tgis or []
                return result
            except Exception:
                return result

        processed = 0
        new_cache_entries = []
        max_workers = max(2, min(8, (os.cpu_count() or 4)))
        cancelled = False
        if total <= 4:
            # Small: run inline
            for p in files:
                if self._cancel_event.is_set():
                    cancelled = True
                    break
                r = _parse_one(p)
                tgis = r['tgis']
                for t, g, i in tgis:
                    s = conflicts.get((t, g, i))
                    if s is None:
                        s = set()
                        conflicts[(t, g, i)] = s
                    s.add(p)
                if tgis:
                    parsed_files += 1
                    total_entries += len(tgis)
                if not r['cache_used'] and r['cache_key'] and tgis is not None:
                    new_cache_entries.append((r['cache_key'], tgis))
                processed += 1
                if processed % 5 == 0 or processed == total:
                    self.progress.setValue(processed)
                    QtWidgets.QApplication.processEvents()
        else:
            # Parallel parse
            from concurrent.futures import ThreadPoolExecutor, as_completed
            with ThreadPoolExecutor(max_workers=max_workers) as ex:
                futs = [ex.submit(_parse_one, p) for p in files]
                for fut in as_completed(futs):
                    if self._cancel_event.is_set():
                        cancelled = True
                        # Cancel remaining futures (pending ones only)
                        for f in futs:
                            f.cancel()
                        break
                    r = fut.result()
                    p = r.get('file')
                    tgis = r.get('tgis') or []
                    for t, g, i in tgis:
                        s = conflicts.get((t, g, i))
                        if s is None:
                            s = set()
                            conflicts[(t, g, i)] = s
                        s.add(p)
                    if tgis:
                        parsed_files += 1
                        total_entries += len(tgis)
                    if (not r.get('cache_used')) and r.get('cache_key') and tgis is not None:
                        new_cache_entries.append((r['cache_key'], tgis))
                    processed += 1
                    if processed % 5 == 0 or processed == total:
                        self.progress.setValue(processed)
                        QtWidgets.QApplication.processEvents()
        # Save cache updates at once
        try:
            for key, tgis in new_cache_entries:
                cache[key] = {"tgis": [[t, g, i] for (t, g, i) in tgis]}
            self._save_id_index_cache(cache)
        except Exception:
            pass

        # Render conflicts only
        count_conflicts = 0
        for (t, g, i), paths in sorted(conflicts.items(), key=lambda kv: (kv[0][0], kv[0][1], kv[0][2])):
            if len(paths) < 2:
                continue
            count_conflicts += 1
            t_hex = f"0x{t:08X}"
            g_hex = f"0x{g:08X}"
            i_hex = f"0x{i:016X}"
            key_text = f"{t_hex}:{g_hex}:{i_hex}"
            top = QtWidgets.QTreeWidgetItem([key_text, str(len(paths)), t_hex, g_hex, i_hex])
            self.tree.addTopLevelItem(top)
            for p in sorted(paths, key=str.casefold):
                try:
                    ts = os.path.getmtime(p)
                    dt = format_datetime(datetime.fromtimestamp(ts))
                except Exception:
                    dt = ""
                child = QtWidgets.QTreeWidgetItem([os.path.basename(p), "", "", p, dt])
                # store path in UserRole
                child.setData(0, QtCore.Qt.UserRole, p)
                top.addChild(child)
            top.setExpanded(True)

        # Provide AI group rebuild within Group View
    def _rebuild_groups_from_ai(self):
        parent_app = self.parent() if hasattr(self, 'parent') else None
        if parent_app is None or not hasattr(parent_app, 'settings'):
            return
        settings = parent_app.settings
        mod_dir = str(settings.get('mod_directory', '') or '')
        if not (mod_dir and os.path.isdir(mod_dir)):
            QtWidgets.QMessageBox.information(self, "Dossier invalide", "Configure un dossier de mods valide dans la Configuration.")
            return
        # Collect results from logs under Mods
        exts = {'.log', '.txt', '.html', '.htm'}
        out = []
        for rootd, _dirs, files in os.walk(mod_dir):
            for fn in files:
                lf = fn.lower()
                if os.path.splitext(lf)[1] in exts:
                    path = os.path.join(rootd, fn)
                    try:
                        content = open(path, 'r', encoding='utf-8', errors='replace').read()
                    except Exception:
                        continue
                    if lf.endswith('.html') or lf.endswith('.htm'):
                        parsed = analyze_last_exception_html(content)
                        results = list(parsed.get('results') or [])
                        text = _strip_html_to_text(content)
                        results += analyze_generic_log_text(text)
                    else:
                        results = analyze_generic_log_text(content)
                    # Dedup by (type, message, first path)
                    seen = set((it.get('type'), it.get('message'), (it.get('paths') or [None])[0]) for it in out)
                    for it in results:
                        key = (it.get('type'), it.get('message'), (it.get('paths') or [None])[0])
                        if key not in seen:
                            out.append(it)
                            seen.add(key)
        if not out:
            QtWidgets.QMessageBox.information(self, "Aucune donnée", "Aucun log exploitable trouvé sous Mods.")
            return
        # Ensure AI available
        if getattr(parent_app, 'mod_ai', None) is None:
            try:
                parent_app.mod_ai = ModAI.load(str(settings.get('ai_model_path', 'mod_ai.json')))
            except Exception:
                parent_app.mod_ai = None
        ai = getattr(parent_app, 'mod_ai', None)
        overrides = dict(settings.get('ai_group_overrides', {}) or {})
        updated = 0
        for item in out:
            paths = list(item.get('paths') or [])
            text = f"{item.get('type') or ''} {item.get('message') or ''}"
            guess = ''
            conf_val = 0.0
            if ai is not None:
                m, conf = ai.guess_from_paths_and_text(paths, text)
                guess = (m or '').strip()
                try:
                    conf_val = float(conf)
                except Exception:
                    conf_val = 0.0
            if not guess:
                for p in paths:
                    norm = ('/' + str(p).replace('\\', '/').lstrip('/'))
                    m = re.search(r'(?i)/mods/([^/]+)/', norm)
                    if m:
                        guess = m.group(1)
                        conf_val = 1.0
                        break
            if not guess or conf_val < 0.99:
                continue
            for p in paths:
                base = os.path.basename(str(p) or '')
                if not base:
                    continue
                n = normalize_mod_basename(base)
                if n and not overrides.get(n):
                    overrides[n] = guess
                    updated += 1
        if updated:
            settings['ai_group_overrides'] = overrides
            save_settings(settings)
            try:
                parent_app.refresh_table_only()
            except Exception:
                pass
            QtWidgets.QMessageBox.information(self, "Groupes (AI)", f"{updated} associations ajoutées. La table principale a été rafraîchie.")
        else:
            QtWidgets.QMessageBox.information(self, "Groupes (AI)", "Aucune association supplémentaire détectée.")
    def _rebuild_groups_from_ai(self):
        parent_app = self.parent() if hasattr(self, 'parent') else None
        if parent_app is None or not hasattr(parent_app, 'settings'):
            return
        settings = parent_app.settings
        mod_dir = str(settings.get('mod_directory', '') or '')
        if not (mod_dir and os.path.isdir(mod_dir)):
            QtWidgets.QMessageBox.information(self, "Dossier invalide", "Configure un dossier de mods valide dans la Configuration.")
            return
        # Collect results from logs under Mods
        exts = {'.log', '.txt', '.html', '.htm'}
        out = []
        for rootd, _dirs, files in os.walk(mod_dir):
            for fn in files:
                lf = fn.lower()
                if os.path.splitext(lf)[1] in exts:
                    path = os.path.join(rootd, fn)
                    try:
                        content = open(path, 'r', encoding='utf-8', errors='replace').read()
                    except Exception:
                        continue
                    if lf.endswith('.html') or lf.endswith('.htm'):
                        parsed = analyze_last_exception_html(content)
                        results = list(parsed.get('results') or [])
                        text = _strip_html_to_text(content)
                        results += analyze_generic_log_text(text)
                    else:
                        results = analyze_generic_log_text(content)
                    # Dedup by (type, message, first path)
                    seen = set((it.get('type'), it.get('message'), (it.get('paths') or [None])[0]) for it in out)
                    for it in results:
                        key = (it.get('type'), it.get('message'), (it.get('paths') or [None])[0])
                        if key not in seen:
                            out.append(it)
                            seen.add(key)
        if not out:
            QtWidgets.QMessageBox.information(self, "Aucune donnée", "Aucun log exploitable trouvé sous Mods.")
            return
        # Ensure AI available
        if getattr(parent_app, 'mod_ai', None) is None:
            try:
                parent_app.mod_ai = ModAI.load(str(settings.get('ai_model_path', 'mod_ai.json')))
            except Exception:
                parent_app.mod_ai = None
        ai = getattr(parent_app, 'mod_ai', None)
        overrides = dict(settings.get('ai_group_overrides', {}) or {})
        updated = 0
        for item in out:
            paths = list(item.get('paths') or [])
            text = f"{item.get('type') or ''} {item.get('message') or ''}"
            guess = ''
            conf_val = 0.0
            if ai is not None:
                m, conf = ai.guess_from_paths_and_text(paths, text)
                guess = (m or '').strip()
                try:
                    conf_val = float(conf)
                except Exception:
                    conf_val = 0.0
            if not guess:
                for p in paths:
                    norm = ('/' + str(p).replace('\\', '/').lstrip('/'))
                    m = re.search(r'(?i)/mods/([^/]+)/', norm)
                    if m:
                        guess = m.group(1)
                        conf_val = 1.0
                        break
            if not guess or conf_val < 0.99:
                continue
            for p in paths:
                base = os.path.basename(str(p) or '')
                if not base:
                    continue
                n = normalize_mod_basename(base)
                if n and not overrides.get(n):
                    overrides[n] = guess
                    updated += 1
        if updated:
            settings['ai_group_overrides'] = overrides
            save_settings(settings)
            try:
                parent_app.refresh_table_only()
            except Exception:
                pass
            QtWidgets.QMessageBox.information(self, "Groupes (AI)", f"{updated} associations ajoutées. La table principale a été rafraîchie.")
        else:
            QtWidgets.QMessageBox.information(self, "Groupes (AI)", "Aucune association supplémentaire détectée.")



    def _end_scan(self, cancelled=False, stats=None):
        try:
            self.progress.setVisible(False)
            self.scan_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)
        except Exception:
            pass
        if cancelled:
            try:
                self.status_label.setText("Analyse annulée")
            except Exception:
                pass
            return
        if stats:
            try:
                txt = f"{stats['count_conflicts']} ressource(s) en conflit • {stats['parsed_files']}/{stats['total']} fichiers parsés • {stats['total_entries']} ressources • {stats['elapsed']:.1f}s"
                self.status_label.setText(txt)
            except Exception:
                pass

    def _selected_file_paths(self):
        paths = []
        for item in self.tree.selectedItems():
            # Child rows store path in UserRole on column 0; top rows do not
            p = item.data(0, QtCore.Qt.UserRole)
            if p:
                paths.append(str(p))
        return paths

    def _context_menu(self, pos):
        item = self.tree.itemAt(pos)
        if item is None:
            return
        menu = QtWidgets.QMenu(self)
        act_open = menu.addAction("Ouvrir le dossier")
        act_prefix = menu.addAction("Préfixer 'zzz_' le fichier")
        act_disable = menu.addAction("Désactiver le dossier du mod")
        selected = menu.exec_(self.tree.viewport().mapToGlobal(pos))
        if selected == act_open:
            p = item.data(0, QtCore.Qt.UserRole)
            target = os.path.dirname(p) if p else None
            if target and os.path.isdir(target):
                if self.parent_app and hasattr(self.parent_app, "_open_in_file_manager"):
                    try:
                        self.logger.debug("IDConflictViewer: open folder %s", target)
                    except Exception:
                        pass
                    self.parent_app._open_in_file_manager(target)
        elif selected == act_prefix:
            p = item.data(0, QtCore.Qt.UserRole)
            if not p or not os.path.isfile(p):
                return
            d = os.path.dirname(p)
            b = os.path.basename(p)
            if b.lower().startswith('zzz_'):
                return
            candidate = os.path.join(d, 'zzz_' + b)
            try:
                os.rename(p, candidate)
                self.logger.info("IDConflictViewer: renommé %s -> %s", p, candidate)
            except OSError as exc:
                QtWidgets.QMessageBox.critical(self, "Erreur", f"Impossible de renommer: {exc}")
                return
            self._run_scan()
        elif selected == act_disable:
            p = item.data(0, QtCore.Qt.UserRole)
            if not p:
                return
            mod_folder = os.path.dirname(p)
            backups_dir = self.parent_app.settings.get("backups_directory", "") if self.parent_app else ""
            if not backups_dir:
                QtWidgets.QMessageBox.warning(self, "Backups manquant", "Définis un dossier de backups dans la configuration.")
                return
            disabled_root = os.path.join(backups_dir, "Disabled Mod")
            os.makedirs(disabled_root, exist_ok=True)
            dest = os.path.join(disabled_root, os.path.basename(mod_folder))
            final = dest
            i = 1
            while os.path.exists(final):
                final = f"{dest}_{i}"
                i += 1
            try:
                shutil.move(mod_folder, final)
                self.logger.info("IDConflictViewer: désactivé %s -> %s", mod_folder, final)
            except OSError as exc:
                QtWidgets.QMessageBox.critical(self, "Erreur", f"Impossible de désactiver: {exc}")
                return
            self._run_scan()

    def _export_excel(self):
        # Build flat rows from current tree
        rows = []
        for i in range(self.tree.topLevelItemCount()):
            top = self.tree.topLevelItem(i)
            key_text = top.text(0)
            t_hex = top.text(2)
            g_hex = top.text(3)
            i_hex = top.text(4)
            for j in range(top.childCount()):
                ch = top.child(j)
                rows.append({
                    'resource': key_text,
                    'type': t_hex,
                    'group': g_hex,
                    'instance': i_hex,
                    'file': ch.text(0),
                    'path': ch.text(3),
                    'date': ch.text(4),
                })
        if not rows:
            QtWidgets.QMessageBox.information(self, "Export", "Aucun conflit à exporter.")
            return
        # Choose save path
        suggested = os.path.join(os.getcwd(), "id_conflicts.xlsx")
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Exporter vers Excel", suggested, "Fichiers Excel (*.xlsx)")
        if not path:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "ID Conflicts"
            headers = ["resource", "type", "group", "instance", "file", "path", "date"]
            ws.append(headers)
            for r in rows:
                ws.append([r[h] for h in headers])
            wb.save(path)
            QtWidgets.QMessageBox.information(self, "Export", f"Exporté: {path}")
            try:
                self.logger.info("IDConflictViewer: export Excel -> %s (%d lignes)", path, len(rows))
            except Exception:
                pass
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Erreur", f"Échec export Excel: {exc}")


class ModManagerApp(QtWidgets.QWidget):
    def __init__(self, splash=None):
        super().__init__()
        self.setObjectName("root")
        try:
            self.setAutoFillBackground(True)
            self.setAttribute(QtCore.Qt.WA_StyledBackground, True)
        except Exception:
            pass

        self.setWindowTitle(f"Gestionnaire de Mods Sims 4 – {APP_VERSION} ({APP_VERSION_DATE})")
        self.setGeometry(100, 100, 800, 600)

        # Initialize settings and logging (default DEBUG, overridable by settings)
        self.settings = load_settings()
        self.logger = setup_logging(self.settings.get("log_level", "DEBUG"))
        try:
            self.logger.debug("Application starting with log level: %s", self.settings.get("log_level", "DEBUG"))
        except Exception:
            pass
        # Log cache metadata at startup
        try:
            if os.path.exists(MOD_SCAN_CACHE_PATH):
                mtime = datetime.fromtimestamp(os.path.getmtime(MOD_SCAN_CACHE_PATH))
                size = os.path.getsize(MOD_SCAN_CACHE_PATH)
                cache = load_mod_scan_cache()
                entries = len(cache.get("entries", [])) if cache else 0
                self.logger.info(
                    "mod_scan_cache.json: mtime=%s, size=%d bytes, entries=%d",
                    mtime.strftime("%Y-%m-%d %H:%M:%S"), size, entries
                )
            else:
                self.logger.info("mod_scan_cache.json: not found")
        except Exception:
            pass
        # Apply background as early as possible
        try:
            self._apply_background()
        except Exception:
            pass
        self.custom_version_releases = load_custom_version_releases()
        self.version_releases = merge_version_releases(self.custom_version_releases)
        self.ignored_mods = set(self.settings.get("ignored_mods", []))
        self.last_scanned_directory = ""
        self.all_data_rows = []
        self._cache_clear_triggered_this_refresh = False
        # Background scan state
        self._scan_thread = None
        self._scan_worker = None
        self._scan_request_id = 0

        self._splash = splash
        self.init_ui()

        auto = bool(self.settings.get("auto_scan_on_start", True))
        if auto or not os.path.exists(MOD_SCAN_CACHE_PATH):
            mod_directory = self.settings.get("mod_directory", "")
            if mod_directory and os.path.isdir(mod_directory):
                QtCore.QTimer.singleShot(300, self.refresh_tree)
        else:
            try:
                self._populate_from_cache()
            except Exception:
                # If cache load fails, keep UI idle without crashing
                pass

        # Start web interface (Flask) in background if enabled
        try:
            if bool(self.settings.get("web_enabled", True)):
                QtCore.QTimer.singleShot(500, self._start_web_interface)
        except Exception:
            pass

        # Initialize cache reload debounce timer (UI thread)
        try:
            self._cache_watch_timer = QtCore.QTimer(self)
            self._cache_watch_timer.setSingleShot(True)
            self._cache_watch_timer.setInterval(300)
            self._cache_watch_timer.timeout.connect(self._on_cache_reload_timeout)
        except Exception:
            self._cache_watch_timer = None

        # Start cache watchdog (monitor mod_scan_cache.json updates)
        try:
            self._start_cache_watchdog()
        except Exception:
            pass

    def _yield_ui_events(self, max_time_ms=25):
        try:
            flags_type = QtCore.QEventLoop.ProcessEventsFlag
            flags = flags_type.ExcludeUserInputEvents | flags_type.ExcludeSocketNotifiers
        except AttributeError:
            flags = QtCore.QEventLoop.ExcludeUserInputEvents | QtCore.QEventLoop.ExcludeSocketNotifiers
        app = QtWidgets.QApplication.instance()
        if app is None:
            return
        app.processEvents(flags, max_time_ms)

    # --- Cache watchdog: update table when mod_scan_cache.json changes ---
    def _start_cache_watchdog(self):
        self._cache_observer = None
        if Observer is None:
            return
        cache_path = os.path.abspath(MOD_SCAN_CACHE_PATH)
        watch_dir = os.path.dirname(cache_path) or os.getcwd()
        app_ref = self

        class _CacheHandler(FileSystemEventHandler):
            def on_modified(self, event):
                if getattr(event, 'is_directory', False):
                    return
                if os.path.abspath(getattr(event, 'src_path', '')) != cache_path:
                    return
                # Post to UI thread to avoid cross-thread QObject/timer usage
                try:
                    QtCore.QTimer.singleShot(0, app_ref._schedule_cache_reload)
                except Exception:
                    pass

            def on_created(self, event):
                self.on_modified(event)

        self._cache_handler = _CacheHandler()
        observer = Observer()
        observer.schedule(self._cache_handler, watch_dir, recursive=False)
        observer.daemon = True
        observer.start()
        self._cache_observer = observer
        try:
            if hasattr(self, 'logger'):
                self.logger.info("Cache watchdog started on %s", cache_path)
        except Exception:
            pass

    def _schedule_cache_reload(self):
        # Debounce successive events; this runs on the UI thread
        try:
            if self._cache_watch_timer is not None:
                self._cache_watch_timer.start()
            else:
                # Fallback: immediate reload
                self._populate_from_cache()
        except Exception:
            # Last resort: async call on UI
            QtCore.QTimer.singleShot(0, self._populate_from_cache)

    def _on_cache_reload_timeout(self):
        try:
            if hasattr(self, 'logger'):
                self.logger.info("Cache modified: reloading table from cache")
        except Exception:
            pass
        try:
            self._populate_from_cache()
        finally:
            self._update_scan_status("")

    def _stop_cache_watchdog(self):
        try:
            if getattr(self, '_cache_observer', None) is not None:
                self._cache_observer.stop()
                self._cache_observer.join(timeout=1.0)
                self._cache_observer = None
        except Exception:
            pass

    def closeEvent(self, event):
        try:
            self._stop_cache_watchdog()
        except Exception:
            pass
        try:
            super().closeEvent(event)
        except Exception:
            pass

    def init_ui(self):
        # Layout
        layout = QtWidgets.QVBoxLayout()

        # Mode sombre - Définir le style global
        self._base_stylesheet = self._compute_base_stylesheet()
        self.setStyleSheet(self._base_stylesheet)

        # Dossier des mods (affichage uniquement)
        mod_dir_layout = QtWidgets.QHBoxLayout()
        mod_dir_layout.addWidget(QtWidgets.QLabel("Dossier des mods :"))
        self.mod_directory_label = QtWidgets.QLabel()
        self.mod_directory_label.setTextInteractionFlags(QtCore.Qt.TextSelectableByMouse)
        mod_dir_layout.addWidget(self.mod_directory_label, stretch=1)

        layout.addLayout(mod_dir_layout)
        self.update_mod_directory_label()

        # Filtrage
        version_range_layout = QtWidgets.QVBoxLayout()
        self.version_filters_checkbox = QtWidgets.QCheckBox("Versions", self)
        self.version_filters_checkbox.setChecked(self.settings.get("enable_version_filters", True))
        self.version_filters_checkbox.toggled.connect(self._on_version_filters_toggled)
        version_row1 = QtWidgets.QHBoxLayout()
        version_row1.addWidget(self.version_filters_checkbox)

        self.version_start_label = QtWidgets.QLabel("Version de départ :", self)
        version_row1.addWidget(self.version_start_label)
        self.version_start_combo = QtWidgets.QComboBox(self)
        self.version_start_combo.setSizeAdjustPolicy(QtWidgets.QComboBox.AdjustToContents)
        version_row1.addWidget(self.version_start_combo)
        version_range_layout.addLayout(version_row1)
        self.version_end_label = QtWidgets.QLabel("Version d'arrivée :", self)
        version_row2 = QtWidgets.QHBoxLayout()
        version_row2.addWidget(self.version_end_label)
        self.version_end_combo = QtWidgets.QComboBox(self)
        self.version_end_combo.setSizeAdjustPolicy(QtWidgets.QComboBox.AdjustToContents)
        version_row2.addWidget(self.version_end_combo)
        version_range_layout.addLayout(version_row2)

        # Filter mode dropdown (replaces show package/ts4script checkboxes)
        self.file_filter_combo = QtWidgets.QComboBox(self)
        self.file_filter_combo.addItem("Show both", "both")
        self.file_filter_combo.addItem("Show Package", "package")
        self.file_filter_combo.addItem("Show TS4Script", "ts4script")
        self.file_filter_combo.addItem("Mod Installer Only", "installer_only")
        current_mode = str(self.settings.get("file_filter_mode", "both")).lower()
        idx = self.file_filter_combo.findData(current_mode)
        self.file_filter_combo.setCurrentIndex(idx if idx != -1 else 0)
        self.file_filter_combo.currentIndexChanged.connect(self.on_file_filter_mode_changed)

        self.show_ignored_checkbox = QtWidgets.QCheckBox("Afficher les mods ignorés", self)
        self.show_ignored_checkbox.setChecked(self.settings.get("show_ignored", False))
        self.show_ignored_checkbox.toggled.connect(lambda: self.toggle_show_ignored())
        # Group box for filters (top-left)
        filters_group = QtWidgets.QGroupBox("Filtres", self)
        filters_layout = QtWidgets.QVBoxLayout(filters_group)
        filters_layout.addLayout(version_range_layout)
        type_row = QtWidgets.QHBoxLayout()
        type_row.addWidget(QtWidgets.QLabel("Type :", self))
        type_row.addWidget(self.file_filter_combo)
        filters_layout.addLayout(type_row)
        # Row: Ignored + Hide Mod Installer side by side
        row_ignored = QtWidgets.QHBoxLayout()
        row_ignored.addWidget(self.show_ignored_checkbox)
        self.hide_installer_checkbox = QtWidgets.QCheckBox("Masquer Mod Installer", self)
        self.hide_installer_checkbox.setChecked(self.settings.get("hide_installer_mods", False))
        self.hide_installer_checkbox.toggled.connect(self.toggle_hide_installer)
        row_ignored.addWidget(self.hide_installer_checkbox)
        row_ignored.addStretch(1)
        filters_layout.addLayout(row_ignored)
        self.show_disabled_only_checkbox = QtWidgets.QCheckBox("Afficher seulement désactivés", self)
        self.show_disabled_only_checkbox.setChecked(self.settings.get("show_disabled_only", False))
        self.show_disabled_only_checkbox.toggled.connect(self.toggle_show_disabled_only)
        filters_layout.addWidget(self.show_disabled_only_checkbox)

        self.populate_version_combos()
        self.version_start_combo.currentIndexChanged.connect(self.on_version_filter_changed)
        self.version_end_combo.currentIndexChanged.connect(self.on_version_filter_changed)
        self._update_version_filter_visibility()

        # Top bar: Filters left, Actions right (buttons will be added after creation)
        top_bar = QtWidgets.QHBoxLayout()
        # Slightly reduce filters group width
        filters_group.setMaximumWidth(460)
        top_bar.addWidget(filters_group, stretch=1)
        self.actions_group = QtWidgets.QGroupBox("Actions", self)
        self.actions_layout = QtWidgets.QGridLayout(self.actions_group)
        try:
            self.actions_layout.setHorizontalSpacing(12)
            self.actions_layout.setVerticalSpacing(10)
            self.actions_layout.setContentsMargins(8, 8, 8, 8)
        except Exception:
            pass
        top_bar.addWidget(self.actions_group, stretch=1)
        # Tools group on the right (adaptive grid)
        self.tools_group = QtWidgets.QGroupBox("Tools", self)
        self.tools_layout = QtWidgets.QGridLayout(self.tools_group)
        tools_buttons = []
        btn_dup = QtWidgets.QPushButton("Find dupplicates", self.tools_group); btn_dup.clicked.connect(self.open_duplicate_finder); tools_buttons.append(btn_dup)
        btn_conflict = QtWidgets.QPushButton("Conflict Checker", self.tools_group); btn_conflict.clicked.connect(self.open_conflict_checker); tools_buttons.append(btn_conflict)
        btn_idconf = QtWidgets.QPushButton("ID Conflict Viewer", self.tools_group); btn_idconf.clicked.connect(self.open_id_conflict_viewer); tools_buttons.append(btn_idconf)
        btn_idconf_v2 = QtWidgets.QPushButton("ID Conflict V2", self.tools_group); btn_idconf_v2.clicked.connect(self.open_id_conflict_viewer_v2); tools_buttons.append(btn_idconf_v2)
        btn_nonmods = QtWidgets.QPushButton("Find non-mods files", self.tools_group); btn_nonmods.clicked.connect(partial(self._show_placeholder_message, "Find non-mods files", "La détection des fichiers non mods sera ajoutée ultérieurement.")); tools_buttons.append(btn_nonmods)
        btn_disable = QtWidgets.QPushButton("Disable all mods", self.tools_group); btn_disable.clicked.connect(partial(self._show_placeholder_message, "Disable all mods", "La désactivation des mods sera proposée dans une future mise à jour.")); tools_buttons.append(btn_disable)
        btn_cfg = QtWidgets.QPushButton("Correct resource.cfg", self.tools_group); btn_cfg.clicked.connect(self.correct_resource_cfg); tools_buttons.append(btn_cfg)
        btn_symlink = QtWidgets.QPushButton("Symlink Mods", self.tools_group); btn_symlink.clicked.connect(partial(self._show_placeholder_message, "Symlink Mods", "La création de liens symboliques vers le dossier Mods sera ajoutée ultérieurement.")); tools_buttons.append(btn_symlink)
        btn_backup = QtWidgets.QPushButton("Backup Mods", self.tools_group); btn_backup.clicked.connect(partial(self._show_placeholder_message, "Backup Mods", "La sauvegarde du dossier Mods sera disponible dans une prochaine version.")); tools_buttons.append(btn_backup)
        btn_check_ts4 = QtWidgets.QPushButton("Check placement .ts4script", self.tools_group); btn_check_ts4.clicked.connect(self.check_ts4script_placement); tools_buttons.append(btn_check_ts4)
        btn_scan_folder = QtWidgets.QPushButton("Scan dossier (mod)", self.tools_group); btn_scan_folder.clicked.connect(self.open_folder_scanner); tools_buttons.append(btn_scan_folder)
        btn_find_ts4 = QtWidgets.QPushButton("Find in ts4script", self.tools_group); btn_find_ts4.clicked.connect(self.open_find_in_ts4script); tools_buttons.append(btn_find_ts4)
        btn_compare = QtWidgets.QPushButton("Comparateur de mods", self.tools_group); btn_compare.clicked.connect(self.open_mod_comparator); tools_buttons.append(btn_compare)
        btn_updates = QtWidgets.QPushButton("Updates Checker", self.tools_group); btn_updates.clicked.connect(self.open_updates_checker); tools_buttons.append(btn_updates)
        btn_log_manager = QtWidgets.QPushButton("Log Manager", self.tools_group); btn_log_manager.clicked.connect(self.open_log_manager); tools_buttons.append(btn_log_manager)
        # Load external tools from ./modules (add-ons)
        try:
            tools_buttons.extend(self._load_external_tools_buttons())
        except Exception:
            pass
        # initial layout
        self._layout_tools_buttons(tools_buttons)
        top_bar.addWidget(self.tools_group, stretch=1)
        # reflow on resize (ensure handler returns None)
        try:
            def _tools_resize_event(e, b=tools_buttons, grp=self.tools_group):
                self._layout_tools_buttons(b)
                try:
                    QtWidgets.QGroupBox.resizeEvent(grp, e)
                except Exception:
                    pass
                return None
            self.tools_group.resizeEvent = _tools_resize_event
        except Exception:
            pass
        layout.addLayout(top_bar)

        # Search controls (two rows)
        self.search_edit = QtWidgets.QLineEdit(self)
        self.search_edit.setPlaceholderText("Nom du mod à rechercher")
        self.search_edit.textChanged.connect(self.apply_search_filter)
        self.show_search_checkbox = QtWidgets.QCheckBox("Afficher recherche", self)
        self.show_search_checkbox.setChecked(self.settings.get("show_search_results", True))
        self.show_search_checkbox.toggled.connect(self.toggle_show_search_results)
        self.instant_search_checkbox = QtWidgets.QCheckBox("Instant search", self)
        self.instant_search_checkbox.setChecked(self.settings.get("instant_search", True))
        self.instant_search_checkbox.toggled.connect(self.toggle_instant_search)
        self.search_button = QtWidgets.QPushButton("Rechercher", self)
        self.search_button.clicked.connect(partial(self.apply_search_filter, forced=True))

        search_row1 = QtWidgets.QHBoxLayout()
        search_row1.addWidget(self.show_search_checkbox)
        search_row1.addWidget(self.instant_search_checkbox)
        search_row1.addStretch(1)

        search_row2 = QtWidgets.QHBoxLayout()
        search_row2.addWidget(QtWidgets.QLabel("Recherche mod :"))
        search_row2.addWidget(self.search_edit)
        search_row2.addWidget(self.search_button)

        self.search_edit.setEnabled(self.show_search_checkbox.isChecked())

        # Move search controls into Filters group
        filters_layout.addLayout(search_row1)
        filters_layout.addLayout(search_row2)

        progress_layout = QtWidgets.QHBoxLayout()
        self.scan_status_label = QtWidgets.QLabel("", self)
        self.scan_status_label.setVisible(False)
        self.scan_progress_bar = QtWidgets.QProgressBar(self)
        self.scan_progress_bar.setVisible(False)
        self.scan_progress_bar.setMinimum(0)
        self.scan_progress_bar.setMaximum(1)
        self.scan_progress_bar.setValue(0)
        self.scan_count_label = QtWidgets.QLabel("", self)
        self.scan_count_label.setVisible(False)
        progress_layout.addWidget(self.scan_status_label)
        progress_layout.addWidget(self.scan_progress_bar, stretch=1)
        progress_layout.addWidget(self.scan_count_label)
        layout.addLayout(progress_layout)

        # Table des mods
        self.table = QtWidgets.QTableWidget(self)
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "État",
            "Mod (groupe)",
            "Fichier .package",
            "Date .package",
            "Fichier .ts4script",
            "Date .ts4script",
            "Version",
            "Confiance",
            "Installer",
            "Ignoré",
        ])
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.MultiSelection)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.DoubleClicked | QtWidgets.QAbstractItemView.EditKeyPressed)

        header = self.table.horizontalHeader()
        # Header context menu for column visibility
        header.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        header.customContextMenuRequested.connect(self._show_header_menu)
        for column in range(self.table.columnCount()):
            resize_mode = QtWidgets.QHeaderView.Stretch
            # Ajuster la largeur au contenu pour colonnes clés
            if column in (0, 1, 3, 5, 6, 7, 8, self.table.columnCount() - 1):
                resize_mode = QtWidgets.QHeaderView.ResizeToContents
            header.setSectionResizeMode(column, resize_mode)
        header.setStretchLastSection(False)
        self.table.setSortingEnabled(True)
        self.table.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        self.table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)

        layout.addWidget(self.table, stretch=1)

        # Apply hidden columns from settings
        self._apply_hidden_columns()

        # Boutons
        self.configuration_button = QtWidgets.QPushButton("Configuration", self)
        self.configuration_button.clicked.connect(self.open_configuration)

        # Actions group: use tool buttons with icon over text
        self.mod_installer_button = QtWidgets.QToolButton(self)
        self.mod_installer_button.setText("Mod Installer")
        self.mod_installer_button.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
        self.mod_installer_button.clicked.connect(self.open_mod_installer)

        self.refresh_button = QtWidgets.QToolButton(self)
        self.refresh_button.setText("Analyser / Rafraîchir")
        self.refresh_button.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
        self.refresh_button.clicked.connect(self.refresh_tree)

        self.export_button = QtWidgets.QToolButton(self)
        self.export_button.setText("Exporter vers Excel")
        self.export_button.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
        self.export_button.clicked.connect(self.export_current)

        self.clear_cache_button = QtWidgets.QToolButton(self)
        self.clear_cache_button.setText("Clear Sims4 Cache")
        self.clear_cache_button.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
        self.clear_cache_button.clicked.connect(self.clear_sims4_cache)

        self.grab_logs_button = QtWidgets.QToolButton(self)
        self.grab_logs_button.setText("Grab Logs")
        self.grab_logs_button.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
        self.grab_logs_button.clicked.connect(self.grab_logs)

        self.launch_button = QtWidgets.QPushButton("Launch Sims 4", self)
        self.launch_button.clicked.connect(self.launch_sims4)

        self.kill_button = QtWidgets.QPushButton("Kill Sims 4", self)
        self.kill_button.clicked.connect(self.kill_sims4)

        # Tools moved to dedicated group; keep dialog function for completeness but remove button from actions
        self.group_view_button = QtWidgets.QToolButton(self)
        self.group_view_button.setText("Group View")
        self.group_view_button.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
        self.group_view_button.clicked.connect(self.open_group_view)

        # Open Mods Folder
        self.open_mods_button = QtWidgets.QToolButton(self)
        self.open_mods_button.setText("Ouvrir Mods")
        self.open_mods_button.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
        self.open_mods_button.clicked.connect(self.open_mods_folder)

        # Populate Actions group (top-right) adaptively in a grid
        self._action_buttons = [
            # Put Refresh first as requested
            self.refresh_button,
            self.mod_installer_button,
            self.group_view_button,
            self.open_mods_button,
            self.export_button,
            self.clear_cache_button,
            self.grab_logs_button,
        ]
        self._layout_action_buttons()

        # Keep Configuration on the left, center game info, and show state/AI + Kill/Launch on the right
        bottom_buttons = QtWidgets.QHBoxLayout()
        # Bottom buttons: smaller height
        try:
            self.configuration_button.setMaximumHeight(32)
            self.kill_button.setMaximumHeight(32)
            self.launch_button.setMaximumHeight(32)
        except Exception:
            pass

        # Center label for game version/build
        self.game_info_label = QtWidgets.QLabel("", self)
        self.game_info_label.setAlignment(QtCore.Qt.AlignCenter)
        try:
            f_small = self.game_info_label.font()
            f_small.setPointSize(max(8, f_small.pointSize()-1))
            self.game_info_label.setFont(f_small)
            self.game_info_label.setStyleSheet("color: #cfd8dc;")
        except Exception:
            pass

        # Sims process state label
        self.sims_state_label = QtWidgets.QLabel("", self)
        try:
            sfont = self.sims_state_label.font()
            sfont.setPointSize(max(8, sfont.pointSize()-1))
            self.sims_state_label.setFont(sfont)
            self.sims_state_label.setStyleSheet("color: #cfd8dc;")
        except Exception:
            pass

        # AI mode label
        self.ai_mode_label = QtWidgets.QLabel("", self)
        try:
            af = self.ai_mode_label.font()
            af.setBold(True)
            self.ai_mode_label.setFont(af)
            self.ai_mode_label.setStyleSheet("color: #80cbc4;")
        except Exception:
            pass

        # Layout: [Config] [stretch] [game info centered] [stretch] [state] [AI] [Kill] [Launch]
        bottom_buttons.addWidget(self.configuration_button)
        bottom_buttons.addStretch(1)
        bottom_buttons.addWidget(self.game_info_label)
        bottom_buttons.addStretch(1)
        bottom_buttons.addWidget(self.sims_state_label)
        bottom_buttons.addSpacing(8)
        bottom_buttons.addWidget(self.ai_mode_label)
        bottom_buttons.addSpacing(12)
        bottom_buttons.addWidget(self.kill_button)
        bottom_buttons.addWidget(self.launch_button)
        layout.addLayout(bottom_buttons)
        # Final
        self.setLayout(layout)

        # Apply icons to actions and bottom buttons
        self._apply_button_icons()
        # Apply background if configured
        self._apply_background()
        # Populate info labels
        try:
            self._update_game_info_label()
        except Exception:
            pass
        self.update_launch_button_state()
        # Initialize AI engine if enabled
        try:
            self.mod_ai = None
            if bool(self.settings.get("ai_enabled", False)):
                self.mod_ai = ModAI.load(str(self.settings.get("ai_model_path", "mod_ai.json")))
        except Exception:
            self.mod_ai = None
        # Initialize state + AI labels and start a small poll timer for process state
        try:
            self._update_sims_state_label()
        except Exception:
            pass
        try:
            self._update_ai_mode_label()
        except Exception:
            pass
        try:
            QtCore.QTimer.singleShot(0, self._apply_background)
        except Exception:
            pass
        try:
            self._sims_state_timer = QtCore.QTimer(self)
            self._sims_state_timer.timeout.connect(self._update_sims_state_label)
            self._sims_state_timer.start(2000)
        except Exception:
            pass

    def _normalize_actions_buttons(self):
        # Use expanding size for grid tiles; no fixed widths
        try:
            for i in range(self.actions_layout.count()):
                item = self.actions_layout.itemAt(i)
                w = item.widget() if item else None
                if w is None:
                    continue
                w.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred)
        except Exception:
            pass

    def _layout_tools_buttons(self, buttons):
        # Adaptive grid for tools
        try:
            # clear existing
            while self.tools_layout.count():
                item = self.tools_layout.takeAt(0)
                w = item.widget()
                if w:
                    self.tools_layout.removeWidget(w)
            # compute cols
            group_width = max(1, self.tools_group.width() or self.width() // 3 or 1)
            min_tile = 160
            cols = max(1, group_width // min_tile)
            row = col = 0
            for btn in buttons:
                btn.setSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred)
                self.tools_layout.addWidget(btn, row, col)
                col += 1
                if col >= cols:
                    col = 0
                    row += 1
        except Exception:
            # fallback: stack vertically
            r = 0
            for btn in buttons:
                self.tools_layout.addWidget(btn, r, 0)
                r += 1

        # Re-apply layout after size change
        try:
            self._layout_action_buttons()
        except Exception:
            pass

    def _layout_action_buttons(self):
        # Adaptive grid: choose number of columns based on available width
        try:
            # Clear grid
            while self.actions_layout.count():
                item = self.actions_layout.takeAt(0)
                w = item.widget()
                if w is not None:
                    self.actions_layout.removeWidget(w)
        except Exception:
            pass

        group_width = max(1, self.actions_group.width() or self.width() // 2 or 1)
        min_tile = 150  # pixels per tile, ensures normalized button size
        cols = max(2, min(len(self._action_buttons), group_width // min_tile))
        try:
            if hasattr(self, "logger"):
                self.logger.debug("Actions layout: group_width=%s, min_tile=%s, cols=%s", group_width, min_tile, cols)
        except Exception:
            pass
        row = col = 0
        for btn in self._action_buttons:
            self.actions_layout.addWidget(btn, row, col)
            col += 1
            if col >= cols:
                col = 0
                row += 1
        # Update action icon sizes based on computed columns/width
        try:
            self._update_action_icons_for_layout()
        except Exception:
            pass

    def resizeEvent(self, event):
        super().resizeEvent(event)
        try:
            self._layout_action_buttons()
        except Exception:
            pass
        try:
            self._update_action_icons_for_layout()
        except Exception:
            pass

    def _update_action_icons_for_layout(self):
        # Adapt icon sizes to available width/columns for a balanced look
        try:
            group_width = max(1, self.actions_group.width() or self.width() // 2 or 1)
            min_tile = 150
            cols = max(2, min(len(self._action_buttons), group_width // min_tile))
            # Estimate tile width excluding margins/spacing
            l, t, r, b = self.actions_layout.getContentsMargins()
            spacing = self.actions_layout.horizontalSpacing() if self.actions_layout.horizontalSpacing() != -1 else 12
            inner_width = max(1, group_width - l - r - max(0, cols - 1) * spacing)
            tile_width = max(100, inner_width // cols)
            # Icon size between 24 and 40 px depending on tile width
            icon_px = max(24, min(40, int(tile_width * 0.45)))
            if hasattr(self, "logger"):
                try:
                    self.logger.debug("Action icon sizing: group=%s, cols=%s, tile=%s, icon=%spx", group_width, cols, tile_width, icon_px)
                except Exception:
                    pass
            size = QtCore.QSize(icon_px, icon_px)
            for btn in self._action_buttons:
                if isinstance(btn, QtWidgets.QToolButton):
                    try:
                        btn.setIconSize(size)
                    except Exception:
                        pass
        except Exception:
            pass

    def open_group_view(self):
        dialog = GroupViewDialog(self, list(self.all_data_rows))
        dialog.exec_()

    def _update_scan_status(self, message):
        if hasattr(self, "scan_status_label") and self.scan_status_label is not None:
            self.scan_status_label.setText(message)
            self.scan_status_label.setVisible(bool(message))
            self._yield_ui_events()
        if getattr(self, "_splash", None) is not None and message:
            try:
                self._splash.update_message(message)
            except Exception:
                pass

    def _start_scan_progress(self):
        if hasattr(self, "scan_progress_bar") and self.scan_progress_bar is not None:
            self.scan_progress_bar.setVisible(True)
            self.scan_progress_bar.setMaximum(0)
            self.scan_progress_bar.setValue(0)
        if hasattr(self, "scan_count_label") and self.scan_count_label is not None:
            self.scan_count_label.setText("")
            self.scan_count_label.setVisible(True)
        self._yield_ui_events()

    def _finish_scan_progress(self):
        if hasattr(self, "scan_progress_bar") and self.scan_progress_bar is not None:
            self.scan_progress_bar.setVisible(False)
        if hasattr(self, "scan_count_label") and self.scan_count_label is not None:
            self.scan_count_label.setVisible(False)
        self._yield_ui_events()
        if getattr(self, "_splash", None) is not None:
            try:
                self._splash.finish(self)
            except Exception:
                pass
            self._splash = None

    def _handle_scan_progress(self, processed, total, current_path):
        if hasattr(self, "scan_progress_bar") and self.scan_progress_bar is not None:
            if total:
                if self.scan_progress_bar.maximum() != total:
                    self.scan_progress_bar.setMaximum(total)
                self.scan_progress_bar.setValue(min(processed, total))
            else:
                self.scan_progress_bar.setMaximum(0)
        if hasattr(self, "scan_count_label") and self.scan_count_label is not None:
            if total:
                self.scan_count_label.setText(f"Objets scannés : {processed}/{total}")
            else:
                self.scan_count_label.setText(f"Objets scannés : {processed}")
            self.scan_count_label.setVisible(True)
        self._yield_ui_events()
        if getattr(self, "_splash", None) is not None:
            try:
                if total:
                    self._splash.update_message(f"Scan {processed}/{total}")
                elif processed:
                    self._splash.update_message(f"Scan {processed}")
            except Exception:
                pass

    def _start_background_scan(self, folder):
        # Cancel/ignore previous by bumping request id
        self._scan_request_id += 1
        request_id = self._scan_request_id
        # Stop previous polling timer if any
        try:
            if hasattr(self, '_scan_poll_timer') and self._scan_poll_timer is not None:
                self._scan_poll_timer.stop()
        except Exception:
            pass
        # Create a fresh queue and thread
        self._scan_progress_queue = queue.Queue()

        def progress_cb(processed, total, path):
            try:
                self._scan_progress_queue.put_nowait((processed, total, path))
            except Exception:
                pass

        def worker_run():
            try:
                rows, scan_changed = generate_data_rows(
                    folder,
                    self.settings,
                    self.version_releases,
                    progress_callback=progress_cb,
                    notify_callback=lambda msg: QtCore.QTimer.singleShot(0, lambda m=msg: self._update_scan_status(m)),
                )
            except Exception as exc:
                # Report error back on UI thread
                QtCore.QTimer.singleShot(0, lambda e=str(exc), rid=request_id: self._on_scan_error(e, rid))
                return
            # Post result on UI thread, but ensure request id matches
            QtCore.QTimer.singleShot(0, lambda r=rows, ch=scan_changed, rid=request_id: self._on_scan_finished(r, ch, rid))

        t = threading.Thread(target=worker_run, name="ScanThread", daemon=True)
        self._scan_thread = t
        t.start()
        # Start a small polling timer to drain progress queue to UI
        self._scan_poll_timer = QtCore.QTimer(self)
        self._scan_poll_timer.setInterval(50)
        def _drain_progress():
            try:
                while True:
                    processed, total, path = self._scan_progress_queue.get_nowait()
                    self._handle_scan_progress(processed, total, path)
            except Exception:
                pass
        self._scan_poll_timer.timeout.connect(_drain_progress)
        self._scan_poll_timer.start()

    def _load_external_tools_buttons(self):
        buttons = []
        modules_dir = os.path.join(os.getcwd(), 'modules')
        if not os.path.isdir(modules_dir):
            return buttons
        if modules_dir not in sys.path:
            sys.path.append(modules_dir)
        for entry in os.listdir(modules_dir):
            if not entry.lower().endswith('.py'):
                continue
            name = os.path.splitext(entry)[0]
            if name.startswith('_'):
                continue
            try:
                mod = importlib.import_module(name)
            except Exception:
                continue
            tools = []
            try:
                if hasattr(mod, 'get_tools') and callable(getattr(mod, 'get_tools')):
                    tools = list(mod.get_tools(self)) or []
                elif hasattr(mod, 'register_tools') and callable(getattr(mod, 'register_tools')):
                    tools = list(mod.register_tools(self)) or []
            except Exception:
                tools = []
            # Normalize tools: accept list of (label, callable) or ready QPushButton
            for t in tools:
                try:
                    if isinstance(t, QtWidgets.QPushButton):
                        buttons.append(t)
                    elif isinstance(t, (list, tuple)) and len(t) >= 2:
                        label, handler = t[0], t[1]
                        btn = QtWidgets.QPushButton(str(label), self.tools_group)
                        if callable(handler):
                            btn.clicked.connect(lambda _=False, h=handler: h(self))
                        buttons.append(btn)
                except Exception:
                    continue
        return buttons

    @SLOT(object, bool, int)
    def _on_scan_finished(self, rows, scan_changed, request_id):
        if request_id != self._scan_request_id:
            # stale result from previous request; ignore
            return
        try:
            try:
                self.logger.info("Scan finished: %d rows (changed=%s)", len(rows or []), bool(scan_changed))
            except Exception:
                pass
            # Always render from cache to avoid heavy UI population issues
            try:
                self.logger.info("Rendering from cache after scan (workaround enabled)")
            except Exception:
                pass
            try:
                self._populate_from_cache()
            except Exception:
                pass
            self._update_scan_status("")
            self._finish_scan_progress()
            # Stop progress polling timer
            try:
                if hasattr(self, '_scan_poll_timer') and self._scan_poll_timer is not None:
                    self._scan_poll_timer.stop()
            except Exception:
                pass
            self._cache_clear_triggered_this_refresh = False
            if scan_changed:
                self._cache_clear_triggered_this_refresh = True
                # Do not clear cache on manual Analyze/Refresh
                # Only clear if explicitly requested by an installation context
                if getattr(self, "_context_forced_cache_clear", False):
                    try:
                        self.clear_sims4_cache()
                    finally:
                        self._context_forced_cache_clear = False
        except Exception as exc:
            try:
                self.logger.exception("Populate/render failed after scan: %s", exc)
            except Exception:
                pass

    @SLOT(str, int)
    def _on_scan_error(self, message, request_id):
        if request_id != self._scan_request_id:
            return
        try:
            QtWidgets.QMessageBox.critical(self, "Erreur de scan", message)
        except Exception:
            pass
        try:
            if hasattr(self, '_scan_poll_timer') and self._scan_poll_timer is not None:
                self._scan_poll_timer.stop()
        except Exception:
            pass
        # Ensure progress UI is reset on error
        try:
            self._update_scan_status("")
            self._finish_scan_progress()
        except Exception:
            pass

    def populate_version_combos(self):
        if not hasattr(self, "version_start_combo"):
            return
        entries = list(self.version_releases.items())
        combo_map = (
            (self.version_start_combo, "version_filter_start"),
            (self.version_end_combo, "version_filter_end"),
        )
        for combo, setting_key in combo_map:
            current_value = self.settings.get(setting_key, "") or ""
            combo.blockSignals(True)
            combo.clear()
            combo.addItem("Aucune (pas de limite)", "")
            for version, release_date in entries:
                combo.addItem(f"{version} – {format_release_date(release_date)}", version)
            index = combo.findData(current_value)
            combo.setCurrentIndex(index if index != -1 else 0)
            combo.blockSignals(False)

    def _on_version_filters_toggled(self, checked):
        self.settings["enable_version_filters"] = bool(checked)
        save_settings(self.settings)
        self._update_version_filter_visibility()

    def _apply_button_icons(self):
        try:
            style = QtWidgets.QApplication.style()
            # Base icon size for action tiles; final size adapted per layout
            icon_size = QtCore.QSize(36, 36)
            mapping = [
                (self.mod_installer_button, QtWidgets.QStyle.SP_DialogOpenButton),
                (self.group_view_button, QtWidgets.QStyle.SP_DirIcon),
                (self.open_mods_button, QtWidgets.QStyle.SP_DirOpenIcon),
                (self.refresh_button, QtWidgets.QStyle.SP_BrowserReload),
                (self.export_button, QtWidgets.QStyle.SP_DialogSaveButton),
                (self.clear_cache_button, QtWidgets.QStyle.SP_TrashIcon),
                (self.grab_logs_button, QtWidgets.QStyle.SP_FileIcon),
                (self.kill_button, QtWidgets.QStyle.SP_MediaStop),
                (self.configuration_button, QtWidgets.QStyle.SP_FileDialogDetailedView),
                (self.launch_button, QtWidgets.QStyle.SP_MediaPlay),
            ]
            for btn, sp in mapping:
                try:
                    btn.setIcon(style.standardIcon(sp))
                    # Smaller icons for bottom bar buttons; action tiles sized adaptively later
                    if btn in (self.configuration_button, self.launch_button, self.kill_button):
                        btn.setIconSize(QtCore.QSize(20, 20))
                    else:
                        btn.setIconSize(icon_size)
                    if isinstance(btn, QtWidgets.QToolButton):
                        btn.setToolButtonStyle(QtCore.Qt.ToolButtonTextUnderIcon)
                except Exception:
                    pass
            # Let layout decide final icon size based on available width
            self._update_action_icons_for_layout()
        except Exception:
            pass
        self.refresh_table_only()

    def open_mods_folder(self):
        root = self.settings.get("mod_directory", "")
        if not root or not os.path.isdir(root):
            QtWidgets.QMessageBox.warning(self, "Dossier des mods invalide", "Définis un dossier des mods valide dans la configuration.")
            return
        self._open_in_file_manager(root)

    def _start_web_interface(self):
        host = str(self.settings.get("web_host", "127.0.0.1") or "127.0.0.1")
        try:
            port = int(self.settings.get("web_port", 5000) or 5000)
        except Exception:
            port = 5000
        debug = bool(self.settings.get("web_debug", False))
        try:
            # Import late to avoid circular import (webapp imports from main)
            import webapp as _web
            self._web_thread = _web.start_in_thread(host=host, port=port, debug=debug)
            try:
                self.logger.info("Web interface started on http://%s:%s", host, port)
            except Exception:
                pass
        except Exception as exc:
            try:
                self.logger.warning("Web interface failed to start: %s", exc)
            except Exception:
                pass

    def _update_version_filter_visibility(self):
        enabled = bool(self.settings.get("enable_version_filters", True))
        if hasattr(self, "version_filters_checkbox"):
            if self.version_filters_checkbox.isChecked() != enabled:
                self.version_filters_checkbox.blockSignals(True)
                self.version_filters_checkbox.setChecked(enabled)
                self.version_filters_checkbox.blockSignals(False)
        widgets = [
            getattr(self, "version_start_label", None),
            getattr(self, "version_start_combo", None),
            getattr(self, "version_end_label", None),
            getattr(self, "version_end_combo", None),
        ]
        for widget in widgets:
            if widget is not None:
                widget.setVisible(enabled)
                widget.setEnabled(enabled)

    def _apply_background(self):
        # Recalcule la feuille de style de base selon l'opacité
        try:
            self._base_stylesheet = self._compute_base_stylesheet()
        except Exception:
            pass
        try:
            path = str(self.settings.get("background_image_path", "") or "").strip()
        except Exception:
            path = ""
        extra = ""
        if path:
            try:
                # Expand env vars and ~, then absolute
                expanded = os.path.expanduser(os.path.expandvars(path))
                abs_path = os.path.abspath(expanded)
            except Exception:
                abs_path = path
            if os.path.isfile(abs_path):
                qpath = abs_path.replace("\\", "/")
                extra = (
                    "\n#root { "
                    f"border-image: url('{qpath}') 0 0 0 0 stretch stretch; "
                    "background-repeat: no-repeat; "
                    "background-position: center; "
                    "}\n"
                )
        self.setStyleSheet(self._base_stylesheet + extra)
        self._bg_applied = True
        try:
            self.update()
        except Exception:
            pass

    def _compute_base_stylesheet(self):
        """Construit la feuille de style de base avec une opacité configurable."""
        try:
            op = int(getattr(self, 'settings', {}).get('ui_frame_opacity', 100))
        except Exception:
            op = 100
        op = max(0, min(100, op))
        alpha = int(255 * (op / 100.0))
        # Couleurs de base du thème sombre avec alpha
        bg_app = f"rgba(46,46,46,{alpha})"       # #2e2e2e
        bg_table = f"rgba(51,51,51,{alpha})"     # #333333
        bg_header = f"rgba(78,78,78,{alpha})"    # #4e4e4e
        # Le contour reste opaque pour la lisibilité
        return f"""
            QWidget {{
                background-color: {bg_app};
                color: white;
            }}
            QWidget#root {{
                background-color: {bg_app};
            }}
            QFrame, QGroupBox, QScrollArea, QSplitter {{
                background-color: {bg_app};
            }}
            QTableWidget {{
                background-color: {bg_table};
                alternate-background-color: #3a3a3a;
                color: white;
                border: 1px solid #444444;
            }}
            QTreeWidget {{
                background-color: {bg_table};
                alternate-background-color: #3a3a3a;
                color: white;
                border: 1px solid #444444;
            }}
            QTreeView, QListView, QListWidget {{
                background-color: {bg_table};
                alternate-background-color: #3a3a3a;
                color: white;
            }}
            QLineEdit, QComboBox, QTextEdit, QPlainTextEdit, QSpinBox, QDoubleSpinBox {{
                background-color: {bg_table};
                color: white;
                border: 1px solid #444444;
            }}
            QPushButton, QToolButton {{
                background-color: #455a64;
                color: white;
                border: 1px solid #1c313a;
                border-radius: 4px;
                padding: 6px 12px;
            }}
            QPushButton:hover, QToolButton:hover {{
                background-color: #546e7a;
            }}
            QPushButton:pressed, QToolButton:pressed {{
                background-color: #29434e;
            }}
            QPushButton:disabled, QToolButton:disabled {{
                background-color: #2e3c43;
                color: #8f9ea4;
                border-color: #2a3135;
            }}
            QHeaderView::section {{
                background-color: {bg_header};
                color: white;
            }}
            QCheckBox {{
                color: white;
            }}
        """

    def showEvent(self, event):
        try:
            super().showEvent(event)
        except Exception:
            pass
        try:
            if not getattr(self, "_bg_applied", False):
                self._apply_background()
        except Exception:
            pass

    def _read_game_version_build(self):
        version = ""
        build = ""
        try:
            cache_dir = str(self.settings.get("sims_cache_directory", "") or "").strip()
            if not cache_dir:
                return version, build
            config_path = os.path.join(cache_dir, "config.log")
            if not os.path.isfile(config_path):
                return version, build
            with open(config_path, "r", encoding="utf-8", errors="ignore") as fh:
                for raw in fh:
                    line = raw.strip()
                    if not version and line.lower().startswith("version:"):
                        parts = line.split(":", 1)
                        if len(parts) == 2:
                            candidate = parts[1].strip()
                            if candidate:
                                version = candidate
                    elif not build and line.lower().startswith("build:"):
                        parts = line.split(":", 1)
                        if len(parts) == 2:
                            candidate = parts[1].strip()
                            if candidate:
                                build = candidate
                    if version and build:
                        break
            try:
                if hasattr(self, "logger"):
                    self.logger.debug("Read game info: version=%s, build=%s from %s", version or "", build or "", config_path)
            except Exception:
                pass
        except Exception:
            pass
        return version, build

    def _update_game_info_label(self):
        try:
            version, build = self._read_game_version_build()
            if version or build:
                display = f"Version: {version or '—'}    Build: {build or '—'}"
                self.game_info_label.setText(display)
                self.game_info_label.setVisible(True)
            else:
                self.game_info_label.setText("")
                self.game_info_label.setVisible(True)
        except Exception:
            if hasattr(self, "game_info_label"):
                try:
                    self.game_info_label.setText("")
                except Exception:
                    pass

    def on_version_filter_changed(self):
        if not hasattr(self, "version_start_combo"):
            return
        start_idx = self.version_start_combo.currentIndex()
        end_idx = self.version_end_combo.currentIndex()
        sender = self.sender()
        if start_idx > 0 and end_idx > 0 and start_idx > end_idx:
            if sender is self.version_start_combo:
                self.version_end_combo.blockSignals(True)
                self.version_end_combo.setCurrentIndex(start_idx)
                self.version_end_combo.blockSignals(False)
            else:
                self.version_start_combo.blockSignals(True)
                self.version_start_combo.setCurrentIndex(end_idx)
                self.version_start_combo.blockSignals(False)
            start_idx = self.version_start_combo.currentIndex()
            end_idx = self.version_end_combo.currentIndex()

        start_value = self.version_start_combo.itemData(start_idx)
        end_value = self.version_end_combo.itemData(end_idx)
        self.settings["version_filter_start"] = start_value or ""
        self.settings["version_filter_end"] = end_value or ""
        save_settings(self.settings)
        self.refresh_table_only()

    def get_version_entries(self):
        return list(self.version_releases.items())

    def add_version_release(self, version, release_date):
        normalized_version = str(version).strip()
        if not normalized_version:
            return False, "Le numéro de version est obligatoire."

        if hasattr(release_date, "toPyDate"):
            release_date = release_date.toPyDate()

        if release_date is None:
            return False, "La date de sortie est obligatoire."

        if normalized_version in self.version_releases:
            return False, "Cette version existe déjà."

        self.custom_version_releases[normalized_version] = release_date
        save_custom_version_releases(self.custom_version_releases)
        self.version_releases = merge_version_releases(self.custom_version_releases)
        self.populate_version_combos()
        self.refresh_table_only()
        return True, f"Version {normalized_version} ajoutée."

    def toggle_setting(self, key):
        self.settings[key] = getattr(self, f"{key}_checkbox").isChecked()
        save_settings(self.settings)
        self.refresh_table_only()

    def toggle_show_ignored(self):
        self.settings["show_ignored"] = self.show_ignored_checkbox.isChecked()
        save_settings(self.settings)
        self.refresh_table_only()

    def toggle_hide_installer(self):
        self.settings["hide_installer_mods"] = self.hide_installer_checkbox.isChecked()
        save_settings(self.settings)
        self.refresh_table_only()

    def toggle_show_search_results(self):
        checked = self.show_search_checkbox.isChecked()
        self.settings["show_search_results"] = checked
        save_settings(self.settings)
        self.search_edit.setEnabled(checked)
        self._apply_search_filter()

    def toggle_instant_search(self):
        checked = self.instant_search_checkbox.isChecked()
        self.settings["instant_search"] = bool(checked)
        save_settings(self.settings)
        # Optionally run a refresh when enabling instant search
        if checked:
            self._apply_search_filter()

    def on_file_filter_mode_changed(self):
        idx = self.file_filter_combo.currentIndex()
        mode = self.file_filter_combo.itemData(idx) or "both"
        self.settings["file_filter_mode"] = str(mode)
        save_settings(self.settings)
        self.refresh_table_only()

    def toggle_show_disabled_only(self):
        self.settings["show_disabled_only"] = bool(self.show_disabled_only_checkbox.isChecked())
        save_settings(self.settings)
        self.refresh_table_only()

    def update_mod_directory_label(self):
        directory = self.settings.get("mod_directory", "")
        display_text = directory if directory else "(non défini)"
        self.mod_directory_label.setText(display_text)

    def open_configuration(self):
        dialog = ConfigurationDialog(self, dict(self.settings))
        dialog.exec_()

    def open_mod_installer(self):
        dialog = ModInstallerDialog(self, self.settings.get("mod_directory", ""))
        dialog.exec_()
        if dialog.installations_performed:
            # Mark that we want a cache clear only for this post-install refresh
            self._context_forced_cache_clear = True
            self.refresh_tree()

    def apply_configuration(self, mod_directory, cache_directory, backups_directory, sims_executable_path, sims_executable_arguments, log_extra_extensions, grab_logs_ignore_files, log_level=None, auto_scan_on_start=None):
        previous_mod_directory = self.settings.get("mod_directory", "")
        previous_auto_scan = bool(self.settings.get("auto_scan_on_start", True))
        self.settings["mod_directory"] = mod_directory
        self.settings["sims_cache_directory"] = cache_directory
        self.settings["backups_directory"] = backups_directory
        self.settings["sims_executable_path"] = sims_executable_path
        self.settings["sims_executable_arguments"] = sims_executable_arguments
        self.settings["log_extra_extensions"] = sorted(set(log_extra_extensions))
        self.settings["grab_logs_ignore_files"] = list(grab_logs_ignore_files)
        if log_level:
            self.settings["log_level"] = str(log_level).upper()
        if auto_scan_on_start is not None:
            self.settings["auto_scan_on_start"] = bool(auto_scan_on_start)
        save_settings(self.settings)
        try:
            self.logger.info("Configuration updated (mod dir: %s)", mod_directory)
        except Exception:
            pass
        self.update_mod_directory_label()
        self.update_launch_button_state()
        try:
            self._update_game_info_label()
        except Exception:
            pass

        # Apply log level change immediately (if provided)
        if log_level:
            try:
                setup_logging(self.settings.get("log_level", "DEBUG"))
                self.logger = logging.getLogger("Sims4ModTool")
                self.logger.info("Log level set to %s", self.settings.get("log_level"))
            except Exception:
                pass

        if previous_mod_directory != mod_directory:
            self.last_scanned_directory = ""
            if hasattr(self, "table"):
                self.table.setRowCount(0)
        # Trigger immediate scan if auto-scan has just been enabled
        try:
            if auto_scan_on_start is not None and bool(auto_scan_on_start) and not previous_auto_scan:
                if mod_directory and os.path.isdir(mod_directory):
                    QtCore.QTimer.singleShot(300, self.refresh_tree)
        except Exception:
            pass

    def refresh_tree(self):
        # Global safety: block operations when Sims 4 is running
        try:
            if self._is_sims_running():
                QtWidgets.QMessageBox.warning(self, "Opération bloquée", "TS4_x64.exe est en cours d'exécution. Fermez le jeu pour analyser les mods.")
                return
        except Exception:
            pass
        folder = self.settings.get("mod_directory", "")
        if not folder or not os.path.isdir(folder):
            QtWidgets.QMessageBox.critical(self, "Erreur", "Sélectionne un dossier valide dans la configuration.")
            return
        self.settings["mod_directory"] = folder
        save_settings(self.settings)
        self.update_mod_directory_label()
        self.ignored_mods = set(self.settings.get("ignored_mods", []))
        self.last_scanned_directory = folder
        try:
            self.logger.info("Refreshing tree for directory: %s", folder)
        except Exception:
            pass
        self._update_scan_status("Scan en cours...")
        self._start_scan_progress()
        self._start_background_scan(folder)

    def refresh_table_only(self):
        # Also block refresh if Sims is running
        try:
            if self._is_sims_running():
                QtWidgets.QMessageBox.warning(self, "Opération bloquée", "TS4_x64.exe est en cours d'exécution. Fermez le jeu pour analyser les mods.")
                return
        except Exception:
            pass
        if self.last_scanned_directory and os.path.isdir(self.last_scanned_directory):
            self.ignored_mods = set(self.settings.get("ignored_mods", []))
            self._update_scan_status("Scan en cours...")
            self._cache_clear_triggered_this_refresh = False
            self._start_scan_progress()
            self._start_background_scan(self.last_scanned_directory)

    def _populate_from_cache(self):
        """Populate main table from mod_scan_cache.json when auto-scan is disabled.
        Tries to mirror core pairing/filters with cached metadata (path, type, mtime)."""
        cache = load_mod_scan_cache()
        if not cache or not cache.get("entries"):
            return
        root = cache.get("root") or self.settings.get("mod_directory", "")
        try:
            root_abs = os.path.abspath(root)
        except Exception:
            root_abs = root

        entries = cache.get("entries") or []
        # Build maps
        pkg_map = {}
        scr_map = {}
        mtime_map = {}
        path_map = {}
        for e in entries:
            rel = str(e.get("path") or "").replace("\\", "/").lstrip("/")
            tp = str(e.get("type") or "").lower()
            abs_path = os.path.join(root_abs, *rel.split("/")) if root_abs else rel
            base = os.path.basename(rel)
            mtime = e.get("mtime")
            if mtime:
                try:
                    mtime_map[abs_path] = datetime.fromtimestamp(float(mtime))
                except Exception:
                    pass
            path_map[base] = abs_path
            if tp == "package":
                pkg_map[base] = abs_path
            elif tp == "ts4script":
                scr_map[base] = abs_path

        # Load installer data to resolve groups/disabled
        group_by_path = {}
        disabled_by_path = {}
        try:
            installed = load_installed_mods()
        except Exception:
            installed = []
        atf_by_name = set()
        for entry in installed:
            target_folder = entry.get("target_folder") or ""
            name = entry.get("name") or os.path.basename(target_folder) or ""
            if entry.get("atf") and name:
                atf_by_name.add(name)
            disabled_flag = bool(entry.get("disabled", False))
            root_target = os.path.normcase(os.path.abspath(target_folder)) if target_folder else ""
            for rel in (entry.get("files") or []):
                rel_norm = str(rel).replace("\\", "/").strip().lstrip("/")
                abs_p = os.path.join(root_target, *rel_norm.split("/")) if root_target else rel_norm
                key = os.path.normcase(os.path.abspath(abs_p))
                group_by_path[key] = name
                disabled_by_path[key] = disabled_flag
            for addon in (entry.get("addons") or []):
                for rel in (addon.get("paths") or []):
                    rel_norm = str(rel).replace("\\", "/").strip().lstrip("/")
                    abs_p = os.path.join(root_target, *rel_norm.split("/")) if root_target else rel_norm
                    key = os.path.normcase(os.path.abspath(abs_p))
                    group_by_path[key] = name
                    disabled_by_path[key] = disabled_flag

        # Pair by normalized name (simple, fast)
        def _norm(name):
            try:
                return normalize_mod_basename(os.path.splitext(name)[0])
            except Exception:
                return os.path.splitext(name)[0].casefold()

        scripts_by_norm = {}
        for s in scr_map.keys():
            scripts_by_norm.setdefault(_norm(s), []).append(s)
        for v in scripts_by_norm.values():
            v.sort(key=str.casefold)

        matched = {}
        unpaired_pkgs = set(pkg_map.keys())
        unpaired_scrs = set(scr_map.keys())
        for pkg in list(unpaired_pkgs):
            n = _norm(pkg)
            if not n:
                continue
            candidates = scripts_by_norm.get(n)
            if not candidates:
                continue
            scr = next((c for c in candidates if c in unpaired_scrs), None)
            if not scr:
                continue
            matched[pkg] = {"script": scr, "confidence": "Élevée", "tooltip": "Appariement basé sur un nom normalisé identique."}
            unpaired_pkgs.remove(pkg)
            unpaired_scrs.remove(scr)

        # Settings / filters
        settings = self.settings
        ignored_mods = set(settings.get("ignored_mods", []))
        show_ignored = settings.get("show_ignored", False)
        mode = (settings.get("file_filter_mode") or "both").strip().lower()
        show_packages = mode in {"both", "package", "installer_only"}
        show_scripts = mode in {"both", "ts4script", "installer_only"}
        hide_installer_mods = bool(settings.get("hide_installer_mods", False)) and mode != "installer_only"
        version_filters_enabled = settings.get("enable_version_filters", True)
        start_version = settings.get("version_filter_start") or ""
        end_version = settings.get("version_filter_end") or ""
        if not version_filters_enabled:
            start_version = ""
            end_version = ""
        start_date = self.version_releases.get(start_version)
        end_date = self.version_releases.get(end_version)
        start_limit = datetime.combine(start_date, time.min) if start_date else None
        latest_version_key = next(reversed(self.version_releases)) if self.version_releases else None
        if end_version and latest_version_key and end_version == latest_version_key:
            end_limit = datetime.combine(date.today(), time.max)
        else:
            end_limit = datetime.combine(end_date, time.max) if end_date else None

        rows = []

        # Helper to format mtime
        def _dt_for(path):
            return mtime_map.get(path)

        # Build rows for matched pairs and package-only
        for pkg in sorted(pkg_map.keys(), key=str.casefold):
            pkg_path = pkg_map.get(pkg)
            pkg_dt = _dt_for(pkg_path) if pkg_path else None
            mi = matched.get(pkg)
            scr = mi.get("script") if mi else ""
            scr_path = scr_map.get(scr) if scr else None
            scr_dt = _dt_for(scr_path) if scr_path else None

            latest = max((d for d in (pkg_dt, scr_dt) if d is not None), default=None)
            if end_limit and latest and latest > end_limit:
                continue
            if start_limit and latest and latest < start_limit:
                continue

            has_pkg = True
            has_scr = scr_path is not None
            if not ((has_pkg and show_packages) or (has_scr and show_scripts)):
                continue

            candidates = [x for x in (pkg, scr if has_scr else None) if x]
            ignored = any(name in ignored_mods for name in candidates)
            if ignored and not show_ignored:
                continue

            status = "X" if has_scr else "MS"
            version = extract_version_from_name(os.path.splitext(pkg)[0]) or (
                extract_version_from_name(os.path.splitext(scr)[0]) if scr else ""
            ) or estimate_version_from_dates(pkg_dt, scr_dt, self.version_releases)
            confidence_value = mi.get("confidence") if mi else "—"
            tooltip = mi.get("tooltip") if mi else "Aucun appariement détecté."

            # Resolve group/disabled via tracked paths
            group_value = ""
            disabled_value = False
            for p in (pkg_path, scr_path):
                if not p:
                    continue
                key = os.path.normcase(os.path.abspath(p))
                if not group_value:
                    group_value = group_by_path.get(key, "")
                if not disabled_value:
                    disabled_value = bool(disabled_by_path.get(key, False))

            rows.append({
                "status": status,
                "group": group_value,
                "disabled": disabled_value,
                "package": pkg,
                "package_date": format_datetime(pkg_dt),
                "script": scr if has_scr else "",
                "script_date": format_datetime(scr_dt),
                "version": version,
                "confidence": confidence_value,
                "confidence_tooltip": tooltip,
                "ignored": ignored,
                "ignore_candidates": candidates or [pkg],
                "paths": [p for p in (pkg_path, scr_path) if p],
                "atf": bool(group_value and group_value in atf_by_name),
            })

        # Remaining script-only
        for scr in sorted(unpaired_scrs, key=str.casefold):
            scr_path = scr_map.get(scr)
            if not scr_path:
                continue
            scr_dt = _dt_for(scr_path)
            if end_limit and scr_dt and scr_dt > end_limit:
                continue
            if start_limit and scr_dt and scr_dt < start_limit:
                continue
            if not show_scripts:
                continue
            candidates = [scr]
            ignored = any(name in ignored_mods for name in candidates)
            if ignored and not show_ignored:
                continue
            version = extract_version_from_name(os.path.splitext(scr)[0]) or estimate_version_from_dates(None, scr_dt, self.version_releases)
            key = os.path.normcase(os.path.abspath(scr_path)) if scr_path else None
            group_value = group_by_path.get(key, "") if key else ""
            disabled_value = bool(disabled_by_path.get(key, False)) if key else False
            rows.append({
                "status": "MP",
                "group": group_value,
                "disabled": disabled_value,
                "package": "",
                "package_date": "",
                "script": scr,
                "script_date": format_datetime(scr_dt),
                "version": version,
                "confidence": "—",
                "confidence_tooltip": "Aucun package correspondant trouvé.",
                "ignored": ignored,
                "ignore_candidates": candidates,
                "paths": [scr_path],
                "atf": bool(group_value and group_value in atf_by_name),
            })

        # Render
        self.populate_table(rows)

    def clear_sims4_cache(self):
        cache_directory = self.settings.get("sims_cache_directory", "")
        if not cache_directory or not os.path.isdir(cache_directory):
            QtWidgets.QMessageBox.warning(self, "Dossier cache invalide", "Configure un dossier cache Sims 4 valide dans la configuration.")
            return

        targets = [
            "localthumbcache.package",
            "localsimtexturecache.package",
            "avatarcache.package",
            "cachestr",
            "onlinethumbnailcache",
        ]

        removed = []
        missing = []
        errors = []

        for item in targets:
            path = os.path.join(cache_directory, item)
            if not os.path.exists(path):
                missing.append(item)
                continue
            try:
                if os.path.isdir(path):
                    shutil.rmtree(path)
                else:
                    os.remove(path)
                removed.append(item)
            except OSError as exc:
                errors.append(f"{item} : {exc}")

        messages = []
        if removed:
            messages.append("Supprimé : " + ", ".join(removed))
        if missing:
            messages.append("Absent : " + ", ".join(missing))
        if errors:
            messages.append("Erreurs :\n" + "\n".join(errors))

        if not messages:
            messages.append("Aucun fichier ou dossier à supprimer.")

        QtWidgets.QMessageBox.information(self, "Nettoyage du cache", "\n".join(messages))
        try:
            self.logger.info("Cache cleared. Removed: %s, Missing: %s, Errors: %s", ", ".join(removed) if removed else "—", ", ".join(missing) if missing else "—", "; ".join(errors) if errors else "—")
        except Exception:
            pass

        launch_response = QtWidgets.QMessageBox.question(
            self,
            "Lancer Les Sims 4",
            "Souhaitez-vous lancer Les Sims 4 maintenant ?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if launch_response == QtWidgets.QMessageBox.Yes:
            self.launch_sims4()

    def grab_logs(self):
        mod_directory = self.settings.get("mod_directory", "")
        if not mod_directory or not os.path.isdir(mod_directory):
            QtWidgets.QMessageBox.warning(self, "Dossier des mods invalide", "Définis un dossier des mods valide dans la configuration avant d'extraire les logs.")
            return

        backups_directory = self.settings.get("backups_directory", "")
        if not backups_directory:
            QtWidgets.QMessageBox.warning(self, "Dossier backups manquant", "Définis un dossier de backups dans la configuration avant d'extraire les logs.")
            return

        os.makedirs(backups_directory, exist_ok=True)

        extensions = {".log", ".txt"}
        extensions.update(self.settings.get("log_extra_extensions", []))
        ignored_log_names = {name.lower() for name in self.settings.get("grab_logs_ignore_files", []) if name}

        normalized_mod_dir = os.path.normpath(mod_directory)
        backups_directory_norm = os.path.normpath(backups_directory)
        found_logs = []

        for current_root, dirs, files in os.walk(normalized_mod_dir):
            dirs[:] = [d for d in dirs if os.path.normpath(os.path.join(current_root, d)) != backups_directory_norm]
            for file_name in files:
                if file_name.lower() in ignored_log_names:
                    continue
                _, ext = os.path.splitext(file_name)
                if ext.lower() in extensions:
                    file_path = os.path.join(current_root, file_name)
                    normalized_file = os.path.normpath(file_path)
                    try:
                        if os.path.commonpath([normalized_file, backups_directory_norm]) == backups_directory_norm:
                            continue
                    except ValueError:
                        pass
                    found_logs.append(file_path)

        if not found_logs:
            QtWidgets.QMessageBox.information(self, "Aucun log", "Aucun fichier de log correspondant n'a été trouvé.")
            try:
                self.logger.info("Grab Logs: no matching logs found under %s", normalized_mod_dir)
            except Exception:
                pass
            return

        timestamp = datetime.now().strftime("Logs_%Y%m%d_%H%M%S")
        destination_root = os.path.join(backups_directory, timestamp)
        os.makedirs(destination_root, exist_ok=True)

        moved_files = []
        errors = []
        for source_path in found_logs:
            relative_path = os.path.relpath(source_path, normalized_mod_dir)
            destination_path = os.path.join(destination_root, relative_path)
            destination_dir = os.path.dirname(destination_path)
            os.makedirs(destination_dir, exist_ok=True)

            final_destination = destination_path
            counter = 1
            while os.path.exists(final_destination):
                name, ext = os.path.splitext(destination_path)
                final_destination = f"{name}_{counter}{ext}"
                counter += 1

            try:
                shutil.move(source_path, final_destination)
                moved_files.append(final_destination)
            except OSError as exc:
                errors.append(f"{source_path} → {exc}")

        if not moved_files:
            QtWidgets.QMessageBox.information(self, "Aucun log déplacé", "Aucun fichier n'a pu être déplacé.")
            try:
                self.logger.info("Grab Logs: moved 0 files; errors: %d", len(errors))
            except Exception:
                pass
            return

        message_lines = [f"{len(moved_files)} fichier(s) de log déplacé(s) vers {destination_root}."]
        if errors:
            message_lines.append("\nErreurs:\n" + "\n".join(errors))
        QtWidgets.QMessageBox.information(self, "Logs sauvegardés", "\n".join(message_lines))
        try:
            self.logger.info("Grab Logs: moved %d files to %s; errors: %d", len(moved_files), destination_root, len(errors))
        except Exception:
            pass
        self._open_in_file_manager(destination_root)

    def launch_sims4(self):
        executable_path = self.settings.get("sims_executable_path", "")
        if not executable_path:
            QtWidgets.QMessageBox.warning(self, "Exécutable manquant", "Configure le chemin de TS4_X64.exe dans la configuration.")
            return

        if not os.path.isfile(executable_path):
            QtWidgets.QMessageBox.critical(self, "Exécutable introuvable", "Le fichier TS4_X64.exe configuré est introuvable.")
            return

        args_text = self.settings.get("sims_executable_arguments", "").strip()
        try:
            args = shlex.split(args_text, posix=not sys.platform.startswith("win")) if args_text else []
        except ValueError as exc:
            QtWidgets.QMessageBox.warning(self, "Arguments invalides", f"Les arguments spécifiés sont invalides : {exc}")
            return

        try:
            if sys.platform == "darwin":
                started = QtCore.QProcess.startDetached(executable_path, args)
            elif sys.platform.startswith("win"):
                started = QtCore.QProcess.startDetached(executable_path, args)
            else:
                started = QtCore.QProcess.startDetached(executable_path, args)
            if not started:
                raise OSError("le processus n'a pas pu être démarré")
        except OSError as exc:
            QtWidgets.QMessageBox.critical(self, "Erreur", f"Impossible de lancer Sims 4 : {exc}")

    def kill_sims4(self):
        process_name = "TS4_x64.exe"
        if sys.platform.startswith("win"):
            # First, try normal taskkill
            try:
                completed = subprocess.run(["taskkill", "/F", "/IM", process_name], capture_output=True, text=True)
                if completed.returncode == 0:
                    QtWidgets.QMessageBox.information(self, "Sims 4 arrêté", "Le processus TS4_x64.exe a été arrêté avec succès.")
                    return
            except FileNotFoundError:
                pass
            # Try elevated kill via UAC prompt
            try:
                ps_cmd = "Start-Process -Verb RunAs powershell -ArgumentList 'Stop-Process -Name TS4_x64 -Force'"
                completed2 = subprocess.run(["powershell", "-NoProfile", "-Command", ps_cmd], capture_output=True, text=True)
                # We can't detect if user canceled UAC; show informational dialog
                QtWidgets.QMessageBox.information(self, "Tentative d'arrêt", "Une demande d'élévation a été envoyée pour arrêter TS4_x64.exe. Si l'arrêt échoue, exécutez l'application en tant qu'administrateur.")
            except Exception as exc:
                QtWidgets.QMessageBox.warning(self, "Arrêt impossible", f"Impossible d'arrêter TS4_x64.exe : {exc}")
        else:
            try:
                completed = subprocess.run(["pkill", "-f", process_name], capture_output=True, text=True)
                if completed.returncode == 0:
                    QtWidgets.QMessageBox.information(self, "Sims 4 arrêté", "Le processus a été arrêté avec succès.")
                else:
                    output = completed.stderr.strip() or completed.stdout.strip() or "La commande a échoué."
                    QtWidgets.QMessageBox.warning(self, "Aucun processus arrêté", output)
            except FileNotFoundError:
                QtWidgets.QMessageBox.critical(self, "Commande introuvable", "La commande pkill est introuvable.")

    def _is_sims_running(self):
        name = "TS4_x64.exe"
        if sys.platform.startswith("win"):
            try:
                completed = subprocess.run(["tasklist"], capture_output=True, text=True)
                out = (completed.stdout or "") + "\n" + (completed.stderr or "")
                return name.lower() in out.lower()
            except Exception:
                return False
        else:
            try:
                completed = subprocess.run(["pgrep", "-f", "TS4_x64"], capture_output=True)
                return completed.returncode == 0
            except Exception:
                return False

    def update_launch_button_state(self):
        if hasattr(self, "launch_button"):
            executable_path = self.settings.get("sims_executable_path", "")
            self.launch_button.setEnabled(bool(executable_path and os.path.isfile(executable_path)))

    def _update_sims_state_label(self):
        try:
            running = self._is_sims_running()
        except Exception:
            running = False
        text = "TS4_x64.exe: running" if running else "TS4_x64.exe: stopped"
        color = "#66bb6a" if running else "#ef9a9a"
        try:
            self.sims_state_label.setText(text)
            self.sims_state_label.setStyleSheet(f"color: {color};")
        except Exception:
            pass

    def _evaluate_ai_model_status(self):
        try:
            enabled = bool(self.settings.get("ai_enabled", False))
        except Exception:
            enabled = False
        if not enabled:
            return False, "IA désactivée", "#90a4ae"

        ai = getattr(self, 'mod_ai', None)
        if ai is None:
            model_path = str(self.settings.get("ai_model_path", "mod_ai.json") or "mod_ai.json")
            if os.path.isfile(model_path):
                try:
                    self.mod_ai = ModAI.load(model_path)
                    ai = self.mod_ai
                except Exception:
                    pass
        if ai is None:
            return True, "IA activée : modèle absent", "#ff7043"

        data = ai.data or {}
        metadata = data.get('metadata', {}) or {}
        ml_data = data.get('ml', {}) or {}
        engine = metadata.get('engine') or ml_data.get('last_engine') or 'tokens'
        trained_at = metadata.get('trained_at') or ''
        mods_known = len(data.get('mods', {}) or {})
        tokens_known = len(data.get('token_to_mod', {}) or {})
        needs_training = bool(metadata.get('needs_training'))

        missing = False
        if engine == 'tfidf':
            model_path = str(ml_data.get('tfidf_model_path') or '').strip()
            if not model_path or not os.path.isfile(model_path):
                missing = True
        elif engine == 'mlp':
            model_path = str(ml_data.get('mlp_model_path') or '').strip()
            if not model_path or not os.path.isfile(model_path):
                missing = True
        else:
            if tokens_known == 0:
                missing = True

        parts = []
        label_map = {
            'tokens': 'IA tokens',
            'tfidf': 'IA TF‑IDF',
            'mlp': 'IA MLP',
        }
        parts.append(label_map.get(engine, f"IA {engine}"))
        if trained_at:
            try:
                dt = datetime.fromisoformat(trained_at)
                parts.append(dt.strftime("%Y-%m-%d %H:%M"))
            except Exception:
                parts.append(trained_at)
        parts.append(f"mods={mods_known}")
        parts.append(f"tokens={tokens_known}")

        needs = needs_training or missing
        if needs and not needs_training:
            try:
                metadata['needs_training'] = True
            except Exception:
                pass
        if needs:
            parts.append("entrainement requis")

        text = " • ".join(parts)
        color = "#ffb74d" if needs else "#80cbc4"
        return needs, text, color

    def _update_ai_mode_label(self):
        needs, text, color = self._evaluate_ai_model_status()
        try:
            self.ai_mode_label.setText(text)
            self.ai_mode_label.setStyleSheet(f"color: {color}; font-weight:bold;")
        except Exception:
            pass

    def open_tools_dialog(self):
        dialog = QtWidgets.QDialog(self)
        try:
            dialog.setWindowFlags(dialog.windowFlags() | QtCore.Qt.Window)
            dialog.setSizeGripEnabled(True)
        except Exception:
            pass
        dialog.setWindowTitle("Tools")
        dialog.setModal(True)

        layout = QtWidgets.QVBoxLayout(dialog)

        btn_dup = QtWidgets.QPushButton("Find dupplicates", dialog)
        btn_dup.clicked.connect(self.open_duplicate_finder)
        layout.addWidget(btn_dup)

        btn_nonmods = QtWidgets.QPushButton("Find non-mods files", dialog)
        btn_nonmods.clicked.connect(partial(self._show_placeholder_message, "Find non-mods files", "La détection des fichiers non mods sera ajoutée ultérieurement."))
        layout.addWidget(btn_nonmods)

        btn_disable = QtWidgets.QPushButton("Disable all mods", dialog)
        btn_disable.clicked.connect(partial(self._show_placeholder_message, "Disable all mods", "La désactivation des mods sera proposée dans une future mise à jour."))
        layout.addWidget(btn_disable)

        btn_cfg = QtWidgets.QPushButton("Correct resource.cfg", dialog)
        btn_cfg.clicked.connect(self.correct_resource_cfg)
        layout.addWidget(btn_cfg)

        btn_symlink = QtWidgets.QPushButton("Symlink Mods", dialog)
        btn_symlink.clicked.connect(partial(self._show_placeholder_message, "Symlink Mods", "La création de liens symboliques vers le dossier Mods sera ajoutée ultérieurement."))
        layout.addWidget(btn_symlink)

        btn_backup = QtWidgets.QPushButton("Backup Mods", dialog)
        btn_backup.clicked.connect(partial(self._show_placeholder_message, "Backup Mods", "La sauvegarde du dossier Mods sera disponible dans une prochaine version."))
        layout.addWidget(btn_backup)

        btn_check_ts4 = QtWidgets.QPushButton("Check placement .ts4script", dialog)
        btn_check_ts4.clicked.connect(self.check_ts4script_placement)
        layout.addWidget(btn_check_ts4)

        btn_scan_folder = QtWidgets.QPushButton("Scan dossier (mod)", dialog)
        btn_scan_folder.clicked.connect(self.open_folder_scanner)
        layout.addWidget(btn_scan_folder)

        btn_find_ts4 = QtWidgets.QPushButton("Find in ts4script", dialog)
        btn_find_ts4.clicked.connect(self.open_find_in_ts4script)
        layout.addWidget(btn_find_ts4)

        close_button = QtWidgets.QPushButton("Fermer", dialog)
        close_button.clicked.connect(dialog.accept)
        layout.addWidget(close_button)

        dialog.exec_()

    def _show_placeholder_message(self, title, message):
        QtWidgets.QMessageBox.information(self, title, message)

    def open_folder_scanner(self):
        start_dir = self.settings.get("last_folder_scan_directory", "") or self.settings.get("mod_directory", "")
        dlg = FolderScannerDialog(self, start_dir)
        dlg.exec_()

    def open_find_in_ts4script(self):
        start_dir = self.settings.get("mod_directory", "")
        dlg = Ts4ScriptSearchDialog(self, start_dir, datetime_formatter=format_datetime)
        dlg.exec_()

    def open_mod_comparator(self):
        start_dir = self.settings.get("mod_directory", "")
        dlg = ModComparatorDialog(self, start_dir)
        dlg.exec_()

    def open_log_manager(self):
        path = str(self.settings.get("last_log_path", "") or "")
        hooks = LogAnalyzerHooks(
            strip_html=_strip_html_to_text,
            analyze_html=analyze_last_exception_html,
            analyze_generic=analyze_generic_log_text,
            normalize_basename=normalize_mod_basename,
            save_settings=save_settings,
        )
        dlg = LogManagerDialog(self, path, hooks=hooks)
        dlg.exec_()

    # AI Training moved to modules/ai_training.py

    def open_updates_checker(self):
        try:
            dlg = UpdatesCheckerDialog(self)
            dlg.exec_()
        except Exception as exc:
            QtWidgets.QMessageBox.warning(self, "Updates Checker", str(exc))

    def check_ts4script_placement(self):
        mods_dir = self.settings.get("mod_directory", "")
        if not mods_dir or not os.path.isdir(mods_dir):
            QtWidgets.QMessageBox.warning(self, "Dossier des mods invalide", "Définis un dossier des mods valide dans la configuration.")
            return
        backups_directory = os.path.normpath(self.settings.get("backups_directory", "") or "")
        offenders = []
        root = os.path.abspath(mods_dir)
        try:
            if hasattr(self, "logger"):
                self.logger.debug("TS4Script placement check start: root=%s, backups=%s", root, backups_directory)
        except Exception:
            pass
        for current_root, dirs, files in os.walk(root):
            # Exclude backups directory subtree
            if backups_directory:
                try:
                    dirs[:] = [d for d in dirs if os.path.normpath(os.path.join(current_root, d)) != backups_directory]
                except Exception:
                    pass
            for file in files:
                # Debug: log every scanned file
                try:
                    if hasattr(self, "logger") and self.logger.isEnabledFor(logging.DEBUG):
                        self.logger.debug("Scan file: %s", os.path.join(current_root, file))
                except Exception:
                    pass
                if not file.lower().endswith(".ts4script"):
                    continue
                full_path = os.path.join(current_root, file)
                rel_path = os.path.relpath(full_path, root).replace("\\", "/")
                # depth is number of directories in rel path (excluding the file itself)
                parts = rel_path.split("/")
                depth = max(0, len(parts) - 1)  # folders count
                # Rule: ts4script cannot be deeper than one folder below Mods
                # valid depths: 0 or 1
                if depth > 1:
                    offenders.append(rel_path)
                    try:
                        if hasattr(self, "logger"):
                            self.logger.debug("Offender: %s (depth=%s)", rel_path, depth)
                    except Exception:
                        pass

        if not offenders:
            try:
                if hasattr(self, "logger"):
                    self.logger.info("TS4Script placement OK: no offenders found")
            except Exception:
                pass
            QtWidgets.QMessageBox.information(self, "Vérification terminée", "Tout va bien, tout le monde est à sa place 👍")
            return

        try:
            if hasattr(self, "logger"):
                self.logger.warning("TS4Script placement violations: %d", len(offenders))
        except Exception:
            pass
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("TS4Script trop profonds")
        dlg.resize(720, 420)
        vbox = QtWidgets.QVBoxLayout(dlg)
        info = QtWidgets.QLabel("Les scripts .ts4script suivants dépassent une profondeur de 1 dossier sous Mods :", dlg)
        info.setWordWrap(True)
        vbox.addWidget(info)
        table = QtWidgets.QTableWidget(dlg)
        table.setColumnCount(1)
        table.setHorizontalHeaderLabels(["Chemin (relatif à Mods)"])
        table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        table.setRowCount(len(offenders))
        for i, rel in enumerate(sorted(offenders, key=str.casefold)):
            item = QtWidgets.QTableWidgetItem(rel)
            table.setItem(i, 0, item)
        vbox.addWidget(table, stretch=1)
        close_btn = QtWidgets.QPushButton("Fermer", dlg)
        close_btn.clicked.connect(dlg.accept)
        vbox.addWidget(close_btn)
        dlg.exec_()

    def correct_resource_cfg(self):
        mods_dir = self.settings.get("mod_directory", "")
        if not mods_dir or not os.path.isdir(mods_dir):
            QtWidgets.QMessageBox.warning(self, "Dossier des mods invalide", "Définis un dossier des mods valide dans la configuration.")
            return
        target = os.path.join(mods_dir, "Resource.cfg")
        expected_lines = [
            "Priority 500",
            "PackedFile *.package",
            "PackedFile */*.package",
            "PackedFile */*/*.package",
            "PackedFile */*/*/*.package",
            "PackedFile */*/*/*/*.package",
            "PackedFile */*/*/*/*/*.package",
        ]
        expected_content = "\r\n".join(expected_lines) + "\r\n"
        current = ""
        if os.path.exists(target):
            try:
                with open(target, "r", encoding="utf-8", errors="ignore") as fh:
                    current = fh.read()
            except OSError:
                current = ""
        def _normalize(text):
            return "\n".join([line.strip() for line in text.replace("\r\n", "\n").replace("\r", "\n").split("\n") if line is not None])
        if _normalize(current) == _normalize(expected_content):
            QtWidgets.QMessageBox.information(self, "Resource.cfg", "Le fichier Resource.cfg est déjà conforme.")
            return
        # Backup existing if present
        if os.path.exists(target):
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup = os.path.join(mods_dir, f"Resource.cfg.bak_{ts}")
            try:
                shutil.copy2(target, backup)
            except OSError:
                pass
        try:
            with open(target, "w", encoding="utf-8") as fh:
                fh.write(expected_content)
            QtWidgets.QMessageBox.information(self, "Resource.cfg", "Le fichier Resource.cfg a été corrigé.")
        except OSError as exc:
            QtWidgets.QMessageBox.critical(self, "Erreur", f"Impossible d'écrire Resource.cfg : {exc}")

    def open_duplicate_finder(self):
        root = self.settings.get("mod_directory", "")
        if not root or not os.path.isdir(root):
            QtWidgets.QMessageBox.warning(self, "Dossier des mods invalide", "Définis un dossier des mods valide dans la configuration.")
            return
        dialog = DuplicateFinderDialog(self, root)
        dialog.exec_()

    def open_conflict_checker(self):
        root = self.settings.get("mod_directory", "")
        if not root or not os.path.isdir(root):
            QtWidgets.QMessageBox.warning(self, "Dossier des mods invalide", "Définis un dossier des mods valide dans la configuration.")
            return
        dialog = ConflictCheckerDialog(self, root)
        dialog.exec_()

    def open_id_conflict_viewer(self):
        root = self.settings.get("mod_directory", "")
        if not root or not os.path.isdir(root):
            QtWidgets.QMessageBox.warning(self, "Dossier des mods invalide", "Définis un dossier des mods valide dans la configuration.")
            return
        try:
            self.logger.info("Open IDConflictViewer for root=%s", root)
        except Exception:
            pass
        dialog = IDConflictViewerDialog(self, root)
        dialog.exec_()

    def open_id_conflict_viewer_v2(self):
        root = self.settings.get("mod_directory", "")
        if not root or not os.path.isdir(root):
            QtWidgets.QMessageBox.warning(self, "Dossier des mods invalide", "Définis un dossier des mods valide dans la configuration.")
            return
        try:
            self.logger.info("Open IDConflictViewerV2 for root=%s", root)
        except Exception:
            pass
        id_index_cache = os.path.abspath(ID_INDEX_CACHE_PATH)
        mod_scan_cache = os.path.abspath(MOD_SCAN_CACHE_PATH)
        installed_mods_path = os.path.abspath(INSTALLED_MODS_PATH)
        version_releases = dict(getattr(self, "version_releases", {}) or {})
        dialog = IDConflictViewerDialogV2(
            self,
            root,
            id_index_cache_path=id_index_cache,
            mod_scan_cache_path=mod_scan_cache,
            installed_mods_path=installed_mods_path,
            version_releases=version_releases,
        )
        dialog.exec_()

    def populate_table(self, data_rows):
        self.all_data_rows = list(data_rows)
        # Force render even when instant search is disabled
        self._apply_search_filter(forced=True)
        
    def apply_search_filter(self, _text=None, forced=False):
        # If instant search disabled, only process when forced=True
        self._apply_search_filter(forced=forced)

    def _apply_search_filter(self, forced=False):
        query = ""
        if hasattr(self, "search_edit"):
            query = self.search_edit.text().strip().lower()

        show_search_results = self.settings.get("show_search_results", True)
        instant = self.settings.get("instant_search", True)
        if not forced and not instant:
            return

        if not show_search_results:
            filtered_rows = list(self.all_data_rows)
        elif not query:
            filtered_rows = list(self.all_data_rows)
        else:
            filtered_rows = [
                row
                for row in self.all_data_rows
                if self._row_matches_query(row, query)
            ]

        # Optionally filter only disabled mods
        if self.settings.get("show_disabled_only", False):
            filtered_rows = [row for row in filtered_rows if bool(row.get("disabled", False))]

        self._render_table(filtered_rows)

    def _row_matches_query(self, row, query):
        for value in self._gather_searchable_values(row):
            if query in value:
                return True
        return False

    def _gather_searchable_values(self, row):
        values = [
            str(row.get("status", "")),
            str(row.get("group", "")),
            str(row.get("package", "")),
            str(row.get("package_date", "")),
            str(row.get("script", "")),
            str(row.get("script_date", "")),
            str(row.get("version", "")),
            str(row.get("confidence", "")),
        ]
        ignored_value = "oui" if row.get("ignored", False) else "non"
        values.append(ignored_value)
        if row.get("confidence_tooltip"):
            values.append(str(row.get("confidence_tooltip")))
        values.extend(str(candidate) for candidate in row.get("ignore_candidates", []))
        values.extend(str(path) for path in row.get("paths", []))
        return [value.lower() for value in values if value]

    def _render_table(self, rows):
        header = self.table.horizontalHeader()
        sorting_enabled = self.table.isSortingEnabled()
        sort_section = header.sortIndicatorSection()
        sort_order = header.sortIndicatorOrder()
        table = self.table
        table.setSortingEnabled(False)
        table.setUpdatesEnabled(False)
        try:
            if table.rowCount():
                table.clearContents()
            table.setRowCount(len(rows))
            for row_index, row in enumerate(rows):
                columns = [
                    row.get("status", ""),
                    row.get("group", ""),
                    row.get("package", ""),
                    row.get("package_date", ""),
                    row.get("script", ""),
                    row.get("script_date", ""),
                    row.get("version", ""),
                    row.get("confidence", ""),
                ]
                for col_idx, value in enumerate(columns):
                    item = QtWidgets.QTableWidgetItem(str(value))
                    if col_idx == 0:
                        item.setData(QtCore.Qt.UserRole, tuple(row.get("ignore_candidates") or []))
                        item.setData(QtCore.Qt.UserRole + 1, tuple(row.get("paths") or []))
                    if col_idx == 7:
                        item.setToolTip(row.get("confidence_tooltip", ""))
                    table.setItem(row_index, col_idx, item)
                # Installer indicator at column 8
                installer_item = QtWidgets.QTableWidgetItem("✓" if row.get("group") else "")
                if row.get("group"):
                    installer_item.setToolTip("Installé via Mod Installer")
                installer_item.setTextAlignment(QtCore.Qt.AlignCenter)
                installer_item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
                table.setItem(row_index, 8, installer_item)

                ignored = row.get("ignored", False)
                ignore_item = QtWidgets.QTableWidgetItem("Oui" if ignored else "Non")
                ignore_item.setData(QtCore.Qt.UserRole, 1 if ignored else 0)
                ignore_item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
                table.setItem(row_index, 9, ignore_item)

                ignore_checkbox = QtWidgets.QCheckBox()
                ignore_checkbox.stateChanged.connect(
                    partial(self.update_ignore_mod, tuple(row.get("ignore_candidates") or []))
                )
                ignore_checkbox.blockSignals(True)
                ignore_checkbox.setChecked(ignored)
                ignore_checkbox.blockSignals(False)
                table.setCellWidget(row_index, 9, ignore_checkbox)

        # Highlight Protected mods (override all)
                if row.get("atf"):
                    bg = QtGui.QBrush(QtGui.QColor("#ffc0cb"))
                    fg = QtGui.QBrush(QtGui.QColor("#000000"))
                    for c in range(self.table.columnCount()):
                        it = self.table.item(row_index, c)
                        if it is not None:
                            it.setBackground(bg)
                            it.setForeground(fg)
                # Highlight disabled mods (override if not ATF)
                if row.get("disabled"):
                    bg = QtGui.QBrush(QtGui.QColor("#8b0000"))
                    fg = QtGui.QBrush(QtGui.QColor("#ffffff"))
                    for c in range(self.table.columnCount()):
                        it = self.table.item(row_index, c)
                        if it is not None:
                            it.setBackground(bg)
                            it.setForeground(fg)
                # Highlight rows installed via Mod Installer (group present)
                elif row.get("group") and not row.get("atf"):
                    bg = QtGui.QBrush(QtGui.QColor("#2e7d32"))
                    fg = QtGui.QBrush(QtGui.QColor("#ffffff"))
                    for c in range(self.table.columnCount()):
                        it = self.table.item(row_index, c)
                        if it is not None:
                            it.setBackground(bg)
                            it.setForeground(fg)

                if row_index % 50 == 0:
                    self._yield_ui_events()
        finally:
            table.setUpdatesEnabled(True)

        table.setSortingEnabled(sorting_enabled)
        if sorting_enabled and rows:
            table.sortByColumn(sort_section, sort_order)
        table.viewport().update()
        self._yield_ui_events()

    def show_context_menu(self, position):
        index = self.table.indexAt(position)
        if not index.isValid():
            return

        row = index.row()
        status_item = self.table.item(row, 0)
        candidates = []
        if status_item is not None:
            stored_candidates = status_item.data(QtCore.Qt.UserRole)
            if stored_candidates:
                candidates = list(stored_candidates)

        menu = QtWidgets.QMenu(self)
        ignore_action = menu.addAction("Ignorer")
        show_in_explorer_action = menu.addAction("Afficher dans l'explorateur")
        delete_action = menu.addAction("Supprimer le mod")
        google_action = menu.addAction("Recherche Google")
        patreon_action = menu.addAction("Chercher sur Patreon")
        # Protected toggle (only if grouped/installed)
        group_item = self.table.item(row, 1)
        group_name = group_item.text().strip() if group_item else ""
        atf_action = None
        current_atf = False
        if group_name:
            try:
                for ent in load_installed_mods():
                    if str(ent.get("name", "")).strip().casefold() == group_name.casefold():
                        current_atf = bool(ent.get("atf", False))
                        break
            except Exception:
                current_atf = False
        label = "Retirer Protected" if current_atf else "Marquer Protected"
        atf_action = menu.addAction(label)

        selected_action = menu.exec_(self.table.viewport().mapToGlobal(position))

        if selected_action == ignore_action:
            checkbox = self.table.cellWidget(row, 9)
            if checkbox is not None:
                checkbox.setChecked(not checkbox.isChecked())
        elif selected_action == show_in_explorer_action:
            self.show_in_explorer(row, candidates)
        elif selected_action == delete_action:
            self.delete_mod(row, candidates)
        elif selected_action == google_action:
            self.launch_google_search(row, candidates)
        elif selected_action == patreon_action:
            self.launch_patreon_search(row, candidates)
        elif atf_action is not None and selected_action == atf_action:
            if not group_name:
                QtWidgets.QMessageBox.information(self, "Protected", "Action réservée aux mods installés via Mod Installer (groupe connu).")
                return
            self._toggle_atf_group(group_name)
            self._apply_search_filter(forced=True)

    def _toggle_atf_group(self, group_name):
        try:
            items = load_installed_mods()
        except Exception:
            items = []
        changed = False
        for ent in items:
            if str(ent.get("name", "")).strip().casefold() == group_name.strip().casefold():
                ent["atf"] = not bool(ent.get("atf", False))
                changed = True
                tf = ent.get("target_folder")
                if tf and os.path.isdir(tf):
                    try:
                        data = {
                            "name": ent.get("name", ""),
                            "type": ent.get("type", ""),
                            "installed_at": ent.get("installed_at", ""),
                            "source": ent.get("source", ""),
                            "app_version": APP_VERSION,
                            "app_version_date": APP_VERSION_DATE,
                            "files": list(ent.get("files", []) or []),
                            "mod_version": ent.get("mod_version", ""),
                            "url": ent.get("url", ""),
                            "atf": bool(ent.get("atf", False)),
                        }
                        with open(os.path.join(tf, MOD_MARKER_FILENAME), "w", encoding="utf-8") as fh:
                            json.dump(data, fh, ensure_ascii=False, indent=2)
                    except Exception:
                        pass
                break
        if changed:
            save_installed_mods(items)

    def _resolve_row_paths(self, row):
        status_item = self.table.item(row, 0)
        if status_item is None:
            return []
        paths = status_item.data(QtCore.Qt.UserRole + 1)
        return list(paths) if paths else []

    def _open_in_file_manager(self, target_path):
        if sys.platform.startswith("win"):
            try:
                os.startfile(target_path)
            except OSError:
                QtWidgets.QMessageBox.warning(self, "Erreur", "Impossible d'ouvrir l'explorateur de fichiers.")
        elif sys.platform == "darwin":
            QtCore.QProcess.startDetached("open", [target_path])
        else:
            QtCore.QProcess.startDetached("xdg-open", [target_path])

    def show_in_explorer(self, row, candidates):
        paths = self._resolve_row_paths(row)
        if not paths:
            return
        target_path = paths[0]
        if not os.path.exists(target_path):
            QtWidgets.QMessageBox.warning(self, "Fichier introuvable", "Le fichier sélectionné est introuvable sur le disque.")
            return

        directory = os.path.dirname(target_path) or target_path
        self._open_in_file_manager(directory)

    def delete_mod(self, row, candidates):
        # Prevent deletion if Sims is running
        try:
            if self._is_sims_running():
                QtWidgets.QMessageBox.warning(self, "Suppression impossible", "TS4_x64.exe est en cours d'exécution. Fermez le jeu avant de supprimer des mods.")
                return
        except Exception:
            pass
        paths = self._resolve_row_paths(row)
        if not paths:
            return

        confirm = QtWidgets.QMessageBox.question(
            self,
            "Confirmer la suppression",
            "Supprimer ce mod supprimera définitivement les fichiers associés. Continuer ?",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )

        if confirm != QtWidgets.QMessageBox.Yes:
            return

        errors = []
        for path in paths:
            if not os.path.exists(path):
                continue
            try:
                os.remove(path)
            except OSError as exc:
                errors.append(str(exc))

        if errors:
            QtWidgets.QMessageBox.warning(
                self,
                "Erreur lors de la suppression",
                "\n".join(errors),
            )

        if candidates:
            for name in candidates:
                self.ignored_mods.discard(name)
            save_ignore_list(self.ignored_mods)
            self.settings["ignored_mods"] = sorted(self.ignored_mods)
            save_settings(self.settings)

        if self.last_scanned_directory and os.path.isdir(self.last_scanned_directory):
            self.refresh_table_only()
        else:
            self.table.removeRow(row)

    def launch_google_search(self, row, candidates):
        file_name = ""
        # columns: 2=package, 4=script after adding group column
        for column in (2, 4):
            item = self.table.item(row, column)
            if item:
                text = item.text().strip()
                if text:
                    file_name = text
                    break

        if not file_name and candidates:
            file_name = candidates[0]

        if not file_name:
            return

        base_name, _ = os.path.splitext(file_name)
        if not base_name:
            return

        search_url = QtCore.QUrl(f"https://www.google.com/search?q={quote_plus(base_name)}")
        QtGui.QDesktopServices.openUrl(search_url)

    def launch_patreon_search(self, row, candidates):
        file_name = ""
        for column in (2, 4):
            item = self.table.item(row, column)
            if item:
                text = item.text().strip()
                if text:
                    file_name = text
                    break
        if not file_name and candidates:
            file_name = candidates[0]
        if not file_name:
            return
        base_name, _ = os.path.splitext(file_name)
        if not base_name:
            return
        q = quote_plus(f"site:patreon.com {base_name}")
        QtGui.QDesktopServices.openUrl(QtCore.QUrl(f"https://www.google.com/search?q={q}"))

    def update_ignore_mod(self, candidates, state):
        candidates = [name for name in candidates if name]
        if not candidates:
            return

        canonical_key = candidates[0]
        if state == QtCore.Qt.Checked:
            for name in candidates[1:]:
                self.ignored_mods.discard(name)
            self.ignored_mods.add(canonical_key)
        else:
            for name in candidates:
                self.ignored_mods.discard(name)

        save_ignore_list(self.ignored_mods)
        self.settings["ignored_mods"] = sorted(self.ignored_mods)
        save_settings(self.settings)
        self.refresh_table_only()

    def export_current(self):
        # Block export while Sims is running
        try:
            if self._is_sims_running():
                QtWidgets.QMessageBox.warning(self, "Opération bloquée", "TS4_x64.exe est en cours d'exécution. Fermez le jeu pour exporter.")
                return
        except Exception:
            pass
        # Build headers and data using only visible columns
        header = self.table.horizontalHeader()
        visible_cols = [c for c in range(self.table.columnCount()) if not header.isSectionHidden(c)]
        # Compose headers
        headers = []
        for c in visible_cols:
            item = self.table.horizontalHeaderItem(c)
            headers.append(item.text() if item else f"Col {c}")
        # Compose data rows
        rows = []
        for r in range(self.table.rowCount()):
            out = []
            for c in visible_cols:
                # Last column is an "Ignoré" checkbox in current table design
                if self.table.cellWidget(r, c) is not None and isinstance(self.table.cellWidget(r, c), QtWidgets.QCheckBox):
                    out.append(bool(self.table.cellWidget(r, c).isChecked()))
                else:
                    it = self.table.item(r, c)
                    out.append(it.text() if it else "")
            rows.append(out)

        save_path = self.settings.get("xls_file_path", "")
        if not save_path:
            save_path = QtWidgets.QFileDialog.getSaveFileName(self, "Sauvegarder sous", "", "Excel Files (*.xlsx)")[0]
            if not save_path:
                return
            self.settings["xls_file_path"] = save_path
            save_settings(self.settings)

        export_to_excel(save_path, rows, headers)
        QtWidgets.QMessageBox.information(self, "Info", f"Export terminé vers : {save_path}")

    def _show_header_menu(self, pos):
        header = self.table.horizontalHeader()
        global_pos = header.mapToGlobal(pos)
        menu = QtWidgets.QMenu(self)
        labels = [
            "État",
            "Mod (groupe)",
            "Fichier .package",
            "Date .package",
            "Fichier .ts4script",
            "Date .ts4script",
            "Version",
            "Confiance",
            "Installer",
            "Ignoré",
        ]
        hidden = set(self.settings.get("hidden_columns", []))
        for col, label in enumerate(labels):
            action = QtWidgets.QAction(label, menu)
            action.setCheckable(True)
            action.setChecked(col not in hidden)
            action.triggered.connect(partial(self._toggle_column_visibility, col))
            menu.addAction(action)
        menu.exec_(global_pos)

    def _toggle_column_visibility(self, col, checked):
        hidden = set(int(c) for c in self.settings.get("hidden_columns", []))
        if checked and col in hidden:
            hidden.remove(col)
        elif not checked:
            hidden.add(col)
        self.settings["hidden_columns"] = sorted(int(c) for c in hidden)
        save_settings(self.settings)
        self.table.setColumnHidden(col, not checked)

    def _apply_hidden_columns(self):
        hidden = set(int(c) for c in self.settings.get("hidden_columns", []))
        for col in range(self.table.columnCount()):
            self.table.setColumnHidden(col, col in hidden)


# ---------- AI Training Dialog ----------
"""AI Training dialog moved to modules/ai_training.py"""

class UpdatesCheckerDialog(QtWidgets.QDialog):
    def __init__(self, parent: 'ModManagerApp'):
        super().__init__(parent)
        self.setWindowTitle("Updates Checker")
        self.setModal(True)
        try:
            self.setWindowFlags(self.windowFlags() | QtCore.Qt.Window)
            self.setSizeGripEnabled(True)
        except Exception:
            pass
        self.resize(900, 600)
        self.parent_app = parent
        self._rows = []
        self._index = []
        self._name_to_row = {}
        self._paths_by_name = {}
        self._paths_by_name = {}
        layout = QtWidgets.QVBoxLayout(self)

        # Toolbar
        tools = QtWidgets.QHBoxLayout()
        self.search_edit = QtWidgets.QLineEdit(self)
        self.search_edit.setPlaceholderText("Filtrer par nom de mod…")
        self.search_edit.textChanged.connect(self._apply_filter)
        tools.addWidget(self.search_edit, 1)
        self.load_index_btn = QtWidgets.QPushButton("Load Index", self)
        self.load_index_btn.clicked.connect(self._load_index)
        tools.addWidget(self.load_index_btn)
        self.check_all_btn = QtWidgets.QPushButton("Check All (beta)", self)
        self.check_all_btn.clicked.connect(self._check_all_online)
        tools.addWidget(self.check_all_btn)
        self.check_obsolete_btn = QtWidgets.QPushButton("Check Obsolete", self)
        self.check_obsolete_btn.clicked.connect(self._check_obsolete)
        tools.addWidget(self.check_obsolete_btn)
        layout.addLayout(tools)

        # Table
        self.table = QtWidgets.QTableWidget(0, 7, self)
        self.table.setHorizontalHeaderLabels([
            "#", "Mod", "Local Version", "Search", "Try (beta)", "Status", "URL"
        ])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        layout.addWidget(self.table, 1)

        self._populate_initial()

    def _apply_filter(self, *_):
        term = (self.search_edit.text() or "").strip().lower()
        for r in range(self.table.rowCount()):
            name = self.table.item(r, 1).text().lower() if self.table.item(r, 1) else ""
            self.table.setRowHidden(r, bool(term and term not in name))

    def _populate_initial(self):
        # Utilise le cache pour grouper d'abord par .ts4script puis par .package
        mods = []
        self._name_to_row = {}
        self._paths_by_name = {}
        cache_path = os.path.join(os.getcwd(), 'mod_scan_cache.json')
        mods_root = ''
        try:
            with open(cache_path, 'r', encoding='utf-8') as fh:
                data = json.load(fh)
            mods_root = str(data.get('root') or '')
            entries = list(data.get('entries') or [])
        except Exception:
            entries = []
        if entries:
            from collections import OrderedDict
            pkg_by_key = {}
            scr_by_key = {}
            order_scripts = []
            order_packages = []
            for e in entries:
                rel = str(e.get('path') or '').replace('\\', '/')
                ftype = (e.get('type') or '').lower()
                if not rel or ftype not in {'package', 'ts4script'}:
                    continue
                base = os.path.basename(rel)
                norm = normalize_mod_basename(base)
                abs_path = os.path.join(mods_root, rel) if mods_root else rel
                if ftype == 'ts4script':
                    if norm not in scr_by_key:
                        order_scripts.append(norm)
                    scr_by_key.setdefault(norm, []).append(abs_path)
                else:
                    if norm not in pkg_by_key:
                        order_packages.append(norm)
                    pkg_by_key.setdefault(norm, []).append(abs_path)
            # Build mods list: all script groups first (with matching packages), then package-only groups
            seen_keys = set()
            for key in order_scripts:
                paths = set(scr_by_key.get(key, []))
                paths.update(pkg_by_key.get(key, []))
                if not paths:
                    continue
                # Display name: prefer first ts4script stem
                try:
                    first = os.path.splitext(os.path.basename(scr_by_key[key][0]))[0]
                except Exception:
                    first = key
                name = first.strip() or key
                low = name.lower()
                self._paths_by_name[low] = set(paths)
                self._name_to_row[low] = {'group': name, 'version': '', 'paths': list(paths)}
                mods.append((name, ''))
                seen_keys.add(key)
            for key in order_packages:
                if key in seen_keys:
                    continue
                paths = set(pkg_by_key.get(key, []))
                if not paths:
                    continue
                try:
                    first = os.path.splitext(os.path.basename(pkg_by_key[key][0]))[0]
                except Exception:
                    first = key
                name = first.strip() or key
                low = name.lower()
                self._paths_by_name[low] = set(paths)
                self._name_to_row[low] = {'group': name, 'version': '', 'paths': list(paths)}
                mods.append((name, ''))
        else:
            # Fallback: utilise les lignes de la vue actuelle
            rows = list(getattr(self.parent_app, 'all_data_rows', []) or [])
            mods_root = str(getattr(self.parent_app, 'settings', {}).get('mod_directory', '') or '')
            mods_root_norm = os.path.abspath(mods_root) if mods_root else ''
            for row in rows:
                # Respecte le groupe tel que construit par la vue principale
                name = str(row.get('group') or '').strip()
                if not name:
                    pkg = (row.get('package') or '').strip()
                    scr = (row.get('script') or '').strip()
                    base = pkg or scr
                    if base:
                        name = os.path.splitext(os.path.basename(base))[0]
                if not name:
                    continue
                key = name.lower()
                self._name_to_row[key] = row
                for p in list(row.get('paths') or []):
                    self._paths_by_name.setdefault(key, set()).add(p)
                mods.append((name, str(row.get('version') or '').strip()))
        # Unique by name preserving order
        seen = set()
        uniq = []
        for name, ver in mods:
            if name.lower() in seen:
                continue
            seen.add(name.lower())
            uniq.append((name, ver))
        self.table.setRowCount(0)
        for i, (name, ver) in enumerate(uniq, start=1):
            r = self.table.rowCount()
            self.table.insertRow(r)
            self.table.setItem(r, 0, QtWidgets.QTableWidgetItem(str(i)))
            self.table.setItem(r, 1, QtWidgets.QTableWidgetItem(name))
            self.table.setItem(r, 2, QtWidgets.QTableWidgetItem(ver))
            # Search button
            btn_search = QtWidgets.QPushButton("Open", self)
            btn_search.clicked.connect(partial(self._open_search, name))
            self.table.setCellWidget(r, 3, btn_search)
            # Try button
            btn_try = QtWidgets.QPushButton("Try", self)
            btn_try.clicked.connect(partial(self._try_fetch, name, r))
            self.table.setCellWidget(r, 4, btn_try)
            self.table.setItem(r, 5, QtWidgets.QTableWidgetItem(""))
            self.table.setItem(r, 6, QtWidgets.QTableWidgetItem(""))

    def _open_search(self, name: str):
        q = quote_plus(f"site:app.ts4modhound.com {name}")
        url = f"https://www.google.com/search?q={q}"
        try:
            QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))
        except Exception:
            pass

    def _set_status(self, row: int, text: str):
        self.table.setItem(row, 5, QtWidgets.QTableWidgetItem(text))

    def _try_fetch(self, name: str, row: int):
        if not self._index:
            self._load_index()
        if not self._index:
            self._set_status(row, "Index unavailable")
            return
        best = self._best_match(name)
        if not best:
            self._set_status(row, "No match")
            return
        remote_date = best.get('date') or ''
        remote_url = best.get('url') or ''
        local_dt = self._local_latest_datetime(name)
        local_s = local_dt.strftime('%Y-%m-%d') if local_dt else '?'
        remote_dt = self._parse_date(remote_date)
        remote_s = remote_dt.strftime('%Y-%m-%d') if remote_dt else (remote_date or '?')
        if remote_dt and local_dt:
            status = 'Update available' if remote_dt > local_dt else 'Up-to-date'
        else:
            status = 'Found'
        info = f"{status} – Remote {remote_s} – Local {local_s}"
        self._set_status(row, info)
        if remote_url:
            btn = self.table.cellWidget(row, 3)
            if isinstance(btn, QtWidgets.QPushButton):
                try:
                    btn.clicked.disconnect()
                except Exception:
                    pass
                # Assure que l'URL est une chaîne HTTP/HTTPS valide avant de connecter
                try:
                    url_str = str(remote_url).strip()
                except Exception:
                    url_str = ""
                if url_str and (url_str.startswith("http://") or url_str.startswith("https://")):
                    qurl = QtCore.QUrl(url_str)
                    btn.clicked.connect(lambda u=qurl: QtGui.QDesktopServices.openUrl(u))
            # Show URL in dedicated column
            try:
                self.table.setItem(row, 6, QtWidgets.QTableWidgetItem(str(remote_url)))
            except Exception:
                pass

    def _check_all_online(self):
        if not self._index:
            self._load_index()
        for r in range(self.table.rowCount()):
            if self.table.isRowHidden(r):
                continue
            name = self.table.item(r, 1).text() if self.table.item(r, 1) else ""
            if name:
                self._try_fetch(name, r)
                QtWidgets.QApplication.processEvents()

    # indexing & matching
    def _load_index(self):
        # Default: Google Sheets CSV (Scarlet's Realm Mod List Checker)
        default_csv = (
            "https://docs.google.com/spreadsheets/d/e/"
            "2PACX-1vRexBc8fcYyfsjbGRo3sH18jj9DuwKH8J7_SvQvpK_fvjsnILKRz1xGOwYz-xtG0wIKQcs1eDN1yw9V/"
            "pub?gid=119778444&single=true&range=A:I&output=csv"
        )
        csv_url = str(getattr(self.parent_app, 'settings', {}).get('updates_checker_csv_url', default_csv))
        entries = []
        try:
            req = urllib.request.Request(csv_url, headers={"User-Agent": "Sims4ModTool/3.40"})
            with urllib.request.urlopen(req, timeout=20) as resp:
                text = resp.read().decode('utf-8', errors='ignore')
            import io, csv as _csv
            buf = io.StringIO(text)
            reader = _csv.reader(buf)
            for row in reader:
                if not row or len(row) < 4:
                    continue
                # Try to detect/skip a header row (first cell not numeric)
                if row[0] and not str(row[0]).strip().isdigit() and len(row) > 1 and str(row[1]).lower().startswith('name'):
                    continue
                name = (row[1] or '').strip()
                if not name:
                    continue
                creator = (row[2] or '').strip() if len(row) > 2 else ''
                link = (row[3] or '').strip() if len(row) > 3 else ''
                status = (row[4] or '').strip() if len(row) > 4 else ''
                date_pretty = (row[5] or '').strip() if len(row) > 5 else ''
                date_iso = (row[6] or '').strip() if len(row) > 6 else ''
                date_str = date_iso or date_pretty
                entries.append({
                    'title': name,
                    'url': link,
                    'date': date_str,
                    'status': status,
                    'creator': creator,
                    'tokens': self._tokenize(name),
                })
            self._index = entries
            QtWidgets.QMessageBox.information(self, "Index", f"Index chargé (Google CSV): {len(entries)} éléments")
            return
        except Exception as exc:
            # Fallback to TS4ModHound scraping
            try:
                url = "https://app.ts4modhound.com/visitor/all_creators_content"
                req = urllib.request.Request(url, headers={"User-Agent": "Sims4ModTool/3.40"})
                with urllib.request.urlopen(req, timeout=15) as resp:
                    html = resp.read().decode('utf-8', errors='ignore')
                try:
                    if BeautifulSoup is not None:
                        soup = BeautifulSoup(html, 'html.parser')
                        for a in soup.find_all('a'):
                            title = (a.get_text(strip=True) or '')
                            href = a.get('href') or ''
                            if not title or not href:
                                continue
                            if 'visitor' not in href and 'app.ts4modhound.com' not in href:
                                continue
                            url_abs = href if href.startswith('http') else ('https://app.ts4modhound.com' + href)
                            date_str = ''
                            parent_text = a.parent.get_text(' ', strip=True) if a.parent else ''
                            m = re.search(r"(\d{4}[-/\.](?:\d{1,2})[-/\.](?:\d{1,2}))", parent_text)
                            if m:
                                date_str = m.group(1)
                            entries.append({'title': title, 'url': url_abs, 'date': date_str, 'tokens': self._tokenize(title)})
                    else:
                        for m in re.finditer(r"<a[^>]+href=\"([^\"]+)\"[^>]*>(.*?)</a>", html, re.I|re.S):
                            href, txt = m.group(1), re.sub('<[^>]+>','',m.group(2))
                            title = (txt or '').strip()
                            if not title or not href:
                                continue
                            url_abs = href if href.startswith('http') else ('https://app.ts4modhound.com' + href)
                            entries.append({'title': title, 'url': url_abs, 'date': '', 'tokens': self._tokenize(title)})
                except Exception:
                    entries = []
                self._index = entries
                QtWidgets.QMessageBox.information(self, "Index", f"Index (fallback) chargé: {len(entries)} éléments\nErreur CSV: {exc}")
                return
            except Exception as exc2:
                self._index = []
                QtWidgets.QMessageBox.warning(self, "Index", f"Impossible de charger l'index (CSV et fallback): {exc2}")
                return

    def _check_obsolete(self):
        # Liste publiée par l'utilisateur (Google Sheet en HTML)
        url = (
            'https://docs.google.com/spreadsheets/u/0/d/e/2PACX-1vTpfOxPBlU44aaCGXESyAqg4t7Vkl1xucXbNtSQAYfKDIwNNeQFynfVLUtDDp4AiEXaIe7okyVYJMnv/pubhtml/'
            'sheet?headers=false&gid=194232938'
        )
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Sims4ModTool/3.40"})
            with urllib.request.urlopen(req, timeout=20) as resp:
                html = resp.read().decode('utf-8', errors='ignore')
        except Exception as exc:
            QtWidgets.QMessageBox.warning(self, 'Obsolete', f'Impossible de charger la liste obsolete: {exc}')
            return
        names = set()
        try:
            if BeautifulSoup is not None:
                soup = BeautifulSoup(html, 'html.parser')
                # récupère cellules de texte significatives
                for td in soup.find_all('td'):
                    title = (td.get_text(strip=True) or '')
                    if title and len(title) > 2:
                        names.add(title.lower())
            else:
                for m in re.finditer(r">([^<]{3,})<", html):
                    names.add(m.group(1).strip().lower())
        except Exception:
            pass
        if not names:
            QtWidgets.QMessageBox.information(self, 'Obsolete', 'Aucun nom détecté dans la liste.')
            return
        matched = 0
        for r in range(self.table.rowCount()):
            if self.table.isRowHidden(r):
                continue
            name = (self.table.item(r, 1).text() or '').strip().lower() if self.table.item(r, 1) else ''
            if not name:
                continue
            if name in names:
                matched += 1
                prev = self.table.item(r, 5).text() if self.table.item(r, 5) else ''
                self._set_status(r, f"OBSOLETE • {prev}" if prev else 'OBSOLETE')
                try:
                    for c in range(self.table.columnCount()):
                        it = self.table.item(r, c)
                        if it is None:
                            it = QtWidgets.QTableWidgetItem('')
                            self.table.setItem(r, c, it)
                        it.setBackground(QtGui.QBrush(QtGui.QColor('#5b1f1f')))
                        it.setForeground(QtGui.QBrush(QtGui.QColor('#ffffff')))
                except Exception:
                    pass
        QtWidgets.QMessageBox.information(self, 'Obsolete', f'{matched} mod(s) obsolète(s) détecté(s).')

    @staticmethod
    def _tokenize(text: str):
        return [t for t in re.split(r"[^a-z0-9]+", (text or '').lower()) if t and len(t) > 1]

    def _best_match(self, name: str):
        tokens = set(self._tokenize(name))
        best = None
        best_score = 0.0
        for e in self._index:
            overlap = len(tokens.intersection(set(e.get('tokens') or [])))
            ratio = SequenceMatcher(None, name.lower(), (e.get('title') or '').lower()).ratio()
            score = overlap * 2 + ratio
            if score > best_score:
                best_score = score
                best = e
        return best

    def _local_latest_datetime(self, name: str):
        key = (name or '').lower()
        paths = list(self._paths_by_name.get(key) or [])
        if not paths:
            row = self._name_to_row.get(key)
            if not row:
                return None
            paths = list(row.get('paths') or [])
        latest = None
        for p in paths:
            try:
                ts = os.path.getmtime(p)
                dt = datetime.fromtimestamp(ts)
                latest = dt if not latest or dt > latest else latest
            except Exception:
                continue
        return latest

    def _parse_date(self, s: str):
        if not s:
            return None
        s = s.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%Y.%m.%d", "%b %d, %Y", "%B %d, %Y", "%b %d %Y", "%B %d %Y"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                continue
        m = re.search(r"(\d{4})[-/\.](\d{1,2})[-/\.]?(\d{1,2})?", s)
        if m:
            y = int(m.group(1)); mth = int(m.group(2)); d = int(m.group(3) or 1)
            try:
                return datetime(y, mth, d)
            except Exception:
                return None
        return None



# Log Manager UI moved to modules/log_manager.py

# (Removed duplicate AITrainingDialog definition — the enhanced version above is the single source of truth.)


if __name__ == "__main__":
    # Ensure high-DPI scaling behaves consistently before QApplication is created
    try:
        QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling)  # type: ignore[attr-defined]
    except Exception:
        pass

    app = QtWidgets.QApplication(sys.argv)

    class StartupSplash(QtWidgets.QSplashScreen):
        def __init__(self, title: str, bg_image_path: Optional[str] = None):
            pix = QtGui.QPixmap(640, 360)
            pix.fill(QtGui.QColor("#263238"))
            super().__init__(pix)
            self.setWindowFlag(QtCore.Qt.WindowStaysOnTopHint)
            self._title = title
            self._message = ""
            try:
                path = str(bg_image_path or "")
                path = os.path.expanduser(os.path.expandvars(path))
            except Exception:
                path = str(bg_image_path or "")
            self._bg_image_path = path
            self._draw()

        def _draw(self):
            pix = self.pixmap()
            painter = QtGui.QPainter(pix)
            try:
                painter.setRenderHint(QtGui.QPainter.Antialiasing)
                rect = pix.rect()
                has_image = False
                bg_path = self._bg_image_path
                if bg_path:
                    try:
                        bg_path = os.path.abspath(bg_path)
                    except Exception:
                        pass
                if bg_path and os.path.isfile(bg_path):
                    try:
                        img = QtGui.QPixmap(bg_path)
                        if not img.isNull():
                            scaled = img.scaled(rect.size(), QtCore.Qt.KeepAspectRatioByExpanding, QtCore.Qt.SmoothTransformation)
                            painter.drawPixmap(rect, scaled)
                            has_image = True
                    except Exception:
                        has_image = False
                if not has_image:
                    gradient = QtGui.QLinearGradient(rect.topLeft(), rect.bottomRight())
                    gradient.setColorAt(0.0, QtGui.QColor("#1b5e20"))
                    gradient.setColorAt(0.4, QtGui.QColor("#004d40"))
                    gradient.setColorAt(1.0, QtGui.QColor("#263238"))
                    painter.fillRect(rect, QtGui.QBrush(gradient))
                else:
                    overlay = QtGui.QLinearGradient(rect.topLeft(), rect.bottomRight())
                    overlay.setColorAt(0.0, QtGui.QColor(27, 94, 32, 140))
                    overlay.setColorAt(0.4, QtGui.QColor(0, 77, 64, 120))
                    overlay.setColorAt(1.0, QtGui.QColor(38, 50, 56, 140))
                    painter.fillRect(rect, QtGui.QBrush(overlay))

                vignette = QtGui.QRadialGradient(rect.center(), max(rect.width(), rect.height()) * 0.75)
                vignette.setColorAt(0.0, QtGui.QColor(0, 0, 0, 0))
                vignette.setColorAt(1.0, QtGui.QColor(0, 0, 0, 110))
                painter.fillRect(rect, QtGui.QBrush(vignette))

                cx = rect.center().x()
                cy = int(rect.height() * 0.44)
                width = max(100, int(rect.width() * 0.22))
                height = max(140, int(rect.height() * 0.42))
                diamond = QtGui.QPainterPath()
                diamond.moveTo(cx, cy - height // 2)
                diamond.lineTo(cx + width // 2, cy)
                diamond.lineTo(cx, cy + height // 2)
                diamond.lineTo(cx - width // 2, cy)
                diamond.closeSubpath()

                fill = QtGui.QLinearGradient(cx, cy - height // 2, cx, cy + height // 2)
                fill.setColorAt(0.0, QtGui.QColor("#00e676"))
                fill.setColorAt(0.5, QtGui.QColor("#1de9b6"))
                fill.setColorAt(1.0, QtGui.QColor("#00c853"))
                painter.setPen(QtGui.QPen(QtGui.QColor("#004d40"), 2))
                painter.setBrush(QtGui.QBrush(fill))
                painter.drawPath(diamond)

                glow = QtGui.QRadialGradient(QtCore.QPointF(cx, cy), height * 0.9)
                glow.setColorAt(0.0, QtGui.QColor(0, 255, 170, 70))
                glow.setColorAt(1.0, QtGui.QColor(0, 0, 0, 0))
                painter.setBrush(QtGui.QBrush(glow))
                painter.setPen(QtCore.Qt.NoPen)
                painter.drawEllipse(QtCore.QRectF(cx - width, cy - width, width * 2, width * 2))

                title_font = QtGui.QFont()
                title_font.setPointSize(18)
                title_font.setBold(True)
                painter.setFont(title_font)
                painter.setPen(QtGui.QColor("#e0f2f1"))
                title_rect = QtCore.QRect(rect.left() + 24, rect.top() + 20, rect.width() - 48, max(40, int(cy - height // 2) - 20))
                painter.drawText(title_rect, QtCore.Qt.AlignHCenter | QtCore.Qt.AlignBottom | QtCore.Qt.TextWordWrap, self._title)

                message_font = QtGui.QFont()
                message_font.setPointSize(11)
                painter.setFont(message_font)
                painter.setPen(QtGui.QColor("#c8e6c9"))
                msg_top = min(rect.bottom() - 60, int(cy + height // 2) + 20)
                message_rect = QtCore.QRect(rect.left() + 24, msg_top, rect.width() - 48, rect.bottom() - msg_top - 24)
                painter.drawText(message_rect, QtCore.Qt.AlignHCenter | QtCore.Qt.AlignTop | QtCore.Qt.TextWordWrap, self._message)
            finally:
                painter.end()
            self.setPixmap(pix)

        def update_message(self, text: str):
            self._message = str(text or "")
            self._draw()

    try:
        initial_settings = load_settings()
        splash_background = initial_settings.get("splash_background_image_path", "")
    except Exception:
        splash_background = ""

    splash = StartupSplash(f"Sims 4 Mod Manager {APP_VERSION}\n{APP_VERSION_DATE}", splash_background)
    splash.show()
    app.processEvents()

    window = ModManagerApp(splash=splash)
    window.show()

    QtCore.QTimer.singleShot(2000, lambda: splash.finish(window) if splash is not None else None)
    sys.exit(app.exec_())
